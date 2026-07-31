from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.modules.tax_invoice.recognition import (
    combine_invoice_and_customs,
    lookup_fx_rate,
    parse_bot_fx_workbook,
    parse_invoice_workbook,
    parse_sample_workbook,
    recompute_line_thb,
)


def test_recompute_line_thb_none_propagation_and_half_up() -> None:
    """THB 折算的唯一口径：None 传播 + ROUND_HALF_UP 到分。

    识别建单与事后汇率重匹配都走 recompute_line_thb，把这条钉死，防止哪天
    有人在重匹配那侧改回银行家舍入——那会让台账和 PDF 差一分且静默无声。
    """
    # 缺 USD 金额或缺汇率时不臆造 THB。
    assert recompute_line_thb(None, Decimal("32.1230")) is None
    assert recompute_line_thb(Decimal("100"), None) is None
    # 常规折算，保留两位小数。
    assert recompute_line_thb(Decimal("100.00"), Decimal("32.1230")) == Decimal("3212.30")
    # 半分一律进位（ROUND_HALF_UP），而不是 Python 默认的银行家舍入——
    # 后者会把 0.125 收成 0.12。
    assert recompute_line_thb(Decimal("0.125"), Decimal("1")) == Decimal("0.13")


def _invoice_workbook(unit_price: object = 30, amount_usd: object = 300) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CI INVOICE"
    sheet["A1"] = "DATE：2026-06-05"
    sheet["A2"] = "INV.NO.：ZWT-TEST-001"
    sheet["A3"] = "INCOTERMS：FCA"
    sheet["A4"] = "PAYMENT TERMS：OA 30 DAYS"
    sheet["A5"] = "PO NO.：PO-001"
    sheet["A6"] = "BUYER："
    sheet["B6"] = "TEST CUSTOMER"
    sheet["B7"] = "BANGKOK"
    headers = [
        "ITEM NO.",
        "PRODUCT NAME",
        "PRODUCT CODE",
        "H.S. CODE",
        "QUANTITY",
        "UNIT",
        "UNIT PRICE",
        "AMOUNT USD",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=10, column=column, value=header)
    values = [1, "ROUTER", "R-1", "85176243", 10, "PIECES", unit_price, amount_usd]
    for column, value in enumerate(values, start=1):
        sheet.cell(row=11, column=column, value=value)
    sheet["A12"] = "TOTAL"
    sheet["E12"] = 10
    sheet["H12"] = amount_usd
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _bot_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["C6"] = "05 JUN 2026"
    sheet["D6"] = "08 JUN 2026"
    sheet["B7"] = "REFERENCE RATE USD"
    sheet["C10"] = 32.1234
    sheet["D10"] = 32.2222
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_invoice_parser_preserves_existing_fields_and_items() -> None:
    parsed = parse_invoice_workbook(_invoice_workbook(), "invoice.xlsx")

    assert parsed["ci_no"] == "ZWT-TEST-001"
    assert parsed["ci_date"] == date(2026, 6, 5)
    assert parsed["customer_name"] == "TEST CUSTOMER"
    assert parsed["quantity_total"] == Decimal("10")
    assert parsed["items"][0]["hs_code"] == "85176243"


def test_invoice_parser_degrades_out_of_range_amounts_instead_of_raising() -> None:
    # 超出 Decimal 上下文精度的单元格（1e100）会让 quantize 抛 InvalidOperation。
    # 识别阶段必须降级为 None，由 _review_status 标记 needs_review 交人工确认；
    # 上传接口不能因为一个畸形单元格返回 500。
    content = _invoice_workbook(unit_price=1e100, amount_usd=1e100)

    parsed = parse_invoice_workbook(content, "invoice.xlsx")

    assert parsed["items"][0]["ci_unit_price"] is None
    assert parsed["fob_amount_usd"] is None


def test_bot_parser_and_weekend_lookback() -> None:
    rates = parse_bot_fx_workbook(_bot_workbook(), "bot.xlsx")

    rate, matched_date = lookup_fx_rate(rates, date(2026, 6, 7))

    assert rate == Decimal("32.1234")
    assert matched_date == date(2026, 6, 5)


def test_combination_separates_invoice_and_matched_exchange_dates() -> None:
    invoice = parse_invoice_workbook(_invoice_workbook(), "invoice.xlsx")
    invoice["fob_amount_usd"] = Decimal("309.00")
    customs = {
        "cdn": "A001234567890",
        "submission_date": date(2026, 6, 7),
        "customs_fob_usd_total": Decimal("309.00"),
        "submission_date_low_confidence": False,
    }
    combined = combine_invoice_and_customs(
        invoice,
        customs,
        {date(2026, 6, 5): Decimal("32.1234")},
    )

    assert combined["invoice_date"] == date(2026, 6, 7)
    assert combined["exchange_target_date"] == date(2026, 6, 7)
    assert combined["exchange_rate_date"] == date(2026, 6, 5)
    assert combined["fob_revenue_usd_total"] == Decimal("309.00")
    assert combined["fob_verification_failed"] is False


def test_combination_attaches_customs_line_fob_usd_by_position() -> None:
    """报关单该行自印的 FOB USD 按行位置挂到发票行上落库，供复核台逐行核对。

    汇总对得上、分行对不上（合计相同但某行金额不同）在总额上完全看不出来，
    所以行级值必须落库、能在复核台并排看到，不能只留一个 fob_verification_failed。
    """
    invoice = parse_invoice_workbook(_invoice_workbook(), "invoice.xlsx")
    invoice["fob_amount_usd"] = Decimal("309.00")
    customs = {
        "cdn": "A001234567890",
        "submission_date": date(2026, 6, 7),
        "customs_fob_usd_total": Decimal("309.00"),
        "submission_date_low_confidence": False,
        # 报关单该行金额与发票行（309.00）不同，仍如实挂上——复核台标红交人看。
        "customs_items": [{"line_number": 1, "fob_usd": Decimal("300.00")}],
    }

    combined = combine_invoice_and_customs(
        invoice,
        customs,
        {date(2026, 6, 5): Decimal("32.1234")},
    )

    assert combined["items"][0]["customs_fob_usd"] == Decimal("300.00")


def test_combination_leaves_customs_line_none_without_customs_items() -> None:
    """报关单读不到明细时，行级 customs_fob_usd 留空——「没得核对」而不是「核对不过」。"""
    invoice = parse_invoice_workbook(_invoice_workbook(), "invoice.xlsx")
    invoice["fob_amount_usd"] = Decimal("309.00")
    customs = {
        "cdn": "A001234567890",
        "submission_date": date(2026, 6, 7),
        "customs_fob_usd_total": Decimal("309.00"),
        "submission_date_low_confidence": False,
    }

    combined = combine_invoice_and_customs(
        invoice,
        customs,
        {date(2026, 6, 5): Decimal("32.1234")},
    )

    assert combined["items"][0]["customs_fob_usd"] is None


def test_sample_import_marks_legacy_fx_date_fallback_for_review() -> None:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "C/I No.",
        "CDN",
        "Customer Name",
        "Customer Address",
        "Product Name or Service",
        "Product Code",
        "Unit",
        "Quantity",
        "FX Date",
        "FX Rate",
        "FOB Rev USD",
        "FOB Rev THB",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "ZWT-TEST-001",
            "A001234567890",
            "TEST CUSTOMER",
            "BANGKOK",
            "ROUTER",
            "R-1",
            "PIECES",
            10,
            "2026-06-07",
            32.1234,
            309,
            9926.13,
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    imported = parse_sample_workbook(output.getvalue(), "sample.xlsx")

    assert imported[0]["invoice_date"] == date(2026, 6, 7)
    assert imported[0]["submission_date_low_confidence"] is True
    assert imported[0]["submission_date_source"] == "legacy_fx_date_fallback"
