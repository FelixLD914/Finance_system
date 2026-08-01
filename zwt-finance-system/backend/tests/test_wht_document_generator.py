from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pdfplumber
import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
from pypdf import PdfReader

from app.modules.wht.document_generator import (
    DocumentGenerationError,
    build_blue_signature,
    export_pdf_from_template,
    render_wht_workbook,
    thai_baht_text,
    thai_date_text,
    validate_signature_image,
)
from app.modules.wht.models import WhtTask


@pytest.mark.parametrize(
    ("amount", "expected"),
    [
        ("0", "ศูนย์บาทถ้วน"),
        ("1", "หนึ่งบาทถ้วน"),
        ("11", "สิบเอ็ดบาทถ้วน"),
        ("21.25", "ยี่สิบเอ็ดบาทยี่สิบห้าสตางค์"),
        ("1000001", "หนึ่งล้านเอ็ดบาทถ้วน"),
    ],
)
def test_thai_baht_text(amount: str, expected: str) -> None:
    assert thai_baht_text(Decimal(amount)) == expected


def test_thai_date_uses_buddhist_year() -> None:
    assert thai_date_text(date(2026, 6, 10)) == "10 มิถุนายน 2569"


def test_renders_old_template_placeholders(tmp_path) -> None:
    template_path = tmp_path / "Template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["RefNo", "BookNo", "PayeeNameTH", "PaymentDate"])
    worksheet.append(["Amount", "TaxRate", "TaxAmount", "AmountTextThai"])
    worksheet.append(["WHT Type3", "WHT Type53", "DateTextThai", "IncomeType"])
    workbook.save(template_path)
    workbook.close()

    task = WhtTask(
        task_no="ZWT202606001",
        book_no="202606",
        period="2026-06",
        issuance_type="normal",
        supplement_run=0,
        status="approved",
        company_name="บริษัท ตัวอย่าง จำกัด",
        payee_address="กรุงเทพมหานคร",
        tax_id="0105540057561",
        wht_type="PND53",
        branch_type="head_office",
        branch_number=None,
        income_type="ค่าบริการ",
        payment_date=date(2026, 6, 10),
        wht_rate=Decimal("0.03"),
        total_amount=Decimal("3000.00"),
        wht_amount=Decimal("90.00"),
    )

    rendered = load_workbook(BytesIO(render_wht_workbook(template_path, task)))
    result = rendered.active
    assert result["A1"].value == "ZWT202606001"
    assert result["B1"].value == "202606"
    assert result["C1"].value == "บริษัท ตัวอย่าง จำกัด(สำนักงานใหญ่)"
    assert result["A2"].value == Decimal("3000.00")
    assert result["B2"].number_format == "0%"
    assert result["A3"].value is None
    assert result["B3"].value == "√"
    assert result["C3"].value == "10 มิถุนายน 2569"
    rendered.close()


def test_validates_png_signature() -> None:
    output = BytesIO()
    Image.new("RGBA", (20, 10), (255, 255, 255, 0)).save(output, format="PNG")

    assert validate_signature_image(output.getvalue()) == ("image/png", ".png")


def test_rejects_non_image_signature() -> None:
    with pytest.raises(DocumentGenerationError, match="valid image"):
        validate_signature_image(b"not-an-image")


def test_signature_processing_removes_long_horizontal_rule(tmp_path: Path) -> None:
    source = tmp_path / "signature-with-rule.png"
    output = tmp_path / "clean-signature.png"
    image = Image.new("RGB", (300, 120), "white")
    draw = ImageDraw.Draw(image)
    draw.line([(35, 65), (75, 25), (115, 70), (165, 30), (220, 60)], fill="black", width=5)
    draw.line([(20, 95), (280, 95)], fill="black", width=3)
    image.save(source)

    build_blue_signature(source, output)

    with Image.open(output) as cleaned:
        alpha = cleaned.convert("RGBA").getchannel("A")

        def alpha_values(y: int):  # noqa: ANN202
            row = alpha.crop((0, y, alpha.width, y + 1))
            flattened_data = getattr(row, "get_flattened_data", None)
            return flattened_data() if flattened_data is not None else row.getdata()

        row_counts = [
            sum(1 for value in alpha_values(y) if value)
            for y in range(alpha.height)
        ]
        assert alpha.getbbox() is not None
        assert max(row_counts) < alpha.width * 0.45


@pytest.mark.parametrize(
    ("wht_type", "check_x_range"),
    [("PND3", (180, 200)), ("PND53", (360, 376))],
)
def test_generates_four_copy_pdf_without_office(
    tmp_path: Path,
    wht_type: str,
    check_x_range: tuple[int, int],
) -> None:
    assets = Path(__file__).parents[1] / "app" / "assets"
    task = WhtTask(
        task_no="ZWT202606001",
        book_no="202606",
        period="2026-06",
        issuance_type="normal",
        supplement_run=0,
        status="approved",
        company_name="บริษัท ตัวอย่าง จำกัด",
        payee_address="กรุงเทพมหานคร",
        tax_id="0105540057561",
        wht_type=wht_type,
        income_type="ค่าบริการ",
        payment_date=date(2026, 6, 10),
        wht_rate=Decimal("0.03"),
        total_amount=Decimal("3000.00"),
        wht_amount=Decimal("90.00"),
    )
    output = tmp_path / "wht.pdf"

    export_pdf_from_template(
        assets / "templates" / "WHT-Template.pdf",
        output,
        task,
        assets / "fonts" / "Sarabun-Regular.ttf",
        None,
    )

    reader = PdfReader(output)
    assert len(reader.pages) == 4
    assert "ZWT202606001" in (reader.pages[0].extract_text() or "")
    with pdfplumber.open(output) as rendered:
        page = rendered.pages[0]
        words = page.extract_words()
        reference = next(word for word in words if word["text"] == "ZWT202606001")
        assert 510 <= reference["x0"] < reference["x1"] <= 571

        payee_tax_id = next(word for word in words if word["text"] == "0105540057561")
        assert payee_tax_id["bottom"] - payee_tax_id["top"] >= 10

        sequence = next(
            word
            for word in words
            if word["text"] == "1" and 50 <= word["x0"] <= 114 and 229 <= word["top"] <= 245
        )
        assert sequence["bottom"] <= 245

        totals = [
            word
            for word in words
            if word["text"] == "3,000.00" and 603 <= word["top"] <= 619
        ]
        assert len(totals) == 1
        assert totals[0]["bottom"] <= 619

        check_lines = [
            line
            for line in page.lines
            if check_x_range[0] <= min(line["x0"], line["x1"])
            and max(line["x0"], line["x1"]) <= check_x_range[1]
            and 573 <= min(line["y0"], line["y1"]) <= 590
        ]
        assert len(check_lines) == 2


def _image_count(path: Path) -> int:
    reader = PdfReader(path)
    total = 0
    for page in reader.pages:
        resources = page.get("/Resources") or {}
        xobjects = resources.get("/XObject") or {}
        total += sum(
            1
            for reference in xobjects.values()
            if reference.get_object().get("/Subtype") == "/Image"
        )
    return total


@pytest.mark.parametrize("source_file_name", [None, "BatchIssue.xlsx"])
def test_signature_is_rendered_for_single_and_batch_tasks(
    tmp_path: Path, source_file_name: str | None
) -> None:
    """单张与批量最终都是任务快照；来源不同不能让签名链路分叉。"""
    assets = Path(__file__).parents[1] / "app" / "assets"
    task = WhtTask(
        task_no="ZWT202606001",
        book_no="202606",
        period="2026-06",
        issuance_type="normal",
        supplement_run=0,
        status="approved",
        company_name="บริษัท ตัวอย่าง จำกัด",
        payee_address="กรุงเทพมหานคร",
        tax_id="0105540057561",
        wht_type="PND53",
        branch_type="branch",
        branch_number="00001",
        income_type="ค่าบริการ",
        payment_date=date(2026, 6, 10),
        wht_rate=Decimal("0.03"),
        total_amount=Decimal("3000.00"),
        wht_amount=Decimal("90.00"),
        source_file_name=source_file_name,
    )
    signature = tmp_path / "signature.png"
    image = Image.new("RGBA", (240, 80), (255, 255, 255, 0))
    for x in range(20, 220):
        image.putpixel((x, 20 + (x % 35)), (18, 45, 145, 255))
    image.save(signature)
    unsigned = tmp_path / f"unsigned-{source_file_name or 'single'}.pdf"
    signed = tmp_path / f"signed-{source_file_name or 'single'}.pdf"

    export_pdf_from_template(
        assets / "templates" / "WHT-Template.pdf",
        unsigned,
        task,
        assets / "fonts" / "Sarabun-Regular.ttf",
        None,
    )
    export_pdf_from_template(
        assets / "templates" / "WHT-Template.pdf",
        signed,
        task,
        assets / "fonts" / "Sarabun-Regular.ttf",
        signature,
    )

    assert _image_count(signed) > _image_count(unsigned)
