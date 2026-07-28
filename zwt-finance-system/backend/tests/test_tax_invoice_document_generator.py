from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader

from app.modules.tax_invoice import pdf_layout
from app.modules.tax_invoice.document_generator import (
    COPY_COLUMN_OFFSETS,
    DATA_COLUMNS,
    DATA_START_ROW,
    ITEM_NUMBER_FORMATS,
    MAX_ITEMS,
    TaxInvoiceDocumentGenerationError,
    export_pdf_from_template,
    format_amount,
    format_quantity,
    format_rate,
    pdf_field_values,
    render_clean_template_workbook,
    render_tax_invoice_workbook,
    tax_invoice_file_stem,
)
from app.modules.tax_invoice.models import TaxInvoice, TaxInvoiceItem

ASSETS = Path(__file__).parents[1] / "app" / "assets"


def _invoice() -> TaxInvoice:
    return TaxInvoice(
        document_no="ZWT-IV20260608-01",
        status="approved",
        ci_no="CI-001",
        cdn="CDN-001",
        invoice_date=date(2026, 6, 8),
        exchange_target_date=date(2026, 6, 8),
        exchange_rate_date=date(2026, 6, 5),
        currency="USD",
        exchange_rate=Decimal("32.4567"),
        customer_name="บริษัท ทดสอบ จำกัด",
        customer_address="กรุงเทพมหานคร",
        tax_id="0105558004821",
        po_no="PO-001",
        payment_term="30 DAYS",
        fob_revenue_usd_total=Decimal("100.00"),
        fob_revenue_thb_total=Decimal("3245.67"),
        created_by_name="Tester",
        updated_by_name="Tester",
    )


def _item(line_number: int = 1) -> TaxInvoiceItem:
    return TaxInvoiceItem(
        line_number=line_number,
        product_name="Product",
        product_code="P-001",
        unit="PCS",
        quantity=Decimal("2"),
        fob_unit_price_usd=Decimal("50"),
        fob_revenue_usd=Decimal("100"),
        fob_revenue_thb=Decimal("3245.67"),
    )


def _template(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoice"
    for coordinate in ("D13", "D14", "D16", "D17", "D19", "Q13", "Q14"):
        worksheet[coordinate] = coordinate
    path = tmp_path / "template.xlsx"
    workbook.save(path)
    return path


def test_workbook_uses_customs_submission_date_for_invoice_date(tmp_path):
    content = render_tax_invoice_workbook(_template(tmp_path), _invoice(), [_item()])
    worksheet = load_workbook(BytesIO(content), data_only=False).active

    assert worksheet["Q13"].value == "ZWT-IV20260608-01"
    assert worksheet["Q14"].value == "08/06/2026"
    assert worksheet["O23"].value == "08/06/2026"
    assert Decimal(str(worksheet["P23"].value)) == Decimal("32.4567")
    assert worksheet["B23"].value == 1
    assert Decimal(str(worksheet["R23"].value)) == Decimal("3245.67")
    assert worksheet["B24"].value in (None, "")


def _template_with_broken_copy_formats(tmp_path):
    """复刻模板原有的缺陷：两联 COPY 沿用会计格式，数量列只到整数位。

    真模板已经修好了（scripts/normalize_tax_inv_number_formats.py），但它是
    runtime 文件、不在仓库里，重新部署可能又拿到旧版本。这里刻意造一份"坏"
    模板，验证 renderer 自己就能把三联拉齐，不依赖模板的状态。
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoice"
    for coordinate in ("D13", "D14", "D16", "D17", "D19", "Q13", "Q14"):
        worksheet[coordinate] = coordinate
    for row in range(DATA_START_ROW, DATA_START_ROW + MAX_ITEMS):
        for column in DATA_COLUMNS:
            worksheet.cell(row=row, column=column).alignment = Alignment(horizontal="right")
        for copy_offset in COPY_COLUMN_OFFSETS[1:]:
            # 数量列：整数位会计格式，1101.25 会印成 "1,101"
            worksheet.cell(row=row, column=9 + copy_offset).number_format = (
                '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"_ ;_ @_ '
            )
            # FOB Rev USD：三位小数，比正本多一位
            worksheet.cell(row=row, column=13 + copy_offset).number_format = (
                '_-* #,##0.000_-;\\-* #,##0.000_-;_-* "-"??_-;_-@_-'
            )
            # 汇率列声明左对齐：会计格式的 "* " 填充本来把它压成右对齐，
            # 一旦换成普通格式就会露出来，所以 renderer 得连对齐一起拉平。
            worksheet.cell(row=row, column=16 + copy_offset).alignment = Alignment(
                horizontal="left"
            )
    path = tmp_path / "broken-template.xlsx"
    workbook.save(path)
    return path


def test_all_three_copies_share_one_set_of_item_formats(tmp_path):
    """三联的数字格式必须逐格一致——副联和正本是同一份税务单据。

    模板里两联 COPY 是 "=B23" 这类镜像公式，公式只搬值不搬格式，所以
    renderer 必须对三联各写一遍。
    """
    items = [_item(number) for number in range(1, 4)]
    content = render_tax_invoice_workbook(
        _template_with_broken_copy_formats(tmp_path), _invoice(), items
    )
    worksheet = load_workbook(BytesIO(content), data_only=False).active

    for index in range(len(items)):
        row = DATA_START_ROW + index
        for column, number_format in ITEM_NUMBER_FORMATS.items():
            formats = {
                worksheet.cell(row=row, column=column + copy_offset).number_format
                for copy_offset in COPY_COLUMN_OFFSETS
            }
            assert formats == {number_format}, (
                f"第 {row} 行第 {column} 列三联格式不一致：{formats}"
            )
        # 对齐同理：普通格式不像会计格式那样能把数字顶到右边，三联的水平
        # 对齐一旦不同，副联的数字就会横向错位。
        for column in DATA_COLUMNS:
            alignments = {
                worksheet.cell(row=row, column=column + copy_offset).alignment.horizontal
                for copy_offset in COPY_COLUMN_OFFSETS
            }
            assert len(alignments) == 1, (
                f"第 {row} 行第 {column} 列三联对齐不一致：{alignments}"
            )


def test_quantity_stays_in_lockstep_between_xlsx_and_pdf(tmp_path):
    """xlsx 的单元格格式和 PDF 的字符串必须同源。

    两条链路各自决定怎么显示数量：xlsx 交给 Excel 按 number_format 渲染，
    PDF 由 format_quantity 直接拼字符串。任一边单独改了，同一张单子的两份
    交付物就会印出不同的数——而且只在有小数的数据上暴露，抽查很难发现。
    """
    item = _item()
    item.quantity = Decimal("4820.5")
    content = render_tax_invoice_workbook(
        _template_with_broken_copy_formats(tmp_path), _invoice(), [item]
    )
    worksheet = load_workbook(BytesIO(content), data_only=False).active

    for copy_offset in COPY_COLUMN_OFFSETS:
        cell = worksheet.cell(row=DATA_START_ROW, column=9 + copy_offset)
        assert cell.number_format == ITEM_NUMBER_FORMATS[9] == "#,##0"
    # Excel 对 "#,##0" 是四舍五入，4820.5 显示成 "4,821"（实测值）；
    # PDF 侧必须给出同一个字符串。
    assert format_quantity(item.quantity) == "4,821"


def test_workbook_rejects_more_than_eighteen_items(tmp_path):
    items = [_item(index) for index in range(1, 20)]
    try:
        render_tax_invoice_workbook(_template(tmp_path), _invoice(), items)
    except TaxInvoiceDocumentGenerationError as exc:
        assert "18" in str(exc)
    else:
        raise AssertionError("expected the 18-line template limit")


def test_file_stem_uses_invoice_date_and_document_number():
    assert (
        tax_invoice_file_stem(_invoice())
        == "20260608-TAX INV(ZWT-IV20260608-01)-THB3245.67"
    )


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        # 产品按个计，数量不带小数，只要千分位分隔符。
        ("1001", "1,001"),
        ("1200", "1,200"),
        ("15", "15"),
        ("0", "0"),
        # 真混进小数时 Excel 的 "#,##0" 是四舍五入而不是截断，
        # 下面四个期望值都是在装了 Excel 的机器上读回来的实际显示结果。
        ("2.5", "3"),
        ("0.5", "1"),
        ("4820.5", "4,821"),
        ("1234.5678", "1,235"),
    ],
)
def test_quantity_format_matches_excel(value: str, expected: str) -> None:
    assert format_quantity(Decimal(value)) == expected


def test_rate_rounds_half_up_like_excel() -> None:
    # 汇率存的是 Numeric(18,6)，模板只显示 4 位；Python 默认银行家舍入会得到
    # 32.4566，Excel 是 32.4567。
    assert format_rate(Decimal("32.456650")) == "32.4567"
    assert format_amount(Decimal("3245.665")) == "3,245.67"


def test_missing_values_render_as_blank() -> None:
    assert format_amount(None) == ""
    assert format_rate(None) == ""
    assert format_quantity(None) == ""


def test_pdf_totals_repeat_the_line_sum_and_thai_text() -> None:
    items = [_item(1), _item(2)]
    _, lines, totals = pdf_field_values(_invoice(), items)

    assert len(lines) == 2
    assert lines[0]["quantity"] == "2"
    assert lines[0]["exchange_rate"] == "32.4567"
    assert lines[0]["fx_date"] == "08/06/2026"
    # 折扣与 VAT 恒为 0，所以合计/折后/总计三格是同一个数。
    assert totals["total_thb"] == "6,491.34"
    assert totals["after_discount_thb"] == "6,491.34"
    assert totals["grand_total_thb"] == "6,491.34"
    assert totals["amount_text_thai"] == (
        "(หกพันสี่ร้อยเก้าสิบเอ็ดบาทสามสิบสี่สตางค์)"
    )


def test_layout_table_covers_every_page_and_line() -> None:
    assert pdf_layout.PAGE_COUNT == 3
    for anchors in (
        *pdf_layout.HEADER_ANCHORS.values(),
        *pdf_layout.ITEM_ANCHORS.values(),
        *pdf_layout.TOTAL_ANCHORS.values(),
    ):
        assert len(anchors) == pdf_layout.PAGE_COUNT
    for baselines in pdf_layout.ITEM_BASELINES.values():
        assert len(baselines) == MAX_ITEMS


def test_generates_three_copy_pdf_without_office(tmp_path) -> None:
    output = tmp_path / "tax-inv.pdf"

    export_pdf_from_template(
        ASSETS / "templates" / "TAX-INV-Template.pdf",
        output,
        _invoice(),
        [_item()],
        ASSETS / "fonts" / "Sarabun-Regular.ttf",
    )

    reader = PdfReader(output)
    # 一页正本 + 两页副本，与模板的三段 print_area 一一对应。
    assert len(reader.pages) == 3
    for page in reader.pages:
        text = page.extract_text() or ""
        assert "ZWT-IV20260608-01" in text
        assert "3,245.67" in text
        assert "P-001" in text


def test_pdf_rejects_more_than_eighteen_lines(tmp_path) -> None:
    items = [_item(index) for index in range(1, MAX_ITEMS + 2)]

    with pytest.raises(TaxInvoiceDocumentGenerationError, match="18"):
        export_pdf_from_template(
            ASSETS / "templates" / "TAX-INV-Template.pdf",
            tmp_path / "tax-inv.pdf",
            _invoice(),
            items,
            ASSETS / "fonts" / "Sarabun-Regular.ttf",
        )


def test_clean_template_clears_data_without_touching_column_headers(tmp_path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoice"
    # 列名和占位符重名，正是这张模板最容易踩的坑：WHT 那种全表按文本匹配的
    # 清理方式会把三联的表头一起抹掉。
    worksheet["H22"] = "Unit"
    worksheet["O22"] = " FX Date"
    worksheet["P22"] = "FX Rate"
    worksheet["D13"] = "Customer Name"
    worksheet["Q13"] = "DocumentNo"
    worksheet["AN13"] = "=V13"
    worksheet["H23"] = "Unit"
    worksheet["AJ23"] = "=R23"
    worksheet["BB40"] = "=AJ40"
    worksheet["R42"] = "=SUM(R23:R40)"
    worksheet["R43"] = "=IF(Q43=0%,0,ROUND(R42*Q43,2))"
    worksheet["R49"] = 0
    path = tmp_path / "template.xlsx"
    workbook.save(path)
    workbook.close()

    cleaned = load_workbook(BytesIO(render_clean_template_workbook(path)))["Invoice"]

    assert cleaned["H22"].value == "Unit"
    assert cleaned["O22"].value == " FX Date"
    assert cleaned["P22"].value == "FX Rate"
    assert cleaned["D13"].value is None
    assert cleaned["Q13"].value is None
    assert cleaned["AN13"].value is None
    assert cleaned["H23"].value is None
    assert cleaned["AJ23"].value is None
    assert cleaned["BB40"].value is None
    assert cleaned["R42"].value is None
    # 折扣与 WHT 恒为 0，Excel 排好的 "-" 留在底版上，叠加层不再画。
    assert cleaned["R43"].value == "=IF(Q43=0%,0,ROUND(R42*Q43,2))"
    assert cleaned["R49"].value == 0


def test_pdf_reports_a_missing_underlay(tmp_path) -> None:
    with pytest.raises(TaxInvoiceDocumentGenerationError, match="PDF template"):
        export_pdf_from_template(
            tmp_path / "absent.pdf",
            tmp_path / "tax-inv.pdf",
            _invoice(),
            [_item()],
            ASSETS / "fonts" / "Sarabun-Regular.ttf",
        )
