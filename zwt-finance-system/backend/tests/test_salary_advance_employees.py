import asyncio
import uuid
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.config import get_settings
from app.modules.salary_advance.employee_excel import (
    build_employee_import_template_workbook,
    parse_employee_sheet,
)
from app.modules.salary_advance.models import SalaryAdvanceEmployee, SalaryAdvanceEvent
from app.modules.salary_advance.schemas import EmployeeCreate, EmployeeUpdate
from app.modules.salary_advance.service import (
    SalaryAdvanceConflictError,
    SalaryAdvanceService,
    SalaryAdvanceStateError,
)


class FakeSession:
    def __init__(self) -> None:
        self.employees: dict[uuid.UUID, SalaryAdvanceEmployee] = {}
        self.events: list[SalaryAdvanceEvent] = []

    def seed(self, emp: SalaryAdvanceEmployee) -> None:
        self.employees[emp.id] = emp

    def add(self, obj: Any) -> None:
        if isinstance(obj, SalaryAdvanceEmployee):
            if obj.id is None:
                obj.id = uuid.uuid4()
            self.employees[obj.id] = obj
        elif isinstance(obj, SalaryAdvanceEvent):
            self.events.append(obj)

    def event_types(self) -> list[str]:
        return [event.event_type for event in self.events]

    async def flush(self) -> None:
        pass

    async def commit(self) -> None:
        pass

    async def refresh(self, obj: Any) -> None:
        pass

    async def scalar(self, statement: Any) -> Any:
        where_clause = str(getattr(statement, "whereclause", ""))
        stmt_str = str(statement).lower()
        if "records" in stmt_str or "salary_advance.records" in stmt_str:
            return None
        is_count = "count(" in stmt_str
        live_only = "deleted_at is null" in where_clause.lower() or "deleted_at is null" in stmt_str

        params = statement.compile().params if hasattr(statement, "compile") else {}
        target_emp_id = None
        target_id = None
        for k, v in params.items():
            if "emp_id" in k:
                target_emp_id = v
            elif k == "id" or k.startswith("id_"):
                target_id = v

        matched = []
        for emp in self.employees.values():
            if live_only and emp.deleted_at is not None:
                continue
            if target_id is not None and emp.id != target_id:
                continue
            if target_emp_id is not None and emp.emp_id != target_emp_id:
                continue
            matched.append(emp)

        if is_count:
            return len(matched)
        if not matched:
            return None
        res = matched[0]
        if hasattr(statement, "_raw_columns") and len(statement._raw_columns) == 1:
            col_str = str(statement._raw_columns[0])
            if "employees.id" in col_str or col_str.endswith(".id"):
                return res.id
        return res

    async def scalars(self, statement: Any) -> Any:
        where_clause = str(getattr(statement, "whereclause", ""))
        live_only = "deleted_at IS NULL" in where_clause or "deleted_at IS NULL" in str(statement)

        compile_params = statement.compile().params if hasattr(statement, "compile") else {}
        param_values = set()
        for v in compile_params.values():
            if isinstance(v, (list, tuple, set)):
                param_values.update(v)
            elif v is not None:
                param_values.add(v)

        matched = []
        for emp in self.employees.values():
            if live_only and emp.deleted_at is not None:
                continue
            if param_values and emp.emp_id not in param_values and emp.id not in param_values:
                continue
            matched.append(emp)

        class _Scalars:
            def __init__(self, items: list[Any]) -> None:
                self._items = items

            def all(self) -> list[Any]:
                return self._items

        return _Scalars(matched)

    async def execute(self, statement: Any) -> Any:
        return await self.scalars(statement)


def test_build_employee_import_template_workbook() -> None:
    content = build_employee_import_template_workbook()
    assert content.startswith(b"PK")

    wb = load_workbook(BytesIO(content))
    assert "员工主数据 Employee Master Data" in wb.sheetnames
    assert "说明 Notes" in wb.sheetnames

    sheet = wb["员工主数据 Employee Master Data"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) >= 2
    assert "工号" in str(rows[0][0])
    assert rows[1][0] == "EMP001"


def test_parse_employee_sheet_success() -> None:
    content = build_employee_import_template_workbook()
    rows = parse_employee_sheet(content)
    assert len(rows) == 1
    row = rows[0]
    assert row["emp_id"] == "EMP001"
    assert row["first_name"] == "Somchai"
    assert row["surname"] == "Saelim"
    assert row["en_name"] == "Somchai Saelim"
    assert row["chinese_name"] == "宋柴"
    assert row["department"] == "IT"
    assert row["position"] == "Software Engineer"
    assert row["start_date"] == date(2024, 1, 15)
    assert row["is_active"] is True


def test_parse_employee_sheet_missing_emp_id_header_raises_error() -> None:
    wb = load_workbook(BytesIO(build_employee_import_template_workbook()))
    ws = wb.active
    ws.cell(row=1, column=1, value="WrongHeader")
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(SalaryAdvanceStateError, match="未找到必需的 '工号'"):
        parse_employee_sheet(buf.getvalue())


def test_parse_employee_sheet_empty_emp_id_cell_raises_error() -> None:
    wb = load_workbook(BytesIO(build_employee_import_template_workbook()))
    ws = wb.active
    ws.cell(row=2, column=1, value="   ")
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(SalaryAdvanceStateError, match="缺失必需的工号"):
        parse_employee_sheet(buf.getvalue())


def test_create_employee_conflict_on_duplicate_active_emp_id() -> None:
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    existing = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP001",
        en_name="Somchai",
        is_active=True,
        created_by_name="TestAdmin",
        updated_by_name="TestAdmin",
    )
    fake_session.seed(existing)

    payload = EmployeeCreate(
        emp_id="EMP001",
        first_name="Somchai",
        surname="Saelim",
        en_name="Somchai Saelim",
        department="IT",
        position="Engineer",
        start_date=date(2024, 1, 15),
    )

    with pytest.raises(SalaryAdvanceConflictError, match="EMP001"):
        asyncio.run(service.create_employee(payload))


def test_create_and_update_employee_success() -> None:
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    payload = EmployeeCreate(
        emp_id="EMP002",
        first_name="Alice",
        surname="Smith",
        en_name="Alice Smith",
        department="Finance",
        position="Accountant",
        start_date=date(2025, 3, 1),
    )

    created = asyncio.run(service.create_employee(payload))
    assert created.emp_id == "EMP002"
    assert created.created_by_name == "TestAdmin"

    update_payload = EmployeeUpdate(
        first_name="Alice",
        surname="Smith",
        en_name="Alice Smith",
        department="Finance",
        position="Senior Accountant",
        start_date=date(2025, 3, 1),
        is_active=True,
    )
    updated = asyncio.run(service.update_employee(created.id, update_payload))
    assert updated.position == "Senior Accountant"


def test_delete_and_restore_employee() -> None:
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    employee = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP003",
        en_name="Bob Jones",
        is_active=True,
        created_by_name="TestAdmin",
        updated_by_name="TestAdmin",
    )
    fake_session.seed(employee)

    # Soft delete
    deleted = asyncio.run(service.delete_employee(employee.id))
    assert deleted.deleted_at is not None
    assert deleted.deleted_by_name == "TestAdmin"

    # Restore
    restored = asyncio.run(service.restore_employee(employee.id))
    assert restored.deleted_at is None
    assert restored.deleted_by_name is None


def test_restore_employee_conflict_if_emp_id_reused() -> None:
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    deleted_emp = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP004",
        en_name="Old Record",
        deleted_at=datetime.now(UTC),
        deleted_by_name="Admin",
        created_by_name="Admin",
        updated_by_name="Admin",
    )
    new_active_emp = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP004",
        en_name="New Active Record",
        created_by_name="Admin",
        updated_by_name="Admin",
    )
    fake_session.seed(deleted_emp)
    fake_session.seed(new_active_emp)

    with pytest.raises(SalaryAdvanceConflictError, match="EMP004"):
        asyncio.run(service.restore_employee(deleted_emp.id))


def test_import_employees_upsert() -> None:
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    existing = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP005",
        en_name="Charlie V1",
        is_active=False,
        created_by_name="Admin",
        updated_by_name="Admin",
    )
    fake_session.seed(existing)

    rows = [
        {
            "emp_id": "EMP005",
            "first_name": "Charlie",
            "surname": "Brown",
            "en_name": "Charlie Brown",
            "chinese_name": "查理",
            "department": "Sales",
            "position": "Manager",
            "start_date": date(2023, 5, 10),
            "is_active": True,
        },
        {
            "emp_id": "EMP006",
            "first_name": "David",
            "surname": "Miller",
            "en_name": "David Miller",
            "chinese_name": "戴维",
            "department": "HR",
            "position": "Specialist",
            "start_date": date(2024, 2, 1),
            "is_active": True,
        },
    ]

    result = asyncio.run(service.import_employees(rows, "import_test.xlsx"))
    assert result.created == 1
    assert result.updated == 1
    assert result.source_file_name == "import_test.xlsx"


def test_import_marks_existing_employee_as_resigned() -> None:
    """导入把在职员工改成离职必须真的生效。

    原先 update 分支在写完 row 之后又硬写了一次 is_active=True，
    模板里的「是否在职=否」被悄悄丢掉——离职的人仍然出现在单张开具的下拉里。
    create 分支没有这个覆盖，所以两条路径的行为当时是不一致的。
    """
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )
    existing = SalaryAdvanceEmployee(
        id=uuid.uuid4(),
        emp_id="EMP007",
        en_name="Leaver",
        is_active=True,
        created_by_name="Admin",
        updated_by_name="Admin",
    )
    fake_session.seed(existing)

    asyncio.run(
        service.import_employees(
            [{"emp_id": "EMP007", "en_name": "Leaver", "is_active": False}],
            "leavers.xlsx",
        )
    )

    assert existing.is_active is False


def test_employee_mutations_write_audit_events() -> None:
    """人员库的增删改必须留痕，与本模块其他写操作一致。

    员工资料直接决定预支单上印的姓名/部门/职位/入职日期，
    改动无痕等于单据来源不可追溯。
    """
    fake_session = FakeSession()
    service = SalaryAdvanceService(
        session=fake_session, actor_name="TestAdmin", settings=get_settings()
    )

    created = asyncio.run(
        service.create_employee(
            EmployeeCreate(
                emp_id="EMP008",
                first_name="Eve",
                surname="Stone",
                department="Finance",
                position="Analyst",
                start_date=date(2024, 6, 1),
            )
        )
    )
    asyncio.run(service.delete_employee(created.id))
    asyncio.run(service.restore_employee(created.id))

    assert fake_session.event_types() == ["created", "deleted", "restored"]
    assert all(event.actor_name == "TestAdmin" for event in fake_session.events)


def test_create_employee_rejects_record_that_could_never_be_issued() -> None:
    """人员库按开单口径必填：存得下、开不出来的员工不该存在。

    这五个字段是预支单上要印的内容，validate_and_normalize_record 当作必填。
    在录入这一步拦下来，报错才落在有修改入口的地方。
    """
    with pytest.raises(ValidationError):
        EmployeeCreate(emp_id="EMP009", first_name="Only", surname="Name")


@pytest.mark.parametrize(
    ("value", "expected"),
    [("是", True), ("否", False), ("YES", True), ("no", False), (None, True)],
)
def test_parse_employee_sheet_reads_is_active_both_ways(value: object, expected: bool) -> None:
    wb = load_workbook(BytesIO(build_employee_import_template_workbook()))
    ws = wb.active
    ws.cell(row=2, column=9, value=value)
    buf = BytesIO()
    wb.save(buf)

    assert parse_employee_sheet(buf.getvalue())[0]["is_active"] is expected


def test_parse_employee_sheet_rejects_unreadable_date() -> None:
    """认不出的日期要当场退回，不能静默变 None。

    静默的代价出现在别处：导入报"成功"，等到开预支单时才炸出
    "入职日期缺失或无法解析"，那时人已经不在导入这一步了。
    """
    wb = load_workbook(BytesIO(build_employee_import_template_workbook()))
    ws = wb.active
    ws.cell(row=2, column=8, value="15 Jan 2024")
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(SalaryAdvanceStateError, match="15 Jan 2024"):
        parse_employee_sheet(buf.getvalue())


def test_parse_employee_sheet_rejects_unreadable_is_active() -> None:
    wb = load_workbook(BytesIO(build_employee_import_template_workbook()))
    ws = wb.active
    ws.cell(row=2, column=9, value="待定")
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(SalaryAdvanceStateError, match="待定"):
        parse_employee_sheet(buf.getvalue())
