from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.modules.wht.batch_import import (
    TEMPLATE_COLUMNS,
    BatchWorkbookError,
    build_template_workbook,
    parse_batch_sheet,
    resolve_rate,
)
from app.modules.wht.income_types import default_rate, find, options_for

HEADERS = list(TEMPLATE_COLUMNS)
VALID_ROW = ["0105540057561", "ค่าบริการ", "2026-06-05", 3000, "2026-06", "normal", None, None]


def workbook_bytes(*rows: list) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def row_errors(*rows: list) -> list[str]:
    """行级问题被汇总进 errors，顶层消息只说"有几行要改"，所以断言看 errors。"""
    with pytest.raises(BatchWorkbookError) as excinfo:
        parse_batch_sheet(workbook_bytes(*rows))
    return excinfo.value.errors


def test_parses_a_valid_row() -> None:
    (row,) = parse_batch_sheet(workbook_bytes(VALID_ROW))

    assert row["payee_tax_id"] == "0105540057561"
    assert row["income_type"] == "ค่าบริการ"
    assert row["payment_date"].isoformat() == "2026-06-05"
    assert row["total_amount"] == Decimal("3000")
    assert row["period"] == "2026-06"
    assert row["issuance_type"] == "normal"
    assert row["supplement_run"] == 0
    assert row["wht_rate"] is None


def test_derives_period_from_payment_date_when_blank() -> None:
    row = ["0105540057561", "ค่าบริการ", "2026-06-05", 3000, None, None, None, None]

    (parsed,) = parse_batch_sheet(workbook_bytes(row))

    assert parsed["period"] == "2026-06"


def test_restores_leading_zero_on_tax_id_stored_as_number() -> None:
    row = [105540057561, "ค่าบริการ", "2026-06-05", 3000, None, None, None, None]

    (parsed,) = parse_batch_sheet(workbook_bytes(row))

    assert parsed["payee_tax_id"] == "0105540057561"


@pytest.mark.parametrize(
    ("raw", "expected"),
    [(0.03, Decimal("0.03")), ("3%", Decimal("0.03")), ("0.05", Decimal("0.05"))],
)
def test_accepts_decimal_and_percent_rates(raw: object, expected: Decimal) -> None:
    row = ["0105540057561", "ค่าบริการ", "2026-06-05", 3000, None, None, None, raw]

    (parsed,) = parse_batch_sheet(workbook_bytes(row))

    assert parsed["wht_rate"] == expected


def test_rejects_a_bare_number_rate_as_ambiguous() -> None:
    # 3 既可能是 3% 也可能是 300%，猜错会静默算错税额，所以直接拒收。
    row = ["0105540057561", "ค่าบริการ", "2026-06-05", 3000, None, None, None, 3]

    assert "TaxRate must be a decimal" in row_errors(row)[0]


def test_rejects_day_first_dates() -> None:
    row = ["0105540057561", "ค่าบริการ", "05/06/2026", 3000, None, None, None, None]

    assert "invalid date value" in row_errors(row)[0]


def test_pads_a_tax_id_whose_leading_zeros_excel_dropped() -> None:
    # Excel 把税号存成数字就会丢前导零，补零是必要的；补出来对不上主数据的，
    # 会在建单阶段被 "no payee with taxId ..." 挡住，不会静默建错单。
    row = ["123", "ค่าบริการ", "2026-06-05", 3000, None, None, None, None]

    (parsed,) = parse_batch_sheet(workbook_bytes(row))

    assert parsed["payee_tax_id"] == "0000000000123"


def test_rejects_a_tax_id_that_is_too_long() -> None:
    row = ["12345678901234", "ค่าบริการ", "2026-06-05", 3000, None, None, None, None]

    assert "PayeeTaxID must be 13 digits" in row_errors(row)[0]


def test_rejects_a_historical_ledger_by_its_refno_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["RefNo", *HEADERS])
    sheet.append(["ZWT202606001", *VALID_ROW])
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(BatchWorkbookError, match="historical ledger"):
        parse_batch_sheet(buffer.getvalue())


def test_reports_every_bad_row_at_once_and_imports_nothing() -> None:
    no_income = ["0105540057561", None, "2026-06-05", 3000, None, None, None, None]
    bad_amount = ["0105540057561", "ค่าบริการ", "2026-06-05", 0, None, None, None, None]

    with pytest.raises(BatchWorkbookError) as excinfo:
        parse_batch_sheet(workbook_bytes(VALID_ROW, no_income, bad_amount))

    # 一次列全，用户改一轮就能通过，而不是传一次修一行。
    assert len(excinfo.value.errors) == 2
    assert "row 3" in excinfo.value.errors[0]
    assert "row 4" in excinfo.value.errors[1]
    assert "nothing was imported" in str(excinfo.value)


def test_rejects_missing_required_columns() -> None:
    workbook = Workbook()
    workbook.active.append(["PayeeTaxID", "Amount"])
    workbook.active.append(["0105540057561", 3000])
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(BatchWorkbookError, match="IncomeType"):
        parse_batch_sheet(buffer.getvalue())


def test_supplement_rows_need_a_run_number() -> None:
    row = ["0105540057561", "ค่าบริการ", "2026-06-05", 3000, None, "supplement", 0, None]

    assert "SupplementRun must be 1-9" in row_errors(row)[0]


def test_template_carries_the_documented_columns_and_notes() -> None:
    workbook = load_workbook(BytesIO(build_template_workbook()))

    assert [cell.value for cell in workbook.worksheets[0][1]] == HEADERS
    assert "说明 Notes" in workbook.sheetnames
    # 模板自身必须能被解析器接受，否则用户下载下来填完必然被退回。
    assert parse_batch_sheet(build_template_workbook())


class TestIncomeTypes:
    def test_catalogue_matches_the_rates_in_the_historical_ledger(self) -> None:
        # 来自 Sample_previous_code/WHT/data历史 的 234 条实际记录。
        assert default_rate("ค่าบริการ", "PND53") == Decimal("0.03")
        assert default_rate("ค่าขนส่ง", "PND53") == Decimal("0.01")
        assert default_rate("ค่าเช่า", "PND3") == Decimal("0.05")

    def test_interest_rate_depends_on_which_return_it_goes_on(self) -> None:
        assert default_rate("ดอกเบี้ย", "PND3") == Decimal("0.15")
        assert default_rate("ดอกเบี้ย", "PND53") == Decimal("0.01")

    def test_options_are_filtered_by_return_type(self) -> None:
        pnd3 = {option.code for option in options_for("PND3")}

        assert "non_life_insurance_premium" not in pnd3
        assert "non_life_insurance_premium" in {
            option.code for option in options_for("PND53")
        }

    def test_lookup_accepts_both_the_code_and_the_thai_label(self) -> None:
        assert find("service_fee") is find("ค่าบริการ")
        assert find("ไม่มีอยู่จริง") is None

    def test_row_rate_wins_over_the_catalogue_default(self) -> None:
        row = {"income_type": "ค่าบริการ", "wht_rate": Decimal("0.02")}

        assert resolve_rate(row, "PND53") == Decimal("0.02")

    def test_custom_income_type_without_a_rate_cannot_be_resolved(self) -> None:
        row = {"income_type": "ค่าอื่น ๆ", "wht_rate": None}

        assert resolve_rate(row, "PND53") is None
