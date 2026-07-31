from io import BytesIO

import pytest
from openpyxl import Workbook

from app.modules.salary_advance.importer import (
    IMPORT_TEMPLATE_SHEET,
    SalaryAdvanceImportError,
    build_import_template_workbook,
    parse_salary_advance_workbook,
)
from app.modules.salary_advance.validation import FIELD_ORDER

HEADERS = (
    "期间",
    "Emp.ID 工号",
    "名 First Name",
    "姓 Surname",
    "Dept 部门",
    "Position 职位",
    "开始工作日期",
    "签字日期",
    "预支金额",
    "每月扣款金额",
    "审批结果",
    "申请人签名方式",
    "财务签名代码",
    "总经理签名代码",
)


def _workbook_bytes(rows: list[tuple[object, ...]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "导入数据模板"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _row(emp_id: str, period: str = "202607") -> tuple[object, ...]:
    return (
        period,
        emp_id,
        "SOMCHAI",
        "TEST",
        "Finance",
        "Accountant",
        "2024-01-01",
        "2026-07-20",
        5000,
        1000,
        "Approve",
        "Handwritten",
        "FIN_TEST",
        "MD_TEST",
    )


def test_import_filters_period_and_keeps_source_row_numbers() -> None:
    parsed = parse_salary_advance_workbook(
        _workbook_bytes([_row("E001"), _row("E002", "202608")]),
        period="202607",
        active_signature_codes={"FIN_TEST", "MD_TEST"},
    )

    assert len(parsed.records) == 1
    assert parsed.records[0].source_row_no == 2
    assert parsed.records[0].emp_id == "E001"
    assert parsed.records[0].validation_status in {"valid", "warning"}
    assert len(parsed.source_sha256) == 64


def test_import_marks_batch_and_existing_duplicates() -> None:
    parsed = parse_salary_advance_workbook(
        _workbook_bytes([_row("E001"), _row("E001"), _row("E002")]),
        period="202607",
        active_signature_codes={"FIN_TEST", "MD_TEST"},
        existing_keys={("202607", "E002")},
    )

    assert parsed.invalid_rows == 2
    codes = {
        issue["code"]
        for record in parsed.records
        for issue in record.validation_errors
    }
    assert "DUPLICATE_PERIOD_EMPLOYEE" in codes
    assert "DUPLICATE_ACTIVE_RECORD" in codes


def test_import_rejects_non_xlsx_zip_content() -> None:
    with pytest.raises(SalaryAdvanceImportError, match="有效"):
        parse_salary_advance_workbook(
            b"not-an-xlsx",
            period="202607",
            active_signature_codes=set(),
        )


def test_import_template_columns_and_example_align() -> None:
    """列头与示例行长度必须都等于 FIELD_ORDER，否则示例行会串列。"""
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(build_import_template_workbook()))
    try:
        assert workbook.sheetnames[0] == IMPORT_TEMPLATE_SHEET
        sheet = workbook[IMPORT_TEMPLATE_SHEET]
        header = [cell.value for cell in sheet[1]]
        example = [cell.value for cell in sheet[2]]
        assert len(header) == len(FIELD_ORDER)
        assert len(example) == len(FIELD_ORDER)
    finally:
        workbook.close()


def test_generated_template_round_trips_through_parser() -> None:
    """把导出的模板原样喂回解析器：示例行必须零错误，模板与解析口径不许脱节。"""
    parsed = parse_salary_advance_workbook(
        build_import_template_workbook(),
        period="202607",
        active_signature_codes={"FIN_XING_LANHUI", "MD_GONG_YAOWEN"},
    )

    assert len(parsed.records) == 1
    record = parsed.records[0]
    assert record.validation_errors == []
    assert record.validation_status in {"valid", "warning"}
    assert record.emp_id == "E001"
