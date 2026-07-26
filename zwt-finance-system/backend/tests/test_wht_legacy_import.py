from io import BytesIO

import pytest
from openpyxl import Workbook

from app.modules.wht.legacy_import import (
    LegacyWorkbookError,
    parse_historical_number,
    parse_payee_sheet,
    parse_task_sheet,
)

TASK_HEADERS = [
    "No",
    "RefNo",
    "BookNo",
    "PayeeNameEN",
    "PayeeNameTH",
    "PayeeAddress",
    "PayeeTaxID",
    "WHT Type3",
    "WHT Type53",
    "IncomeType",
    "PaymentDate",
    "Amount",
    "TaxRate",
    "TaxAmount",
    "AmountTextThai",
    "DateTextThai",
]


def workbook_bytes() -> bytes:
    workbook = Workbook()
    task_sheet = workbook.active
    task_sheet.title = "Sheet1"
    task_sheet.append(TASK_HEADERS)
    task_sheet.append(
        [
            1,
            "ZWT202606001",
            "202606",
            "Example Co., Ltd.",
            "บริษัท ตัวอย่าง จำกัด",
            "กรุงเทพมหานคร",
            "0105540057561",
            None,
            "√",
            "ค่าบริการ",
            "2026/6/5",
            3000,
            0.03,
            90,
            "-- เก้าสิบบาทถ้วน --",
            "5 มิถุนายน 2569",
        ]
    )
    payee_sheet = workbook.create_sheet("Sheet2")
    payee_sheet.append(["NO", "TAXID", "Name", "Address", "Type", None])
    payee_sheet.append(
        [
            1,
            "0105540057561",
            "บริษัท ตัวอย่าง จำกัด",
            "กรุงเทพมหานคร",
            "PND53",
            "Example Co. / บริษัทตัวอย่าง",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_parses_second_sheet_as_payee_master_data() -> None:
    rows = parse_payee_sheet(workbook_bytes())

    assert rows == [
        {
            "tax_id": "0105540057561",
            "name_th": "บริษัท ตัวอย่าง จำกัด",
            "address_th": "กรุงเทพมหานคร",
            "wht_type": "PND53",
            "aliases": ["Example Co.", "บริษัทตัวอย่าง"],
        }
    ]


def test_parses_first_sheet_as_historical_task() -> None:
    rows = parse_task_sheet(workbook_bytes())

    assert len(rows) == 1
    assert rows[0]["task_no"] == "ZWT202606001"
    assert rows[0]["period"] == "2026-06"
    assert rows[0]["wht_type"] == "PND53"
    assert str(rows[0]["wht_amount"]) == "90"


def test_parses_historical_supplement_number() -> None:
    scope = parse_historical_number("ZWT202606BK201")

    assert scope == {
        "period": "2026-06",
        "issuance_type": "supplement",
        "supplement_run": 2,
        "sequence": 1,
    }


def test_rejects_workbook_without_second_sheet() -> None:
    workbook = Workbook()
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(LegacyWorkbookError, match="second payee data sheet"):
        parse_payee_sheet(output.getvalue())
