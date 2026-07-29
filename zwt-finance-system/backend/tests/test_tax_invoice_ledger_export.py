"""台账导出与批量导入冲突汇总。

导出的重点不是"能生成一个 Excel"，而是生成出来的 Excel 能被
parse_sample_workbook 原样吃回去——导出→核对→改→导回是这个功能存在的理由，
表头一旦两边走样，整条链路就断了。所以这里的主测试是往返，不是快照。
"""

from __future__ import annotations

import asyncio
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.config import get_settings
from app.main import app
from app.modules.tax_invoice.ledger_export import (
    MONEY_FORMAT,
    PRICE_FORMAT,
    QUANTITY_FORMAT,
    build_ledger_workbook,
)
from app.modules.tax_invoice.models import TaxInvoice, TaxInvoiceItem
from app.modules.tax_invoice.recognition import parse_sample_workbook
from app.modules.tax_invoice.router import get_tax_invoice_service
from app.modules.tax_invoice.service import (
    APPROVAL_ITEM_LIMIT,
    TaxInvoiceConflictError,
    TaxInvoiceService,
    check_approval_readiness,
)


def _invoice(**overrides: object) -> TaxInvoice:
    fields: dict[str, object] = {
        "document_no": "ZWT-IV20260605-01",
        "status": "approved",
        "ci_no": "ZWT-TEST-001",
        "cdn": "A0099887766",
        "ci_date": date(2026, 6, 1),
        "invoice_date": date(2026, 6, 5),
        "exchange_target_date": date(2026, 6, 4),
        "exchange_rate": Decimal("36.1234"),
        "revenue_period": "202606",
        "customer_name": "TEST CUSTOMER",
        "customer_address": "BANGKOK",
        "tax_id": "0105500000001",
        "po_no": "PO-001",
        "incoterms": "FCA",
        "payment_term": "OA 30 DAYS",
    }
    fields.update(overrides)
    return TaxInvoice(**fields)


def _item(line_number: int = 1, **overrides: object) -> TaxInvoiceItem:
    fields: dict[str, object] = {
        "line_number": line_number,
        "product_name": "ROUTER",
        "product_code": "R-1",
        "hs_code": "85176243",
        "unit": "PIECES",
        "quantity": Decimal("10"),
        "ci_unit_price": Decimal("30.0000"),
        "fob_unit_price_usd": Decimal("30.0000"),
        "fob_revenue_usd": Decimal("300.00"),
        "fob_revenue_thb": Decimal("10837.02"),
    }
    fields.update(overrides)
    return TaxInvoiceItem(**fields)


def test_exported_workbook_can_be_parsed_back_by_the_sample_importer() -> None:
    """导出 → 再导入的往返。表头两边走样时这条会先炸。"""
    content = build_ledger_workbook(
        [(_invoice(), [_item(1), _item(2, product_name="SWITCH", product_code="S-2")])]
    )

    parsed = parse_sample_workbook(content, "ledger.xlsx")

    assert len(parsed) == 1
    invoice = parsed[0]
    assert invoice["document_no"] == "ZWT-IV20260605-01"
    assert invoice["ci_no"] == "ZWT-TEST-001"
    assert invoice["cdn"] == "A0099887766"
    assert invoice["customer_name"] == "TEST CUSTOMER"
    assert invoice["exchange_target_date"] == date(2026, 6, 4)
    assert invoice["exchange_rate"] == Decimal("36.1234")
    # 两条商品来自两行，且顺序保持导出时的顺序。
    assert [item["product_name"] for item in invoice["items"]] == ["ROUTER", "SWITCH"]
    assert invoice["fob_revenue_usd_total"] == Decimal("600.00")


def test_two_invoices_stay_separate_through_the_round_trip() -> None:
    """分组靠 DocumentNo/CDN/C-I No.，导出时每行都重复税票级字段也不能粘连。"""
    content = build_ledger_workbook(
        [
            (_invoice(), [_item(1)]),
            (
                _invoice(
                    document_no="ZWT-IV20260605-02",
                    ci_no="ZWT-TEST-002",
                    cdn="B0011223344",
                    customer_name="OTHER CUSTOMER",
                ),
                [_item(1), _item(2)],
            ),
        ]
    )

    parsed = parse_sample_workbook(content, "ledger.xlsx")

    assert len(parsed) == 2
    assert {invoice["ci_no"] for invoice in parsed} == {"ZWT-TEST-001", "ZWT-TEST-002"}
    assert sorted(len(invoice["items"]) for invoice in parsed) == [1, 2]


def test_number_formats_follow_the_agreed_precision() -> None:
    """数量不带小数、单价 4 位、金额 2 位。人照着 Excel 显示的精度对账，
    格式写错等于把错的口径写进了对账依据。"""
    content = build_ledger_workbook([(_invoice(), [_item()])])
    sheet = load_workbook(BytesIO(content)).active

    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(row=2, column=headers["Quantity"]).number_format == QUANTITY_FORMAT
    assert sheet.cell(row=2, column=headers["FOB Unit Price USD"]).number_format == PRICE_FORMAT
    assert sheet.cell(row=2, column=headers["FOB Rev USD"]).number_format == MONEY_FORMAT
    assert sheet.cell(row=2, column=headers["FOB Rev THB"]).number_format == MONEY_FORMAT


def test_an_invoice_without_items_still_occupies_a_row() -> None:
    """没有明细的税票不能在导出里凭空消失，否则对账看不出缺口。"""
    content = build_ledger_workbook([(_invoice(), [])])
    sheet = load_workbook(BytesIO(content)).active

    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=headers["C/I No."]).value == "ZWT-TEST-001"
    assert sheet.cell(row=2, column=headers["Product Name or Service"]).value is None


def test_status_column_is_exported_but_ignored_on_reimport() -> None:
    """Status 是给人核对用的，解析器不认识这一列，多带它不该影响回导。"""
    content = build_ledger_workbook([(_invoice(status="needs_review"), [_item()])])
    sheet = load_workbook(BytesIO(content)).active
    headers = {cell.value: cell.column for cell in sheet[1]}

    assert sheet.cell(row=2, column=headers["Status"]).value == "needs_review"
    # 依然解析得出来，且状态不会被当成税票字段带进去。
    parsed = parse_sample_workbook(content, "ledger.xlsx")
    assert "status" not in parsed[0]


def test_export_route_wins_over_the_uuid_route(admin_client) -> None:  # noqa: ANN001
    """/invoices/export 必须排在 /invoices/{invoice_id} 前面。

    后者的 invoice_id 是 UUID，"export" 命中它只会得到 422。FastAPI 按注册顺序
    匹配，所以这条纯粹是在守住 router.py 里两个装饰器的先后——有人重排文件时
    不该悄无声息地把导出打成 422。
    """

    class _StubService:
        async def export_entries(self, **_kwargs: object) -> list:
            return [(_invoice(), [_item()])]

    app.dependency_overrides[get_tax_invoice_service] = lambda: _StubService()
    try:
        response = admin_client.get("/api/v1/tax-invoice/invoices/export")
    finally:
        app.dependency_overrides.pop(get_tax_invoice_service, None)

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "tax-inv-ledger-" in response.headers["content-disposition"]
    # 回读一遍，确认下载到的确实是能再导入的工作簿而不是空壳。
    assert parse_sample_workbook(response.content, "ledger.xlsx")[0]["ci_no"] == "ZWT-TEST-001"


class _EmptyLedgerSession:
    """假会话：台账里什么都没有，所有存在性查询都返回 None。

    这样测出来的就是"同一份文件内部的冲突有没有被一次性收全"，
    不用连库，也不用去猜 SQLAlchemy 语句长什么样。
    """

    async def scalar(self, *_args: object, **_kwargs: object) -> None:
        return None


def _service() -> TaxInvoiceService:
    return TaxInvoiceService(
        session=_EmptyLedgerSession(),  # type: ignore[arg-type]
        actor_name="测试审批人",
        settings=get_settings(),
    )


def _row(ci_no: str, cdn: str | None, source_rows: list[int], **extra: object) -> dict:
    """一行没有编号的普通导入行——完整性由之后的人工批准把关。"""
    return {"ci_no": ci_no, "cdn": cdn, "source_rows": source_rows, **extra}


def _numbered_row(
    ci_no: str,
    cdn: str | None,
    source_rows: list[int],
    document_no: str,
    **overrides: object,
) -> dict:
    """一行带编号的导入行，其余字段填满到「够格直接批准」。

    编号现在一律被拒，所以这个夹具只用来构造"该被拒"的输入。其余字段之所以
    仍然填满：要证明拒绝的理由是编号本身，而不是顺带缺了别的字段。
    """
    row: dict = {
        "ci_no": ci_no,
        "cdn": cdn,
        "source_rows": source_rows,
        "document_no": document_no,
        "invoice_date": date(2026, 6, 5),
        "exchange_target_date": date(2026, 6, 4),
        "exchange_rate": Decimal("36.1234"),
        "customer_name": "TEST CUSTOMER",
        "customer_address": "BANGKOK",
        "items": [{"line_number": 1}],
    }
    row.update(overrides)
    return row


def _check(rows: list[dict]) -> None:
    """同步地跑一遍 _assert_importable。

    测试栈里没装 pytest-asyncio / anyio 插件（CI 也只装 pytest+cov+ruff），
    直接写 async def 测试会被静默跳过。这里自己起一个事件循环，不引入新依赖。
    """
    asyncio.run(_service()._assert_importable(rows))


def test_all_conflicts_are_reported_in_one_pass() -> None:
    """以前查到第一条就抛，用户改一行再导一次；现在一次列全。"""
    rows = [
        _row("CI-1", "CDN-1", [2]),
        _row("CI-1", "CDN-1", [3]),  # 与第 2 行同一张税票
        _row("CI-2", "CDN-2", [4]),
        _row("CI-2", "CDN-2", [5]),  # 与第 4 行同一张税票
    ]

    with pytest.raises(TaxInvoiceConflictError) as excinfo:
        _check(rows)

    issues = excinfo.value.issues
    assert len(issues) == 2, "两处冲突必须一次全报出来，而不是只报第一处"
    assert [issue["rows"] for issue in issues] == [[3], [5]]
    assert {issue["reason"] for issue in issues} == {"duplicate_in_file"}
    # 摘要里要带上总数，前端不展开也能看出规模。
    assert "2 problem" in str(excinfo.value)


def test_a_clean_file_raises_nothing() -> None:
    """留空编号才是导入的正常形态。

    没编号的行不受完整性门槛约束：它落成 needs_review，等人工批准时再校验。
    """
    _check(
        [
            _row("CI-1", "CDN-1", [2]),
            _row("CI-2", "CDN-2", [3]),
            _row("CI-3", "CDN-3", [4]),
        ]
    )


def test_an_incomplete_row_without_a_number_is_allowed_through() -> None:
    """缺字段但没编号，就没绕过任何东西，它会落成 needs_review 等人工补。
    在导入这一步拦它等于把「先导进来再慢慢补」这条正常路堵死。"""
    _check([_row("CI-1", "CDN-1", [2])])


def test_import_rejects_any_supplied_number() -> None:
    """导入一律不接受文件里的编号，没有例外。

    「编号在批准时由数据库事务生成」这条规则的实际执行点就在这里；界面上那句
    「请留空」只是提示，拦不住一张手工改过的表。曾经有过一条允许沿用旧编号的
    历史迁移通道，已于 2026-07-29 摘除——补开以前月份的票靠行里的 Invoice
    Date，编号会跟着那一天发，不需要外部指定。
    """
    rows = [_numbered_row("CI-1", "CDN-1", [2], "ZWT-IV20260605-01")]

    with pytest.raises(TaxInvoiceConflictError) as excinfo:
        _check(rows)

    issue = excinfo.value.issues[0]
    assert issue["reason"] == "number_not_allowed"
    assert issue["rows"] == [2]
    assert issue["key"] == "ZWT-IV20260605-01"


def test_every_numbered_row_is_reported_not_just_the_first() -> None:
    """整份表里每一行带编号的都要报出来，一次改完。

    这也顺带钉住「不再有编号相关的其他分支」：两行同号在迁移时代会报
    duplicate_number_in_file，现在两行各自只报一次 number_not_allowed。
    """
    rows = [
        _numbered_row("CI-1", "CDN-1", [2], "ZWT-IV20260605-01"),
        _numbered_row("CI-2", "CDN-2", [3], "ZWT-IV20260605-01"),
    ]

    with pytest.raises(TaxInvoiceConflictError) as excinfo:
        _check(rows)

    issues = excinfo.value.issues
    assert [issue["reason"] for issue in issues] == ["number_not_allowed"] * 2
    assert [issue["rows"] for issue in issues] == [[2], [3]]


def test_a_rejected_number_is_not_also_reported_as_incomplete() -> None:
    """编号本身就不该出现时，不必再抱怨这行缺哪些字段——那是噪音，
    会让用户以为补齐字段就能过。"""
    rows = [
        _numbered_row(
            "CI-1",
            "CDN-1",
            [2],
            "ZWT-IV20260605-01",
            customer_address=None,
        )
    ]

    with pytest.raises(TaxInvoiceConflictError) as excinfo:
        _check(rows)

    assert [issue["reason"] for issue in excinfo.value.issues] == ["number_not_allowed"]


def test_sample_import_records_its_mode_and_supplies_no_number_flag() -> None:
    """导入模式要在审计里分得开，且不再有"允许带编号"这档开关。

    `allow_existing_numbers` 参数随历史迁移一并删掉了。这里断言它没被悄悄
    加回来——参数一旦回来，例外就又可能被打开。
    """
    from app.modules.tax_invoice import service as service_module

    recorded: list[dict] = []

    async def _capture(self, **kwargs: object) -> None:  # noqa: ANN001
        recorded.append(kwargs)

    original = service_module.TaxInvoiceService._create_import
    service_module.TaxInvoiceService._create_import = _capture
    try:
        asyncio.run(_service().import_sample(rows=[], file_name="a.xlsx", content=b""))
    finally:
        service_module.TaxInvoiceService._create_import = original

    assert recorded[0]["import_mode"] == "sample"
    assert "allow_existing_numbers" not in recorded[0]
    assert not hasattr(service_module.TaxInvoiceService, "import_migration")


# --- 可批准门槛 ---------------------------------------------------------------
#
# check_approval_readiness 是通往 approved 的唯一守门人（历史迁移那条能绕过它
# 的路已摘除）。这几条直接测它，不经过导入——原先 18 行上限只被导入里的
# incomplete_for_number 分支覆盖，那个分支随迁移一起没了。


def test_approval_readiness_reports_every_missing_field() -> None:
    readiness = check_approval_readiness(
        invoice_date=date(2026, 6, 5),
        exchange_target_date=None,
        exchange_rate=None,
        customer_name="TEST CUSTOMER",
        customer_address=None,
        cdn="A0099887766",
        item_count=1,
    )

    assert not readiness.ok
    assert set(readiness.blockers()) == {
        "exchangeTargetDate",
        "exchangeRate",
        "customerAddress",
    }


def test_approval_readiness_rejects_more_than_the_template_limit() -> None:
    """19 条商品的税票要等到生成文件时才炸，那时编号早发出去了。

    模板物理容量是 18 行，超过必须禁止批准、不得截断（已锁定的业务规则）。
    """
    complete = {
        "invoice_date": date(2026, 6, 5),
        "exchange_target_date": date(2026, 6, 4),
        "exchange_rate": Decimal("36.1234"),
        "customer_name": "TEST CUSTOMER",
        "customer_address": "BANGKOK",
        "cdn": "A0099887766",
    }

    assert check_approval_readiness(**complete, item_count=APPROVAL_ITEM_LIMIT).ok

    over = check_approval_readiness(**complete, item_count=APPROVAL_ITEM_LIMIT + 1)
    assert not over.ok
    assert over.blockers() == ["itemLimit"]


def test_approval_readiness_requires_at_least_one_item() -> None:
    readiness = check_approval_readiness(
        invoice_date=date(2026, 6, 5),
        exchange_target_date=date(2026, 6, 4),
        exchange_rate=Decimal("36.1234"),
        customer_name="TEST CUSTOMER",
        customer_address="BANGKOK",
        cdn="A0099887766",
        item_count=0,
    )

    assert readiness.blockers() == ["items"]
