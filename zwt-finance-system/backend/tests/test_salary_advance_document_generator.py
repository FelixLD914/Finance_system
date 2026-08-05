import dataclasses
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw
from pypdf import PdfReader

from app.core.config import get_settings
from app.modules.salary_advance import pdf_layout
from app.modules.salary_advance.document_generator import (
    FORM_FONT_NAME,
    WRAPPED_FIELDS,
    GenerationSnapshot,
    export_pdf_from_template,
    render_salary_advance_workbook,
    sha256_path,
)


def _signature(path: Path, color: tuple[int, int, int, int]) -> None:
    image = Image.new("RGBA", (220, 70), (255, 255, 255, 0))
    draw = ImageDraw.Draw(image)
    draw.line((10, 55, 70, 12, 125, 48, 210, 14), fill=color, width=5)
    image.save(path)


def _snapshot(tmp_path: Path, position: str = "Accountant") -> GenerationSnapshot:
    finance = tmp_path / "finance.png"
    managing_director = tmp_path / "md.png"
    _signature(finance, (20, 35, 120, 255))
    _signature(managing_director, (120, 25, 25, 255))
    normalized = {
        "period": "202607",
        "emp_id": "E001",
        "en_name": "SOMCHAI TEST",
        "first_name": "SOMCHAI",
        "surname": "TEST",
        "chinese_name": "",
        "applicant_display_name": "SOMCHAI TEST",
        "department": "Finance",
        "position": position,
        "start_date": "2024-01-15",
        "reason": "Living expenses",
        "advance_amount": "12500.50",
        "advance_amount_words_th": "หนึ่งหมื่นสองพันห้าร้อยบาทห้าสิบสตางค์",
        "monthly_deduction": "2500.10",
        "monthly_deduction_words_th": "สองพันห้าร้อยบาทสิบสตางค์",
        "request_date": "2026-07-20",
        "finance_comment": "Verified",
        "finance_display_name": "FINANCE TEST",
        "finance_date": "2026-07-20",
        "approval_status": "Approve",
        "md_display_name": "MD TEST",
        "md_date": "2026-07-20",
        "applicant_signature_mode": "Handwritten",
        "finance_signature_code": "FIN_TEST",
        "md_signature_code": "MD_TEST",
        "output_filename": "salary-advance-test",
    }
    return GenerationSnapshot(
        normalized_data=normalized,
        finance_signature_path=finance,
        md_signature_path=managing_director,
        finance_signature_version={"assetVersion": 1},
        md_signature_version={"assetVersion": 1},
    )


def test_template_triple_hashes_match_checked_in_assets() -> None:
    assets = Path(__file__).parents[1] / "app" / "assets" / "templates"

    assert sha256_path(assets / "Salary-Advance-Template.xlsx") == (
        pdf_layout.SOURCE_XLSX_SHA256
    )
    assert sha256_path(assets / "Salary-Advance-Template.pdf") == (
        pdf_layout.PDF_UNDERLAY_SHA256
    )
    assert pdf_layout.PAGE_COUNT == 1


def test_overlay_font_ships_with_repo() -> None:
    """叠加层字体必须随仓库发布，不能依赖机器上装没装。

    底版是 Excel 从模板导出的，表里的标签用的是 TH SarabunPSK；叠加层若换了
    别的字体，同字号下填入的字会和表格明显不是一套。这里钉住两件事：配置指向的
    字体文件确实签入在 assets/fonts 下，且 reportlab 能按 FORM_FONT_NAME 注册它。
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = get_settings().salary_advance_font_path
    assert font_path.is_file(), font_path
    fonts_dir = Path(__file__).parents[1] / "app" / "assets" / "fonts"
    assert font_path.resolve().parent == fonts_dir.resolve()

    # 换了文件却复用旧注册名，reportlab 会继续吐旧字体——这里真注册一次兜底。
    if FORM_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(FORM_FONT_NAME, str(font_path)))
    assert FORM_FONT_NAME in pdfmetrics.getRegisteredFontNames()


def test_renders_approved_xlsx_without_applicant_signature(tmp_path: Path) -> None:
    template = (
        Path(__file__).parents[1]
        / "app"
        / "assets"
        / "templates"
        / "Salary-Advance-Template.xlsx"
    )
    rendered = load_workbook(
        BytesIO(render_salary_advance_workbook(template, _snapshot(tmp_path)))
    )

    assert rendered.sheetnames == ["表单模板", "当前记录"]
    assert rendered["当前记录"].sheet_state == "hidden"
    assert rendered["当前记录"]["B2"].value == "202607"
    assert rendered["当前记录"]["B3"].value == "E001"
    assert rendered["当前记录"]["B13"].value == 12500.5
    for row in range(26, 29):
        for column in range(8, 11):
            assert rendered["表单模板"].cell(row=row, column=column).value is None
    assert len(rendered["表单模板"]._images) >= 3
    rendered.close()


def _stamped_rects(pdf_path: Path) -> list[tuple[float, float, float, float]]:
    """量出这一页上每张贴图的实际矩形 (x, y, 宽, 高)，单位 PDF 点。

    量成品而不是复算一遍公式：复算只能证明"我抄对了自己写的算式"。
    底版右上角本身有一张 logo 贴图，按 y 排除掉。
    """
    import pdfplumber

    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        return sorted(
            (
                (image["x0"], image["y0"], image["x1"] - image["x0"], image["y1"] - image["y0"])
                for image in page.images
                if image["y0"] < 500  # 页眉 logo 在 y≈734，两个签名位都在 400 以下
            ),
            key=lambda rect: -rect[1],
        )


def test_signature_with_margins_still_fills_the_box(tmp_path: Path) -> None:
    """导出的签名 PNG 四周基本都有留白，留白不能占掉签名框。

    drawImage 是把**整张图**等比适配进签名框，所以图里的留白会照比例吃掉框。
    工资预支原先不做裁剪，一张笔画只占 61% 宽的签名印出来只有 34pt（框宽 90pt，
    38%），用户反馈"签名过小"。现在与 WHT / TAX INV 共用 core.signature_image，
    先裁到墨迹外接框再套印。

    钉的是"撑满"这个结果而不是"调了某个函数"：换实现只要签名还是该多大就多大，
    这条就该继续绿。
    """
    assets = Path(__file__).parents[1] / "app" / "assets"
    source = tmp_path / "margined.png"
    canvas = Image.new("RGBA", (600, 300), (255, 255, 255, 0))
    draw = ImageDraw.Draw(canvas)
    # 笔画只占画布中间：横向约 61%、纵向约 40%，四周都是留白。
    draw.line((120, 200, 190, 90, 265, 195, 340, 85, 420, 190, 480, 130), fill="black", width=9)
    canvas.save(source)

    output = tmp_path / "margined-signature.pdf"
    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        dataclasses.replace(
            _snapshot(tmp_path), finance_signature_path=source, md_signature_path=source
        ),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets / "templates" / "Salary-Advance-Template.xlsx",
    )

    for rect, role in zip(_stamped_rects(output), ("finance", "md"), strict=True):
        box_w, box_h = pdf_layout.SIGNATURE_BOXES[role][2:]
        width, height = rect[2], rect[3]
        # 等比适配后至少有一个方向要贴到框边；留白没裁掉的话两个方向都远不到。
        assert max(width / box_w, height / box_h) == pytest.approx(1.0, abs=0.01), (
            f"{role} 没有撑满签名框：{width:.1f}x{height:.1f}pt / 框 {box_w}x{box_h}pt。"
            "签名图四周的留白多半没被裁掉。"
        )
        assert width > box_w * 0.8, (
            f"{role} 只印了 {width:.1f}pt，框宽 {box_w}pt——这就是「签名过小」那个回归。"
        )


def test_workbook_signature_survives_the_temporary_prepared_file(tmp_path: Path) -> None:
    """xlsx 侧同样要用裁过的图，且不能踩到临时文件的生命周期。

    openpyxl 到 workbook.save() 才真正去读图片；套印用的是"用完即删"的临时文件，
    把路径交给它就会在保存时扑空。这条守的是"能出得来"，删掉那层内存读取会红。
    """
    assets = Path(__file__).parents[1] / "app" / "assets"
    source = tmp_path / "margined-xlsx.png"
    _signature(source, (20, 35, 120, 255))

    data = render_salary_advance_workbook(
        assets / "templates" / "Salary-Advance-Template.xlsx",
        dataclasses.replace(
            _snapshot(tmp_path), finance_signature_path=source, md_signature_path=source
        ),
    )

    rendered = load_workbook(BytesIO(data))
    try:
        assert len(rendered["表单模板"]._images) >= 3
    finally:
        rendered.close()


@pytest.mark.parametrize("scale", [60, 100, 150])
def test_signature_scale_shrinks_around_the_box_centre(tmp_path: Path, scale: int) -> None:
    """签名库维护页存的比例必须真的改变工资预支单上的盖章尺寸，且围绕框中心缩放。

    两点各钉一条：
    1. **比例真的生效**——加这个参数之前 _draw_signature 根本没有 scale 形参，
       维护页照样提示"已应用至系统开票"，出票尺寸一动不动。
    2. **缩放围绕签名框中心**——与 wht / tax_invoice 两处 drawImage 前的算式
       逐字一致，也与预览的 stampRectPt 一致。若改成钉住左下角，滑到 60% 时
       预览与出票会差半个框，而用户正是照着预览在调。
    """
    assets = Path(__file__).parents[1] / "app" / "assets"
    output = tmp_path / f"salary-advance-{scale}.pdf"

    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        dataclasses.replace(
            _snapshot(tmp_path), finance_scale_percent=scale, md_scale_percent=scale
        ),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets / "templates" / "Salary-Advance-Template.xlsx",
    )

    rects = _stamped_rects(output)
    assert len(rects) == 2, f"应当只有财务与董事两个签名位，实测 {rects}"
    for rect, role in zip(rects, ("finance", "md"), strict=True):
        box_x, box_y, box_w, box_h = pdf_layout.SIGNATURE_BOXES[role]
        x, y, width, height = rect
        # 等比适配后总有一个方向撑满缩放后的框，另一个方向留白并居中。
        assert width <= box_w * scale / 100 + 0.01, role
        assert height <= box_h * scale / 100 + 0.01, role
        assert max(width / (box_w * scale / 100), height / (box_h * scale / 100)) == pytest.approx(
            1.0, abs=0.01
        ), f"{role} 没有撑满缩放后的签名框：{rect}"
        # 中心不动才是"围绕框中心缩放"。
        assert x + width / 2 == pytest.approx(box_x + box_w / 2, abs=0.01), role
        assert y + height / 2 == pytest.approx(box_y + box_h / 2, abs=0.01), role


def test_signature_scale_is_independent_per_position(tmp_path: Path) -> None:
    """财务与董事常是两个人的章，各调各的尺寸；合成一个值就是让一个跟着另一个走。"""
    assets = Path(__file__).parents[1] / "app" / "assets"
    output = tmp_path / "salary-advance-mixed.pdf"

    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        dataclasses.replace(
            _snapshot(tmp_path), finance_scale_percent=60, md_scale_percent=150
        ),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets / "templates" / "Salary-Advance-Template.xlsx",
    )

    finance_rect, md_rect = _stamped_rects(output)
    assert md_rect[2] > finance_rect[2] * 2, (
        f"两个位的比例没有各自生效：财务 {finance_rect}，董事 {md_rect}"
    )


def test_generates_single_page_pdf_without_office(tmp_path: Path) -> None:
    assets = Path(__file__).parents[1] / "app" / "assets"
    output = tmp_path / "salary-advance.pdf"

    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        _snapshot(tmp_path),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets
        / "templates"
        / "Salary-Advance-Template.xlsx",
    )

    reader = PdfReader(output)
    assert len(reader.pages) == 1
    text = reader.pages[0].extract_text() or ""
    assert "SOMCHAI" in text
    assert "12,500.50" in text
    assert "{{finance_signature}}" not in text
    assert "{{md_signature}}" not in text
    assert output.stat().st_size > 100_000


def _position_lines(pdf: Path) -> list[tuple[float, float, float, str]]:
    """把职位区（第 8-9 行右侧）的文字按行取出来：(基线, 左, 右, 文字)。

    左边界卡在数据区左沿，否则会把底版上 "ตำแหน่ง Position" 标签一起框进来；
    右边不设界，这样文字真溢出格子还测得出来（居中排版右溢即左溢）。
    """
    import pdfplumber

    anchor = pdf_layout.TEXT_ANCHORS["position"]
    left = anchor.x - anchor.max_width / 2
    with pdfplumber.open(pdf) as document:
        page = document.pages[0]
        height = float(page.height)
        rows: dict[float, list] = {}
        for word in page.extract_words():
            baseline = round(height - float(word["bottom"]), 1)
            if float(word["x0"]) >= left and 600.0 <= baseline <= 650.0:
                rows.setdefault(baseline, []).append(word)
    lines = []
    for baseline in sorted(rows, reverse=True):
        words = sorted(rows[baseline], key=lambda item: float(item["x0"]))
        lines.append(
            (
                baseline,
                float(words[0]["x0"]),
                float(words[-1]["x1"]),
                " ".join(word["text"] for word in words),
            )
        )
    return lines


@pytest.mark.parametrize(
    "position",
    [
        "MANUFACTURING TECHNOLOGY SUPERVISOR 2",
        # 导入数据里真实存在的漏空格写法，没有可用断点，只能按字符硬断
        "MANUFACTURING TECHNOLOGYSUPERVISOR 1",
    ],
)
def test_long_position_is_printed_in_full(tmp_path: Path, position: str) -> None:
    """长职位以前会被压到 7pt 再逐字符砍掉，现在必须折行印全。"""
    assets = Path(__file__).parents[1] / "app" / "assets"
    output = tmp_path / "long-position.pdf"

    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        _snapshot(tmp_path, position=position),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets / "templates" / "Salary-Advance-Template.xlsx",
    )

    lines = _position_lines(output)
    assert "".join(text for _, _, _, text in lines).replace(" ", "") == (
        position.replace(" ", "")
    )
    assert len(lines) <= WRAPPED_FIELDS["position"]


def test_position_block_stays_inside_its_box(tmp_path: Path) -> None:
    """折行数和模板里 K8:L9 的高度必须对得上。

    行高够不够放不下没人会报错，只会静静印到格子外面去。这里钉住两条边界：
    横向不越出制版量到的可绘宽度，纵向不高过同一行 Name/Surname 的基线——
    再往上就是第 7 行的留白，说明第 9 行行高和 WRAPPED_FIELDS 脱节了。
    """
    assets = Path(__file__).parents[1] / "app" / "assets"
    output = tmp_path / "boxed-position.pdf"

    export_pdf_from_template(
        assets / "templates" / "Salary-Advance-Template.pdf",
        output,
        _snapshot(tmp_path, position="MANUFACTURING TECHNOLOGY SUPERVISOR 2"),
        get_settings().salary_advance_font_path,
        xlsx_template_path=assets / "templates" / "Salary-Advance-Template.xlsx",
    )

    anchor = pdf_layout.TEXT_ANCHORS["position"]
    lines = _position_lines(output)
    assert len(lines) > 1, "样例职位应当折行，否则这个测试测不到东西"
    for baseline, left, right, text in lines:
        assert right - left <= anchor.max_width, text
        assert right <= anchor.x + anchor.max_width / 2 + 1, text
        assert baseline <= pdf_layout.TEXT_ANCHORS["first_name"].y, text
