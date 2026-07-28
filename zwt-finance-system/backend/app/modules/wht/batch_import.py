"""批量开具导入：一张 Excel 一次建多条 WHT 草稿。

与 legacy_import 的区别（两个功能长得像，混了会出事）：

    legacy_import  迁移旧系统**已经开过**的票 → 表里必须有 RefNo 正式编号，
                   落库直接是 issued，并把编号计数器推到该序号之后。
    batch_import   批量录入**还没开**的票 → 表里不能有编号，落库是 draft，
                   编号照旧只在批准时由服务器事务生成。

解析策略是**全表校验、有错全退**：草稿没有唯一编号，一旦"成功 8 条失败 2 条"，
用户修好那 2 行再传一次，前 8 条就会重复建一遍且无法自动去重。
所以宁可整张表退回并一次性列出所有行的问题，让用户改完再传。
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

from app.modules.wht.income_types import default_rate, find
from app.modules.wht.workbook import (
    WorkbookError,
    date_value,
    decimal_value,
    load,
    text,
)
from app.modules.wht.workbook import (
    tax_id as normalize_tax_id,
)

# 模板列名。前 4 列必填，后 4 列可留空由系统推导。
REQUIRED_COLUMNS = ("PayeeTaxID", "IncomeType", "PaymentDate", "Amount")
OPTIONAL_COLUMNS = ("Period", "IssuanceType", "SupplementRun", "TaxRate")
TEMPLATE_COLUMNS = (*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS)

MAX_ROWS = 500


class BatchWorkbookError(WorkbookError):
    """批量开具工作簿不合格。errors 里是逐行问题，前端整张表退回时全部展示。"""

    def __init__(self, message: str, errors: list[str] | None = None) -> None:
        super().__init__(message)
        self.errors = errors or []


def _rate(value: Any, row_number: int) -> Decimal | None:
    """税率按小数填写（0.03 = 3%），也接受 "3%" 这种带百分号的文本。

    不接受裸的 3 —— 3 既可能是 3% 也可能是 300%，历史台账用的是 0.03，
    这里靠拒绝歧义值把错误挡在导入之前，而不是靠猜。
    """
    raw = text(value)
    if not raw:
        return None
    if raw.endswith("%"):
        parsed = decimal_value(raw[:-1]) / Decimal("100")
    else:
        parsed = decimal_value(raw)
        if parsed > 1:
            raise BatchWorkbookError(
                f"row {row_number}: TaxRate must be a decimal such as 0.03, "
                f'or written with a percent sign such as "3%"'
            )
    if not Decimal(0) < parsed <= Decimal(1):
        raise BatchWorkbookError(f"row {row_number}: TaxRate must be greater than 0 and at most 1")
    return parsed


def parse_batch_sheet(content: bytes) -> list[dict[str, Any]]:
    """把批量开具工作簿解析成 create_task 用的行。任何一行有问题就整张表退回。"""
    workbook = load(content)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise BatchWorkbookError("the first worksheet is empty")
        columns = {text(value): index for index, value in enumerate(header) if text(value)}

        if "RefNo" in columns:
            raise BatchWorkbookError(
                "this sheet contains a RefNo column, so it looks like an already-issued "
                "historical ledger. Use the historical-migration import instead; batch "
                "issuance creates drafts and the server assigns numbers on approval"
            )
        missing = [name for name in REQUIRED_COLUMNS if name not in columns]
        if missing:
            raise BatchWorkbookError(
                f"the first worksheet is missing columns: {', '.join(missing)}"
            )

        def cell(row: tuple[Any, ...], name: str) -> Any:
            index = columns.get(name)
            return row[index] if index is not None and index < len(row) else None

        records: list[dict[str, Any]] = []
        errors: list[str] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            if len(records) >= MAX_ROWS:
                raise BatchWorkbookError(
                    f"a single batch is limited to {MAX_ROWS} rows; split the file"
                )
            try:
                records.append(_parse_row(cell, row, row_number))
            except WorkbookError as exc:
                errors.append(str(exc))

        if errors:
            raise BatchWorkbookError(
                f"{len(errors)} row(s) need fixing; nothing was imported",
                errors=errors,
            )
        if not records:
            raise BatchWorkbookError("the first worksheet has no data rows")
        return records
    finally:
        workbook.close()


def _parse_row(cell: Any, row: tuple[Any, ...], row_number: int) -> dict[str, Any]:
    payee_tax_id = normalize_tax_id(cell(row, "PayeeTaxID"))
    if len(payee_tax_id) != 13:
        raise BatchWorkbookError(f"row {row_number}: PayeeTaxID must be 13 digits")

    income_type = text(cell(row, "IncomeType"))
    if not income_type:
        raise BatchWorkbookError(f"row {row_number}: IncomeType is required")
    # 目录里能查到就用泰文原文落库，保证同一类型在台账里写法一致。
    option = find(income_type)
    if option is not None:
        income_type = option.label_th

    try:
        payment_date = date_value(cell(row, "PaymentDate"))
    except WorkbookError as exc:
        raise BatchWorkbookError(f"row {row_number}: {exc}") from exc
    if payment_date is None:
        raise BatchWorkbookError(f"row {row_number}: PaymentDate is required")

    try:
        total_amount = decimal_value(cell(row, "Amount"))
    except WorkbookError as exc:
        raise BatchWorkbookError(f"row {row_number}: {exc}") from exc
    if total_amount <= 0:
        raise BatchWorkbookError(f"row {row_number}: Amount must be greater than 0")

    period = text(cell(row, "Period")) or payment_date.strftime("%Y-%m")
    if len(period) != 7 or period[4] != "-" or not period.replace("-", "").isdigit():
        raise BatchWorkbookError(f"row {row_number}: Period must look like 2026-06")

    issuance_type = (text(cell(row, "IssuanceType")) or "normal").lower()
    if issuance_type not in {"normal", "supplement"}:
        raise BatchWorkbookError(f"row {row_number}: IssuanceType must be normal or supplement")
    supplement_run = int(decimal_value(cell(row, "SupplementRun")))
    if issuance_type == "normal":
        supplement_run = 0
    elif not 1 <= supplement_run <= 9:
        raise BatchWorkbookError(
            f"row {row_number}: SupplementRun must be 1-9 for supplement issuance"
        )

    return {
        "row_number": row_number,
        "payee_tax_id": payee_tax_id,
        "income_type": income_type,
        "payment_date": payment_date,
        "total_amount": total_amount,
        "period": period,
        "issuance_type": issuance_type,
        "supplement_run": supplement_run,
        "wht_rate": _rate(cell(row, "TaxRate"), row_number),
    }


def resolve_rate(row: dict[str, Any], wht_type: str | None) -> Decimal | None:
    """表里没填税率时，按收入类型 + 收款方的 PND 类型带出法定默认值。"""
    if row.get("wht_rate") is not None:
        return row["wht_rate"]
    return default_rate(row["income_type"], wht_type)  # type: ignore[arg-type]


TEMPLATE_NOTES = (
    ("PayeeTaxID", "必填。收款方 13 位税号，必须已存在于「收款方主数据」。"),
    ("IncomeType", "必填。泰文收入类型，例如 ค่าบริการ / ค่าขนส่ง / ค่าเช่า。"),
    ("PaymentDate", "必填。日期格式 YYYY-MM-DD，不接受 DD/MM/YYYY。"),
    ("Amount", "必填。付款总额，大于 0，不要带千分位以外的符号。"),
    ("Period", "可留空。留空时按 PaymentDate 推导，例如 2026-06。"),
    ("IssuanceType", "可留空。normal（默认）或 supplement。"),
    ("SupplementRun", "补开时必填 1-9；normal 请留空或填 0。"),
    ("TaxRate", '可留空。留空按收入类型带出法定税率。填写请用小数 0.03 或 "3%"。'),
    ("—", "本表只建草稿，不能填 RefNo：正式编号仍在批准时由服务器生成。"),
)

EXAMPLE_ROW = (
    "0105540057561",
    "ค่าบริการ",
    "2026-06-05",
    3000,
    "2026-06",
    "normal",
    None,
    None,
)


def build_template_workbook() -> bytes:
    """生成批量开具模板：第一张表是列头 + 一行示例，第二张表是填写说明。"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BatchIssue"
    sheet.append(list(TEMPLATE_COLUMNS))
    sheet.append(list(EXAMPLE_ROW))
    for index, name in enumerate(TEMPLATE_COLUMNS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(
            14, len(name) + 4
        )

    notes = workbook.create_sheet("说明 Notes")
    notes.append(["列名 Column", "说明 Note"])
    for name, note in TEMPLATE_NOTES:
        notes.append([name, note])
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 76

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
