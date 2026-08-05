from __future__ import annotations

import re
from collections.abc import Mapping
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.signature_image import SignatureImageError
from app.core.signature_image import build_blue_signature as prepare_signature_stamp
from app.modules.wht.branching import display_payee_name
from app.modules.wht.models import WhtTask

CHECKBOX_OFFSET = -1
CHECKBOX_COLUMN_OFFSETS = (0, 18, 36, 54)
PDF_CHECKBOX_CENTERS = {
    "PND3": (195.5, 584.25),
    "PND53": (372.0, 584.25),
}
PDF_SIGNATURE_BOX = (201.5, 83.0, 95.0, 42.0)
TEMPLATE_FIELDS = {
    "No",
    "RefNo",
    "BookNo",
    "PayeeNameEN",
    "PayeeNameTH",
    "PayeeAddress",
    "PayeeTaxID",
    "WHT Type3",
    "WHT Type53",
    "IncomeType",
    "PaymentDate",
    "Amount",
    "TaxRate",
    "TaxAmount",
    "AmountTextThai",
    "DateTextThai",
}
THAI_DIGITS = (
    "ศูนย์",
    "หนึ่ง",
    "สอง",
    "สาม",
    "สี่",
    "ห้า",
    "หก",
    "เจ็ด",
    "แปด",
    "เก้า",
)
THAI_POSITIONS = ("", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน")
THAI_MONTHS = (
    "",
    "มกราคม",
    "กุมภาพันธ์",
    "มีนาคม",
    "เมษายน",
    "พฤษภาคม",
    "มิถุนายน",
    "กรกฎาคม",
    "สิงหาคม",
    "กันยายน",
    "ตุลาคม",
    "พฤศจิกายน",
    "ธันวาคม",
)


class DocumentGenerationError(RuntimeError):
    pass


def _thai_under_million(number: int) -> str:
    if not 0 <= number < 1_000_000:
        raise ValueError("number must be between 0 and 999999")
    if number == 0:
        return ""
    digits = str(number)
    result: list[str] = []
    for index, character in enumerate(digits):
        digit = int(character)
        if digit == 0:
            continue
        position = len(digits) - index - 1
        if position == 1 and digit == 1:
            word = ""
        elif position == 1 and digit == 2:
            word = "ยี่"
        elif position == 0 and digit == 1 and number > 10:
            word = "เอ็ด"
        else:
            word = THAI_DIGITS[digit]
        result.append(f"{word}{THAI_POSITIONS[position]}")
    return "".join(result)


def thai_integer_text(number: int) -> str:
    if number < 0:
        raise ValueError("number must not be negative")
    if number == 0:
        return THAI_DIGITS[0]
    millions, remainder = divmod(number, 1_000_000)
    result = ""
    if millions:
        result = f"{thai_integer_text(millions)}ล้าน"
    if remainder:
        result += "เอ็ด" if millions and remainder == 1 else _thai_under_million(remainder)
    return result


def thai_baht_text(amount: Decimal) -> str:
    rounded = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if rounded < 0:
        raise ValueError("amount must not be negative")
    baht = int(rounded)
    satang = int((rounded - Decimal(baht)) * 100)
    text = f"{thai_integer_text(baht)}บาท"
    if satang:
        return f"{text}{thai_integer_text(satang)}สตางค์"
    return f"{text}ถ้วน"


def thai_date_text(value: date) -> str:
    return f"{value.day} {THAI_MONTHS[value.month]} {value.year + 543}"


def _task_sequence(task_no: str) -> str:
    supplement = re.search(r"BK\d(\d{2})$", task_no)
    if supplement:
        return str(int(supplement.group(1)))
    normal = re.search(r"(\d{1,3})$", task_no)
    return str(int(normal.group(1))) if normal else ""


def task_template_values(task: WhtTask) -> dict[str, Any]:
    if not task.task_no or not task.book_no or not task.payment_date:
        raise DocumentGenerationError(
            "approved task number, book number and payment date are required"
        )
    amount_text = task.amount_text_thai or f"-- {thai_baht_text(task.wht_amount)} --"
    date_text = task.date_text_thai or thai_date_text(task.payment_date)
    return {
        "No": _task_sequence(task.task_no),
        "RefNo": task.task_no,
        "BookNo": task.book_no,
        "PayerName": "บริษัท นำคนต่างด้าวมาทำงานในประเทศ ซีดับเบิ้ลยูที จำกัด",
        "PayerTaxId": "0105562001234",
        "PayerAddress": "เลขที่ 414 ถนนพหลโยธิน แขวงสามเสนใน เขตพญาไท กรุงเทพมหานคร 10400",
        "PayeeNameEN": task.company_name_en or "",
        "PayeeNameTH": display_payee_name(
            task.company_name,
            getattr(task, "branch_type", None),
            getattr(task, "branch_number", None),
        ),
        "PayeeAddress": task.payee_address or "",
        "PayeeTaxID": task.tax_id or "",
        "WHT Type3": "√" if task.wht_type == "PND3" else "",
        "WHT Type53": "√" if task.wht_type == "PND53" else "",
        "IncomeType": task.income_type or "",
        "PaymentDate": task.payment_date,
        "Amount": task.total_amount,
        "TaxRate": task.wht_rate or Decimal("0"),
        "TaxAmount": task.wht_amount,
        "AmountTextThai": amount_text,
        "DateTextThai": date_text,
    }


def render_wht_workbook(template_path: Path, task: WhtTask) -> bytes:
    if not template_path.is_file():
        raise DocumentGenerationError(f"WHT template was not found: {template_path}")
    values = task_template_values(task)
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    placeholders: dict[str, list[tuple[int, int]]] = {}
    backup_coords: dict[str, tuple[int, int]] = {}

    for row in worksheet.iter_rows():
        for cell in row:
            text = cell.value.strip() if isinstance(cell.value, str) else ""
            if text in values:
                placeholders.setdefault(text, []).append((cell.row, cell.column))
            if "ภ.ง.ด. 3" in text and "3ก" not in text:
                backup_coords["PND3"] = (cell.row, cell.column + CHECKBOX_OFFSET)
            if "ภ.ง.ด. 53" in text:
                backup_coords["PND53"] = (cell.row, cell.column + CHECKBOX_OFFSET)

    for field, coordinates in placeholders.items():
        value = values[field]
        for row, column in coordinates:
            cell = worksheet.cell(row=row, column=column)
            cell.value = value
            if field == "TaxRate":
                cell.number_format = "0%"
            elif field in {"Amount", "TaxAmount"}:
                cell.number_format = "#,##0.00"

    for form_type, field in (("PND3", "WHT Type3"), ("PND53", "WHT Type53")):
        if form_type in backup_coords and field not in placeholders:
            row, column = backup_coords[form_type]
            checked = "√" if task.wht_type == form_type else ""
            for offset in CHECKBOX_COLUMN_OFFSETS:
                worksheet.cell(row=row, column=column + offset).value = checked

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def render_clean_template_workbook(template_path: Path) -> bytes:
    """Remove data placeholders before an administrator exports the PDF underlay."""
    if not template_path.is_file():
        raise DocumentGenerationError(f"WHT template was not found: {template_path}")
    workbook = load_workbook(template_path)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                text = cell.value.strip() if isinstance(cell.value, str) else ""
                if text in TEMPLATE_FIELDS:
                    cell.value = None
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def validate_signature_image(content: bytes) -> tuple[str, str]:
    try:
        from PIL import Image

        with Image.open(BytesIO(content)) as image:
            image.verify()
            image_format = (image.format or "").upper()
    except Exception as exc:
        raise DocumentGenerationError("signature file is not a valid image") from exc
    formats = {
        "PNG": ("image/png", ".png"),
        "JPEG": ("image/jpeg", ".jpg"),
    }
    if image_format not in formats:
        raise DocumentGenerationError("signature image must be PNG or JPEG")
    return formats[image_format]


def build_blue_signature(source_path: Path, target_path: Path) -> None:
    """三种单据共用的签名处理，实现在 app.core.signature_image。

    这里保留同名薄封装：TAX INV 与既有测试都从这里 import，且调用方期待的是
    本模块的 DocumentGenerationError。
    """
    try:
        prepare_signature_stamp(source_path, target_path)
    except SignatureImageError as exc:
        raise DocumentGenerationError(str(exc)) from exc


def export_pdf_from_template(
    pdf_template_path: Path,
    pdf_path: Path,
    task: WhtTask,
    thai_font_path: Path,
    signature_source_path: Path | None,
    signature_scale_percent: int = 100,
) -> None:
    """Create a four-copy WHT PDF without automating Microsoft Office."""
    if not pdf_template_path.is_file():
        raise DocumentGenerationError(f"WHT PDF template was not found: {pdf_template_path}")
    if not thai_font_path.is_file():
        raise DocumentGenerationError(f"Thai PDF font was not found: {thai_font_path}")

    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.lib.colors import black
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise DocumentGenerationError(
            "PDF generation requires reportlab and pypdf"
        ) from exc

    values = task_template_values(task)
    pdfmetrics.registerFont(TTFont("ZwtSarabun", str(thai_font_path)))
    reader = PdfReader(str(pdf_template_path))
    writer = PdfWriter()
    prepared_signature = pdf_path.with_suffix(".signature.png")
    if signature_source_path:
        build_blue_signature(signature_source_path, prepared_signature)

    for page in reader.pages:
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        overlay_buffer = BytesIO()
        overlay = canvas.Canvas(overlay_buffer, pagesize=(width, height))
        overlay.setFillColor(black)

        def draw_text(
            text: object,
            x: float,
            y: float,
            *,
            font: str = "Helvetica",
            size: float = 8,
            _canvas: Any = overlay,
        ) -> None:
            _canvas.setFont(font, size)
            _canvas.drawString(x, y, str(text))

        def draw_right(
            text: object,
            x: float,
            y: float,
            *,
            size: float = 8,
            _canvas: Any = overlay,
        ) -> None:
            _canvas.setFont("Helvetica", size)
            _canvas.drawRightString(x, y, str(text))

        def draw_centered_thai(
            text: object,
            x: float,
            y: float,
            *,
            size: float = 10,
            _canvas: Any = overlay,
        ) -> None:
            _canvas.setFont("ZwtSarabun", size)
            _canvas.drawCentredString(x, y, str(text))

        def draw_centered(
            text: object,
            x: float,
            y: float,
            *,
            size: float = 8,
            _canvas: Any = overlay,
        ) -> None:
            _canvas.setFont("Helvetica", size)
            _canvas.drawCentredString(x, y, str(text))

        # 付款方 (Payer) 100% 保持 WHT-Template.pdf 模板原版预印付款公司 ('บริษัท เจินเวสเทิร์น เทคโนโลยี (ไทยแลนด์) จำกัด(สำนักงานใหญ่)', 税号 '0 1055 66051 02 1') 完整无损，不做白块擦除也不叠加额外公司名称

        # 右对齐到页框内侧，给泰文标签留出固定间距；三位流水号不会再压住标签。
        draw_right(values["BookNo"], 570, 799)
        draw_right(values["RefNo"], 570, 785)
        # 收款方 (Payee)：
        # 1. 字体字号保持与付款方完全一致 (ZwtSarabun, 9pt)
        # 2. 收款方税号 (PayeeTaxID) 字号为 10pt，基线 y=667.0pt 精确对齐左侧“เลขประจำตัวผู้เสียภาษีอากร”泰语标签基线
        draw_text(values["PayeeNameTH"], 37, 660, font="ZwtSarabun", size=9)
        draw_text(values["PayeeTaxID"], 482, 667.0, size=10)
        draw_text(values["PayeeAddress"], 37, 628, font="ZwtSarabun", size=9)
        # Bug 1 修复：模板中「ลำดับที่ [   ] ในแบบ」框内部按合规要求留空，绝不打印序号（这里不用填写）

        # PND3 / PND53 下排复选框中心点来自模板矢量坐标。勾号四周
        # 至少保留约 2pt，不再压住左边框或下边框。
        check_x, check_y = PDF_CHECKBOX_CENTERS[task.wht_type]
        overlay.setLineWidth(1.2)
        overlay.setLineCap(1)
        overlay.line(check_x - 5, check_y - 1.25, check_x - 1.5, check_y - 4.75)
        overlay.line(check_x - 1.5, check_y - 4.75, check_x + 5, check_y + 4.75)

        # Bug 4 彻底修复：精准缩窄遮盖框宽度至 60pt (x=440..500 与 x=515..575)，距右边框 (579.5) 留出 4.5pt 安全间距，
        # 既完全擦除 2-4 页第 5 行预印减号 '-'，又绝对不破损右侧竖直边框线！
        overlay.setFillColorRGB(1, 1, 1)
        overlay.rect(440, 260, 60, 13, fill=1, stroke=0)
        overlay.rect(515, 260, 60, 13, fill=1, stroke=0)
        overlay.setFillColorRGB(0, 0, 0)

        # 第 6 行 (6. อื่นๆ(ระบุ)) 完完全全与图二一致：x=181.0 居中，基线设为 y=252.5pt 优雅悬浮于虚线上方 3pt 处，预印虚线在下方全长贯穿
        draw_text(values["IncomeType"], 181.0, 252.5, font="ZwtSarabun", size=9)
        payment_date = task.payment_date
        if payment_date:
            draw_text(
                f"{payment_date.year}/{payment_date.month}/{payment_date.day}",
                368,
                251.5,
            )
        amount = f"{task.total_amount:,.2f}"
        tax_amount = f"{task.wht_amount:,.2f}"
        draw_right(amount, 499, 251.5)
        draw_right(tax_amount, 571, 251.5)
        # 汇总行上下边界为 y=222.5..238.6；使用居中基线避开边框。
        draw_right(amount, 499, 227)
        draw_right(tax_amount, 571, 227)
        draw_centered_thai(values["AmountTextThai"], 364, 198, size=10)
        draw_centered_thai(values["DateTextThai"], 233, 78, size=10)

        if prepared_signature.exists():
            image = ImageReader(str(prepared_signature))
            signature_x, signature_y, signature_width, signature_height = PDF_SIGNATURE_BOX
            scale_ratio = (signature_scale_percent or 100) / 100.0
            scaled_w = signature_width * scale_ratio
            scaled_h = signature_height * scale_ratio
            scaled_x = signature_x + (signature_width - scaled_w) / 2.0
            scaled_y = signature_y + (signature_height - scaled_h) / 2.0
            overlay.drawImage(
                image,
                scaled_x,
                scaled_y,
                width=scaled_w,
                height=scaled_h,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )

        overlay.save()
        overlay_buffer.seek(0)
        overlay_page = PdfReader(overlay_buffer).pages[0]
        writer.add_page(page)
        writer.pages[-1].merge_page(overlay_page)

    with pdf_path.open("wb") as output:
        writer.write(output)
    if prepared_signature.exists():
        prepared_signature.unlink()


def document_file_stem(task: WhtTask) -> str:
    if not task.task_no:
        raise DocumentGenerationError("task number has not been allocated")
    payment = task.payment_date.strftime("%Y%m%d") if task.payment_date else "NoDate"
    tax_amount = f"{task.wht_amount:,.2f}"
    raw = f"{task.task_no}-{payment}-THB-{tax_amount}"
    return re.sub(r'[\\/:*?"<>|]', "", raw)


def signature_path(root: Path, storage_key: str) -> Path:
    resolved_root = root.resolve()
    resolved = (resolved_root / Path(storage_key)).resolve()
    try:
        resolved.relative_to(resolved_root)
    except ValueError as exc:
        raise DocumentGenerationError("stored signature path escaped the attachment root") from exc
    return resolved


def merge_template_values(
    base: Mapping[str, Any],
    overrides: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Small test helper used by future module-specific template adapters."""
    return {**base, **(overrides or {})}
