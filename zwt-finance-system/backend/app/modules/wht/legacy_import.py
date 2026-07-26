import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.modules.wht.numbering import compact_period

NORMAL_NUMBER_PATTERN = re.compile(r"^ZWT(?P<period>\d{6})(?P<sequence>\d{1,3})$")
SUPPLEMENT_NUMBER_PATTERN = re.compile(r"^ZWT(?P<period>\d{6})BK(?P<run>[1-9])(?P<sequence>\d{2})$")


class LegacyWorkbookError(ValueError):
    pass


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _tax_id(value: Any) -> str:
    tax_id = re.sub(r"\D", "", _text(value))
    if tax_id and len(tax_id) < 13:
        tax_id = tax_id.zfill(13)
    return tax_id


def _aliases(value: Any) -> list[str]:
    raw = _text(value)
    if not raw:
        return []
    return list(dict.fromkeys(part.strip() for part in re.split(r"[/\n\r]+", raw) if part.strip()))


def _decimal(value: Any, default: str = "0") -> Decimal:
    raw = _text(value).replace(",", "")
    if not raw:
        return Decimal(default)
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise LegacyWorkbookError(f"invalid decimal value: {value}") from exc


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value).replace("/", "-")
    for date_format in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    raise LegacyWorkbookError(f"invalid date value: {value}")


def _load(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise LegacyWorkbookError("the uploaded file is not a readable XLSX workbook") from exc


def parse_payee_sheet(content: bytes) -> list[dict[str, Any]]:
    workbook = _load(content)
    try:
        if len(workbook.worksheets) < 2:
            raise LegacyWorkbookError("the workbook must contain a second payee data sheet")
        rows = workbook.worksheets[1].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise LegacyWorkbookError("the second worksheet is empty")
        normalized = [_text(value).upper() for value in header]
        expected = ["NO", "TAXID", "NAME", "ADDRESS", "TYPE"]
        if normalized[:5] != expected:
            raise LegacyWorkbookError(
                "the second worksheet must begin with NO, TAXID, Name, Address, Type"
            )

        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            tax_id = _tax_id(row[1] if len(row) > 1 else None)
            name_th = _text(row[2] if len(row) > 2 else None)
            address_th = _text(row[3] if len(row) > 3 else None)
            wht_type = re.sub(r"[\s.]+", "", _text(row[4] if len(row) > 4 else None)).upper()
            if not tax_id or not name_th or not address_th or wht_type not in {"PND3", "PND53"}:
                raise LegacyWorkbookError(
                    f"invalid payee data at second worksheet row {row_number}"
                )
            records.append(
                {
                    "tax_id": tax_id,
                    "name_th": name_th,
                    "address_th": address_th,
                    "wht_type": wht_type,
                    "aliases": _aliases(row[5] if len(row) > 5 else None),
                }
            )
        return records
    finally:
        workbook.close()


def parse_task_sheet(content: bytes) -> list[dict[str, Any]]:
    workbook = _load(content)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise LegacyWorkbookError("the first worksheet is empty")
        columns = {_text(value): index for index, value in enumerate(header)}
        required = {
            "RefNo",
            "BookNo",
            "PayeeNameTH",
            "PayeeAddress",
            "PayeeTaxID",
            "IncomeType",
            "PaymentDate",
            "Amount",
            "TaxRate",
            "TaxAmount",
        }
        missing = sorted(required.difference(columns))
        if missing:
            raise LegacyWorkbookError(
                f"the first worksheet is missing columns: {', '.join(missing)}"
            )

        def value(row: tuple[Any, ...], name: str) -> Any:
            index = columns.get(name)
            return row[index] if index is not None and index < len(row) else None

        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(cell not in (None, "") for cell in row):
                continue
            task_no = _text(value(row, "RefNo"))
            number_scope = parse_historical_number(task_no)
            if number_scope is None:
                raise LegacyWorkbookError(
                    f"invalid historical WHT number at first worksheet row {row_number}"
                )
            pnd3 = bool(_text(value(row, "WHT Type3")))
            pnd53 = bool(_text(value(row, "WHT Type53")))
            if pnd3 == pnd53:
                raise LegacyWorkbookError(
                    f"exactly one WHT type is required at first worksheet row {row_number}"
                )
            records.append(
                {
                    **number_scope,
                    "task_no": task_no,
                    "book_no": _text(value(row, "BookNo")),
                    "company_name": _text(value(row, "PayeeNameTH")),
                    "company_name_en": _text(value(row, "PayeeNameEN")) or None,
                    "payee_address": _text(value(row, "PayeeAddress")),
                    "tax_id": _tax_id(value(row, "PayeeTaxID")),
                    "wht_type": "PND3" if pnd3 else "PND53",
                    "income_type": _text(value(row, "IncomeType")),
                    "payment_date": _date(value(row, "PaymentDate")),
                    "total_amount": _decimal(value(row, "Amount")),
                    "wht_rate": _decimal(value(row, "TaxRate")),
                    "wht_amount": _decimal(value(row, "TaxAmount")),
                    "amount_text_thai": _text(value(row, "AmountTextThai")) or None,
                    "date_text_thai": _text(value(row, "DateTextThai")) or None,
                }
            )
        return records
    finally:
        workbook.close()


def parse_historical_number(task_no: str) -> dict[str, Any] | None:
    supplement = SUPPLEMENT_NUMBER_PATTERN.fullmatch(task_no)
    if supplement:
        compact = supplement.group("period")
        return {
            "period": f"{compact[:4]}-{compact[4:]}",
            "issuance_type": "supplement",
            "supplement_run": int(supplement.group("run")),
            "sequence": int(supplement.group("sequence")),
        }
    normal = NORMAL_NUMBER_PATTERN.fullmatch(task_no)
    if normal:
        compact = normal.group("period")
        period = f"{compact[:4]}-{compact[4:]}"
        compact_period(period)
        return {
            "period": period,
            "issuance_type": "normal",
            "supplement_run": 0,
            "sequence": int(normal.group("sequence")),
        }
    return None
