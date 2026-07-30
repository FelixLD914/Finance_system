"""台账导出成 Sample 格式的 Excel。

刻意与 recognition.parse_sample_workbook 用同一套表头：导出的文件核对、改完
之后能原样再导回去，不用手工重排列。表头改动必须两边一起动，
tests/test_tax_invoice_ledger_export.py 里有一条测试专门盯住这个往返。

数字格式按业务已确认的口径写死：数量不带小数、单价 4 位、金额 2 位。
Excel 里显示成什么样，人就会照着那个精度核对，格式错了对账就会错。
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.tax_invoice.models import TaxInvoice, TaxInvoiceItem

QUANTITY_FORMAT = "0"
PRICE_FORMAT = "0.0000"
MONEY_FORMAT = "0.00"
DATE_FORMAT = "yyyy-mm-dd"
# 海关汇率按报关单原样显示 6 位：单子上印的就是 31.062700，显示成 31.0627
# 会让人以为系统改过数——核对列的第一原则是和原件逐字符对得上。
RATE_FORMAT = "0.000000"

# (表头, 宽度, 数字格式)。顺序即列顺序。
# 前 14 列是税票级字段，其余是商品级字段——一张税票的多条商品写成多行，
# 税票级字段在每行都重复一遍：parse_sample_workbook 只读每组的第一行，重复
# 不影响回导，但人在 Excel 里筛选排序时不会看到一片空白。
COLUMNS: tuple[tuple[str, int, str | None], ...] = (
    ("DocumentNo", 22, None),
    ("Status", 12, None),
    ("C/I No.", 18, None),
    ("CDN", 20, None),
    ("CI/PI Date", 13, DATE_FORMAT),
    ("Invoice Date", 13, DATE_FORMAT),
    ("FX Date", 13, DATE_FORMAT),
    ("FX Rate", 11, PRICE_FORMAT),
    ("RevRec Period", 13, None),
    ("Customer Name", 32, None),
    ("Customer Address", 40, None),
    ("TAX ID", 18, None),
    ("PO No", 18, None),
    ("INCOTERMS", 11, None),
    ("Payment Term", 20, None),
    ("Product Name or Service", 34, None),
    ("Product Code", 18, None),
    ("HS CODE", 14, None),
    ("Unit", 10, None),
    ("Quantity", 10, QUANTITY_FORMAT),
    ("CI Unit Price", 13, PRICE_FORMAT),
    ("FOB Unit Price USD", 16, PRICE_FORMAT),
    ("FOB Rev USD", 14, MONEY_FORMAT),
    ("FOB Rev THB", 14, MONEY_FORMAT),
    # ── 报关单侧核对列（业务 2026-07-30 新增）──────────────────────────────
    # 追加在末尾而不是插进中间：parse_sample_workbook 按表头名取列、忽略不认识的
    # 列，所以往后加不影响回导；插在中间则会让所有存量导出文件的列序对不上。
    # "海关汇率"是报关单自印的，不是计价用的 FX Rate——两列并排放，差异一眼可见。
    ("Declaration Ref No", 18, None),
    ("Forwarder", 32, None),
    ("Forwarder TAX ID", 16, None),
    ("Customs FX Rate", 14, RATE_FORMAT),
    ("Customs FOB USD", 16, MONEY_FORMAT),
    ("Customs FOB THB (line)", 20, MONEY_FORMAT),
    ("Customs FOB THB (printed)", 22, MONEY_FORMAT),
    ("Customs FOB THB Diff", 20, MONEY_FORMAT),
)

HEADER_FILL = PatternFill("solid", fgColor="EFE7DC")
HEADER_FONT = Font(bold=True, size=10)


def _invoice_cells(invoice: TaxInvoice) -> dict[str, Any]:
    return {
        "DocumentNo": invoice.document_no,
        "Status": invoice.status,
        "C/I No.": invoice.ci_no,
        "CDN": invoice.cdn,
        "CI/PI Date": invoice.ci_date,
        "Invoice Date": invoice.invoice_date,
        "FX Date": invoice.exchange_target_date,
        "FX Rate": invoice.exchange_rate,
        "RevRec Period": invoice.revenue_period,
        "Customer Name": invoice.customer_name,
        "Customer Address": invoice.customer_address,
        "TAX ID": invoice.tax_id,
        "PO No": invoice.po_no,
        "INCOTERMS": invoice.incoterms,
        "Payment Term": invoice.payment_term,
        "Declaration Ref No": invoice.declaration_ref_no,
        # 报关单印英文就是英文、只印泰文就是泰文——写进去的一定是单子上那几个字。
        "Forwarder": invoice.forwarder_name,
        "Forwarder TAX ID": invoice.forwarder_tax_no,
        "Customs FX Rate": invoice.customs_exchange_rate,
        "Customs FOB USD": invoice.customs_fob_usd_total,
        "Customs FOB THB (line)": invoice.customs_fob_thb_line_total,
        "Customs FOB THB (printed)": invoice.customs_fob_thb_printed_total,
        # 差额直接算出来给人看。让对账的人自己在 Excel 里减一遍，就总有人不减。
        # 两边任一为空时留空，而不是写 0——空值和"核对通过"是两回事。
        "Customs FOB THB Diff": _difference(
            invoice.customs_fob_thb_line_total,
            invoice.customs_fob_thb_printed_total,
        ),
    }


def _difference(left: Decimal | None, right: Decimal | None) -> Decimal | None:
    if left is None or right is None:
        return None
    return left - right


def _item_cells(item: TaxInvoiceItem) -> dict[str, Any]:
    return {
        "Product Name or Service": item.product_name,
        "Product Code": item.product_code,
        "HS CODE": item.hs_code,
        "Unit": item.unit,
        "Quantity": item.quantity,
        "CI Unit Price": item.ci_unit_price,
        "FOB Unit Price USD": item.fob_unit_price_usd,
        "FOB Rev USD": item.fob_revenue_usd,
        "FOB Rev THB": item.fob_revenue_thb,
    }


def build_ledger_workbook(
    entries: list[tuple[TaxInvoice, list[TaxInvoiceItem]]],
) -> bytes:
    """把 (税票, 商品明细) 摊平成一行一条商品的 Sample 格式工作簿。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TAX INV"

    for column, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(column)].width = width

    row_number = 2
    for invoice, items in entries:
        invoice_cells = _invoice_cells(invoice)
        # 没有明细的税票也要占一行，否则它在导出里直接消失，对账时看不出缺口。
        for item in items or [None]:
            values = dict(invoice_cells)
            if item is not None:
                values.update(_item_cells(item))
            for column, (header, _, number_format) in enumerate(COLUMNS, start=1):
                value = values.get(header)
                if isinstance(value, Decimal):
                    value = float(value)
                cell = sheet.cell(row=row_number, column=column, value=value)
                if number_format and value is not None:
                    cell.number_format = number_format
            row_number += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(row_number - 1, 1)}"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
