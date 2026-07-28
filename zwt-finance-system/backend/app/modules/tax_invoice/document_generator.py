from __future__ import annotations

import re
from copy import copy
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.modules.tax_invoice import pdf_layout
from app.modules.tax_invoice.models import TaxInvoice, TaxInvoiceItem

# 模板里的大写金额是 BAHTTEXT()，PDF 链路没有 Excel 求值，得自己算。
# WHT 早就实现了同一套泰文数字，直接复用，避免两份实现各写各的。
from app.modules.wht.document_generator import build_blue_signature, thai_baht_text

DATA_START_ROW = 23
MAX_ITEMS = 18

# 「ผู้มีอำนาจลงนาม/Authorized Signature」签名框，(x, y, 宽, 高)，单位 PDF 点，
# 原点在页面左下角。从 TAX-INV-Template.pdf 底版量出：该框横跨右下角
# "On behalf of ZHENWESTERN..." 与 "Authorized Signature" 标签之间。
# 三联的 y 相同（见 pdf_layout 头部说明），所以三页共用这一个框。
# 底版重新制版后必须连同 pdf_layout 一起复核。
SIGNATURE_BOX = (398.0, 112.0, 150.0, 46.0)

# 模板一页横排三联：ORIGINAL 在 B:R，两联 COPY 分别右移 18 / 36 列，
# 且整片区域都是 "=B23" 这类镜像公式。print_area 也正是这三段，
# 所以 Excel 导出的 PDF 恰好是三页、每页一联，行坐标三页完全相同。
COPY_COLUMN_OFFSETS = (0, 18, 36)
COPY_COUNT = len(COPY_COLUMN_OFFSETS)

# 明细列：字段名 -> 第一联的列号。PDF 叠加层按同样的字段名取坐标。
ITEM_COLUMN_FIELDS: tuple[tuple[str, int], ...] = (
    ("line_number", 2),
    ("product_name", 3),
    ("product_code", 5),
    ("unit", 8),
    ("quantity", 9),
    ("fob_unit_price_usd", 11),
    ("fob_revenue_usd", 13),
    ("fx_date", 15),
    ("exchange_rate", 16),
    ("fob_revenue_thb", 18),
)
DATA_COLUMNS = tuple(column for _, column in ITEM_COLUMN_FIELDS)

# 明细列的数字格式，键是第一联的列号。三联必须逐字符一致：税务副联和正本
# 是同一份单据，数量列少一位小数就成了两份不同的文件。
#
# 模板自带的格式做不到这一点——两联 COPY 沿用了会计格式，其中数量列是整数位
# （1101.25 印成 "1,101"，4820.5 甚至进位成 "4,821"），FOB Rev USD 的 31~40 行
# 是三位小数。所以这里对三联都显式覆盖一遍，不依赖模板的状态。
# 模板本身已由 scripts/normalize_tax_inv_item_styles.py 一并修正，
# 这里的覆盖是防止模板被重新部署回旧版本时又退化。
#
# 按业务口径分三类：数量按个计、不带小数，只要千分位；单价与汇率四位小数；
# 金额两位小数。数量用 "#,##0" 而不是 "#,##0.####"，后者格式里的小数点是
# 字面量，Excel 会把 1200 印成 "1,200."，末尾那个点真的会印出来。
ITEM_NUMBER_FORMATS: dict[int, str] = {
    2: "0",  # 序号
    9: "#,##0",  # 数量
    11: "#,##0.0000",  # 单价
    13: "#,##0.00",  # 金额 USD
    15: "@",  # 汇率日期，写的是字符串
    16: "#,##0.0000",  # 汇率
    18: "#,##0.00",  # 金额 THB
}

# 表头字段在三联中的单元格。第二、三联不是整齐的列偏移
# （作者把 DocumentNo. 挪到了 AJ/BB），所以逐个列出而不是算偏移。
HEADER_CELLS: dict[str, tuple[str, str, str]] = {
    "customer_name": ("D13", "V13", "AN13"),
    "customer_address": ("D14", "V14", "AN14"),
    "tax_id": ("D16", "V16", "AN16"),
    "po_no": ("D17", "V17", "AN17"),
    "payment_term": ("D19", "V19", "AN19"),
    "document_no": ("Q13", "AJ13", "BB13"),
    "invoice_date": ("Q14", "AJ14", "BB14"),
}

# 合计区在 xlsx 里是公式（R42=SUM(R23:R40)、C42=BAHTTEXT(R51) 等）。
# PDF 链路没有 Excel 求值，随数据变化的那几格必须从底版清掉、改由 Python
# 算好再叠加。折扣 R43、VAT R47、WHT R49 恒为 0（Q43 空、Q47/Q49 是 0%），
# 渲染 xlsx 时也不改写，就把 Excel 排好的 "-" 留在底版里。
TOTAL_CELLS: dict[str, tuple[str, str, str]] = {
    "amount_text_thai": ("C42", "U42", "AM42"),
    "total_thb": ("R42", "AJ42", "BB42"),
    "after_discount_thb": ("R45", "AJ45", "BB45"),
    "grand_total_thb": ("R51", "AJ51", "BB51"),
}


# 模板把单据号和日期排成粗体；其余字段用常规字重。
BOLD_FIELDS = frozenset({"document_no", "invoice_date"})
THAI_FONT_NAME = "ZwtSarabun"
# 超宽文字最多缩到原字号的这个比例，再放不下就截断。
MIN_FONT_SCALE = 0.62


class TaxInvoiceDocumentGenerationError(RuntimeError):
    pass


def _display_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def _quantize(value: Decimal, places: int) -> Decimal:
    # Excel 四舍五入，Python 默认银行家舍入；汇率是 Numeric(18,6) 而模板只显示
    # 4 位，差异会真的显出来，所以显式指定 ROUND_HALF_UP。
    return Decimal(value).quantize(Decimal(1).scaleb(-places), rounding=ROUND_HALF_UP)


def format_amount(value: Decimal | None) -> str:
    """还原 render_tax_invoice_workbook 设的 "#,##0.00"。"""
    return "" if value is None else f"{_quantize(value, 2):,.2f}"


def format_rate(value: Decimal | None) -> str:
    """还原 "#,##0.0000"（单价与汇率列）。"""
    return "" if value is None else f"{_quantize(value, 4):,.4f}"


def format_quantity(value: Decimal | None) -> str:
    """还原数量列的 "#,##0"。

    产品按个计，数量不带小数，只要千分位分隔符。真出现了小数，Excel 的
    "#,##0" 会四舍五入（4820.5 -> "4,821"），这里用同样的 ROUND_HALF_UP
    跟上，保证 PDF 和 xlsx 印出来是同一个数。
    """
    return "" if value is None else f"{_quantize(value, 0):,.0f}"


def render_tax_invoice_workbook(
    template_path: Path,
    invoice: TaxInvoice,
    items: list[TaxInvoiceItem],
) -> bytes:
    if not template_path.is_file():
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV template was not found: {template_path}"
        )
    if not invoice.document_no or not invoice.invoice_date:
        raise TaxInvoiceDocumentGenerationError(
            "approved document number and invoice date are required"
        )
    if len(items) > MAX_ITEMS:
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV template supports at most {MAX_ITEMS} product lines"
        )

    workbook = load_workbook(template_path)
    worksheet = workbook.active
    worksheet["D13"] = invoice.customer_name
    worksheet["D14"] = invoice.customer_address
    worksheet["D16"] = invoice.tax_id or ""
    worksheet["D17"] = invoice.po_no or ""
    worksheet["D19"] = invoice.payment_term or ""
    worksheet["Q13"] = invoice.document_no
    # Business rule: TAX invoice date follows the customs submission date.
    worksheet["Q14"] = _display_date(invoice.invoice_date)

    for offset in range(MAX_ITEMS):
        row_number = DATA_START_ROW + offset
        for column in DATA_COLUMNS:
            worksheet.cell(row=row_number, column=column).value = ""
        if offset >= len(items):
            continue
        item = items[offset]
        values = {
            2: item.line_number,
            3: item.product_name or "",
            5: item.product_code or "",
            8: item.unit or "",
            9: item.quantity,
            11: item.fob_unit_price_usd,
            13: item.fob_revenue_usd,
            # FX target date keeps the original exchange-rate display rule.
            15: _display_date(invoice.exchange_target_date),
            16: invoice.exchange_rate,
            18: item.fob_revenue_thb,
        }
        for column, value in values.items():
            worksheet.cell(row=row_number, column=column).value = value
        # 值只写第一联，两联 COPY 靠 "=B23" 这类公式取；但公式不带样式，
        # 所以格式必须三联各写一遍。
        for column, number_format in ITEM_NUMBER_FORMATS.items():
            for copy_offset in COPY_COLUMN_OFFSETS:
                cell = worksheet.cell(row=row_number, column=column + copy_offset)
                cell.number_format = number_format
        # 对齐得跟着格式一起统一。模板给副联留的会计格式靠 "* " 重复填充把
        # 数字顶到右边，把单元格自己声明的水平对齐压住了；换成普通格式之后
        # 那个声明就生效了，而三联的声明并不一致（COPY#1 的汇率列是 left）。
        for column in DATA_COLUMNS:
            wanted = worksheet.cell(row=row_number, column=column).alignment
            for copy_offset in COPY_COLUMN_OFFSETS[1:]:
                cell = worksheet.cell(row=row_number, column=column + copy_offset)
                cell.alignment = copy(wanted)

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def invoice_total_thb(items: list[TaxInvoiceItem]) -> Decimal:
    """明细 THB 金额合计——对应模板里的 =SUM(R23:R40)。

    刻意按明细行相加，而不是取 invoice.fob_revenue_thb_total，
    这样 PDF 的合计和 xlsx 里 Excel 算出来的一定是同一个数。
    """
    return sum((item.fob_revenue_thb or Decimal("0") for item in items), Decimal("0"))


def pdf_field_values(
    invoice: TaxInvoice,
    items: list[TaxInvoiceItem],
) -> tuple[dict[str, str], list[dict[str, str]], dict[str, str]]:
    """PDF 叠加层要写的所有字符串，与 xlsx 里 Excel 显示的结果一致。"""
    total = invoice_total_thb(items)
    header = {
        "customer_name": invoice.customer_name,
        "customer_address": invoice.customer_address,
        "tax_id": invoice.tax_id or "",
        "po_no": invoice.po_no or "",
        "payment_term": invoice.payment_term or "",
        "document_no": invoice.document_no or "",
        "invoice_date": _display_date(invoice.invoice_date),
    }
    fx_date = _display_date(invoice.exchange_target_date)
    exchange_rate = format_rate(invoice.exchange_rate)
    lines = [
        {
            "line_number": str(item.line_number),
            "product_name": item.product_name or "",
            "product_code": item.product_code or "",
            "unit": item.unit or "",
            "quantity": format_quantity(item.quantity),
            "fob_unit_price_usd": format_rate(item.fob_unit_price_usd),
            "fob_revenue_usd": format_amount(item.fob_revenue_usd),
            "fx_date": fx_date,
            "exchange_rate": exchange_rate,
            "fob_revenue_thb": format_amount(item.fob_revenue_thb),
        }
        for item in items
    ]
    totals = {
        # 底版上 R43 折扣、R47 VAT、R49 WHT 恒为 0，Excel 已经把 "-" 印好了；
        # 因此 R45=R42、R51=R45，三格都是明细合计。
        "amount_text_thai": f"({thai_baht_text(total)})",
        "total_thb": format_amount(total),
        "after_discount_thb": format_amount(total),
        "grand_total_thb": format_amount(total),
    }
    return header, lines, totals


def _font_for(text: str) -> str:
    # 底版正文是 Times New Roman，ReportLab 内置的 Times-Roman 与之度量兼容，
    # 不用嵌入字体就能对齐；泰文（客户名、地址、大写金额）走 Sarabun。
    return "Times-Roman" if text.isascii() else THAI_FONT_NAME


def export_pdf_from_template(
    pdf_template_path: Path,
    pdf_path: Path,
    invoice: TaxInvoice,
    items: list[TaxInvoiceItem],
    thai_font_path: Path,
    signature_source_path: Path | None = None,
) -> None:
    """在静态底版上叠加业务数据，生成三联 TAX INV，全程不碰 Office。

    signature_source_path 给定时，在「ผู้มีอำนาจลงนาม/Authorized Signature」框内
    盖章。三联都盖：副联必须与正本一致。
    """
    if not pdf_template_path.is_file():
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV PDF template was not found: {pdf_template_path}"
        )
    if not thai_font_path.is_file():
        raise TaxInvoiceDocumentGenerationError(f"Thai PDF font was not found: {thai_font_path}")
    if not invoice.document_no or not invoice.invoice_date:
        raise TaxInvoiceDocumentGenerationError(
            "approved document number and invoice date are required"
        )
    if len(items) > MAX_ITEMS:
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV template supports at most {MAX_ITEMS} product lines"
        )

    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.lib.colors import black
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise TaxInvoiceDocumentGenerationError(
            "PDF generation requires reportlab and pypdf"
        ) from exc

    pdfmetrics.registerFont(TTFont(THAI_FONT_NAME, str(thai_font_path)))
    # 与 WHT 共用同一套蓝笔化处理：去掉近白底、只留笔画并统一成蓝色墨迹，
    # 两种单据上的签名观感必须一致。
    prepared_signature: Path | None = None
    if signature_source_path is not None:
        prepared_signature = pdf_path.with_suffix(".signature.png")
        build_blue_signature(signature_source_path, prepared_signature)
    header, lines, totals = pdf_field_values(invoice, items)
    reader = PdfReader(str(pdf_template_path))
    if len(reader.pages) != pdf_layout.PAGE_COUNT:
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV PDF template must have {pdf_layout.PAGE_COUNT} pages, "
            f"found {len(reader.pages)}"
        )
    writer = PdfWriter()

    def draw(
        target: Any,
        text: str,
        anchor: pdf_layout.TextAnchor,
        baseline: float,
        *,
        bold: bool = False,
    ) -> None:
        if not text:
            return
        font = _font_for(text)
        if bold and font == "Times-Roman":
            font = "Times-Bold"
        size, text = _fit(text, font, anchor.size, anchor.max_width)
        target.setFont(font, size)
        if anchor.align == "right":
            target.drawRightString(anchor.x, baseline, text)
        elif anchor.align == "center":
            target.drawCentredString(anchor.x, baseline, text)
        else:
            target.drawString(anchor.x, baseline, text)

    for index, page in enumerate(reader.pages):
        buffer = BytesIO()
        overlay = canvas.Canvas(
            buffer,
            pagesize=(float(page.mediabox.width), float(page.mediabox.height)),
        )
        overlay.setFillColor(black)

        for field, value in header.items():
            draw(
                overlay,
                value,
                pdf_layout.HEADER_ANCHORS[field][index],
                pdf_layout.HEADER_BASELINES[field],
                bold=field in BOLD_FIELDS,
            )
        for line_index, line in enumerate(lines):
            for field, value in line.items():
                draw(
                    overlay,
                    value,
                    pdf_layout.ITEM_ANCHORS[field][index],
                    pdf_layout.ITEM_BASELINES[field][line_index],
                )
        for field, value in totals.items():
            draw(
                overlay,
                value,
                pdf_layout.TOTAL_ANCHORS[field][index],
                pdf_layout.TOTAL_BASELINES[field],
            )

        if prepared_signature is not None and prepared_signature.exists():
            overlay.drawImage(
                ImageReader(str(prepared_signature)),
                SIGNATURE_BOX[0],
                SIGNATURE_BOX[1],
                width=SIGNATURE_BOX[2],
                height=SIGNATURE_BOX[3],
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )

        overlay.save()
        buffer.seek(0)
        writer.add_page(page)
        writer.pages[-1].merge_page(PdfReader(buffer).pages[0])

    with pdf_path.open("wb") as output:
        writer.write(output)
    if prepared_signature is not None:
        prepared_signature.unlink(missing_ok=True)


def _fit(text: str, font: str, size: float, max_width: float | None) -> tuple[float, str]:
    """把文字塞进格子：先缩字号，缩到下限还超宽就截断。

    Excel 的做法是直接裁掉超出的部分；税务单据上宁可字小一点也要显示完整，
    所以先缩到 MIN_FONT_SCALE，实在放不下才截。
    """
    from reportlab.pdfbase.pdfmetrics import stringWidth

    if max_width is None:
        return size, text
    width = stringWidth(text, font, size)
    if width <= max_width:
        return size, text
    size = max(size * MIN_FONT_SCALE, size * max_width / width)
    while text and stringWidth(text, font, size) > max_width:
        text = text[:-1]
    return size, text


def data_cell_coordinates() -> list[str]:
    """所有承载业务数据的单元格（三联全算），供制版时清空。"""
    coordinates: list[str] = []
    for cells in (*HEADER_CELLS.values(), *TOTAL_CELLS.values()):
        coordinates.extend(cells)
    for offset in COPY_COLUMN_OFFSETS:
        for row_offset in range(MAX_ITEMS):
            row = DATA_START_ROW + row_offset
            for column in DATA_COLUMNS:
                coordinates.append(f"{get_column_letter(column + offset)}{row}")
    return coordinates


def render_clean_template_workbook(template_path: Path) -> bytes:
    """清掉数据占位符，供管理员在开发机上一次性导出 PDF 底版。

    WHT 的同名函数是全表扫描 "strip() 后等于字段名" 的单元格，这里不能照搬：
    TAX INV 模板的表头 H22="Unit"、O22=" FX Date"、P22="FX Rate" 与占位符同名，
    全表匹配会把三联的列名一起抹掉。改成只清 renderer 真正写入的坐标。
    """
    if not template_path.is_file():
        raise TaxInvoiceDocumentGenerationError(
            f"TAX INV template was not found: {template_path}"
        )
    workbook = load_workbook(template_path)
    worksheet = workbook["Invoice"]
    for coordinate in data_cell_coordinates():
        worksheet[coordinate] = None
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def tax_invoice_file_stem(invoice: TaxInvoice) -> str:
    if not invoice.document_no or not invoice.invoice_date:
        raise TaxInvoiceDocumentGenerationError(
            "approved document number and invoice date are required"
        )
    amount = f"{Decimal(invoice.fob_revenue_thb_total):,.2f}"
    raw = (
        f"{invoice.invoice_date:%Y%m%d}-TAX INV"
        f"({invoice.document_no})-THB{amount}"
    )
    return re.sub(r'[\\/:*?"<>|,]', "", raw)
