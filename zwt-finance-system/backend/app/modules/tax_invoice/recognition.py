from __future__ import annotations

import re
from collections.abc import Mapping
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.modules.tax_invoice.customs_declaration import (
    CustomsDeclarationError,
    parse_export_declaration,
)

MONEY = Decimal("0.01")
PRICE = Decimal("0.0001")
MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


class TaxInvoiceRecognitionError(ValueError):
    pass


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: object, *, quantum: Decimal | None = None) -> Decimal | None:
    text = _safe_text(value).replace(",", "")
    if not text:
        return None
    try:
        parsed = Decimal(text)
        # quantize 必须留在 try 内：单元格值超出 Decimal 上下文精度时（例如
        # 1e100）它同样抛 InvalidOperation。识别阶段任何解析不了的数值都返回
        # None，由 _review_status 标记 needs_review 交人工确认，不能让上传接口 500。
        return parsed.quantize(quantum, rounding=ROUND_HALF_UP) if quantum else parsed
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _safe_text(value)
    if not text:
        return None
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        try:
            return date(*(int(part) for part in match.groups()))
        except ValueError:
            return None
    match = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{4})", text)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    if year >= 2400:
        year -= 543
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _xlsx_grid(content: bytes) -> tuple[list[list[object]], str]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except (BadZipFile, InvalidFileException, KeyError) as exc:
        # 扩展名对不代表内容对：改名的 PDF、下载了一半的文件、真 zip 但不是
        # xlsx，都会在这里炸。openpyxl 抛的是 BadZipFile / InvalidFileException
        # /（缺 workbook.xml 时）KeyError，都不是 TaxInvoiceRecognitionError，
        # 于是一路冒到 FastAPI 变成 500。用户传错文件是 4xx，不是服务器故障。
        raise TaxInvoiceRecognitionError(
            "the file is not a readable .xlsx workbook; "
            "re-save it from Excel and upload again"
        ) from exc
    worksheet = next(
        (
            workbook[name]
            for name in workbook.sheetnames
            if re.search(r"\bCI\b|INVOICE", name, re.IGNORECASE)
        ),
        workbook.active,
    )
    rows = [
        [
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        for row in range(1, worksheet.max_row + 1)
    ]
    title = worksheet.title
    workbook.close()
    return rows, title


def _xls_grid(content: bytes) -> tuple[list[list[object]], str]:
    try:
        import xlrd
    except ImportError as exc:
        raise TaxInvoiceRecognitionError("legacy .xls import requires xlrd") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except xlrd.XLRDError as exc:
        # 同 _xlsx_grid：内容不是真的 .xls 时给 4xx，不要冒成 500。
        raise TaxInvoiceRecognitionError(
            "the file is not a readable .xls workbook; "
            "re-save it from Excel and upload again"
        ) from exc
    sheet = next(
        (
            workbook.sheet_by_name(name)
            for name in workbook.sheet_names()
            if re.search(r"\bCI\b|INVOICE", name, re.IGNORECASE)
        ),
        workbook.sheet_by_index(0),
    )
    return ([sheet.row_values(index) for index in range(sheet.nrows)], sheet.name)


def _grid(content: bytes, file_name: str) -> tuple[list[list[object]], str]:
    if file_name.lower().endswith(".xlsx"):
        return _xlsx_grid(content)
    if file_name.lower().endswith(".xls"):
        return _xls_grid(content)
    raise TaxInvoiceRecognitionError("invoice source must be .xlsx or .xls")


def parse_invoice_workbook(content: bytes, file_name: str) -> dict[str, Any]:
    rows, sheet_name = _grid(content, file_name)
    if not rows:
        raise TaxInvoiceRecognitionError("invoice workbook is empty")
    max_columns = max((len(row) for row in rows), default=0)

    def value(row: int, column: int) -> object:
        if row < 0 or row >= len(rows) or column < 0 or column >= len(rows[row]):
            return None
        return rows[row][column]

    def right_text(row: int, column: int, distance: int = 8) -> str:
        for offset in range(1, distance + 1):
            text = _safe_text(value(row, column + offset))
            if text:
                return text
        return ""

    result: dict[str, Any] = {
        "ci_date": None,
        "ci_no": "",
        "incoterms": "",
        "payment_term": "",
        "po_no": "",
        "customer_name": "",
        "customer_address": "",
        "items": [],
        "fob_amount_usd": None,
        "quantity_total": None,
        "freight_usd": None,
        "insurance_usd": None,
        "source_sheet": sheet_name,
    }

    buyer_row: int | None = None
    buyer_column: int | None = None
    for row_index, row in enumerate(rows):
        for column_index, raw in enumerate(row):
            text = _safe_text(raw)
            if not text:
                continue
            if result["ci_date"] is None:
                match = re.match(r"DATE\s*[：:]\s*(.*)", text, re.IGNORECASE)
                if match:
                    date_value = match.group(1) or right_text(
                        row_index,
                        column_index,
                        4,
                    )
                    result["ci_date"] = _parse_date(date_value)
            if not result["ci_no"]:
                match = re.search(
                    r"INV\.?\s*NO\.?\s*[：:]\s*(ZWT-\S+)",
                    text,
                    re.IGNORECASE,
                )
                if match:
                    result["ci_no"] = match.group(1).rstrip(".")
                elif re.search(r"INV\.?\s*NO\.?\s*[：:]\s*$", text, re.IGNORECASE):
                    candidate = right_text(row_index, column_index, 4)
                    if candidate.upper().startswith("ZWT-"):
                        result["ci_no"] = candidate.rstrip(".")
            if not result["incoterms"]:
                match = re.search(r"INCOTERMS\s*[：:]\s*([A-Z]{2,5})", text, re.IGNORECASE)
                if match:
                    result["incoterms"] = match.group(1).upper()
                elif re.search(r"^INCOTERMS\s*[：:]\s*$", text, re.IGNORECASE):
                    match = re.match(r"([A-Z]{2,5})", right_text(row_index, column_index, 4))
                    if match:
                        result["incoterms"] = match.group(1).upper()
            if not result["payment_term"]:
                match = re.search(r"PAYMENT\s*TERMS?\s*[：:]\s*(.+)", text, re.IGNORECASE)
                if match:
                    result["payment_term"] = match.group(1).strip()
                elif re.search(r"^PAYMENT\s*TERMS?\s*[：:]\s*$", text, re.IGNORECASE):
                    result["payment_term"] = right_text(row_index, column_index)
            if not result["po_no"]:
                match = re.search(r"PO\.?\s*NO\.?\s*[：:]\s*(\S+)", text, re.IGNORECASE)
                if match:
                    result["po_no"] = match.group(1).rstrip(".")
                elif re.search(r"^PO\.?\s*NO\.?\s*[：:]\s*$", text, re.IGNORECASE):
                    result["po_no"] = right_text(row_index, column_index, 4).rstrip(".")
            if buyer_row is None and re.match(r"BUYER\s*[：:]\s*$", text, re.IGNORECASE):
                customer = right_text(row_index, column_index, max_columns)
                if customer:
                    result["customer_name"] = customer
                    buyer_row = row_index
                    for offset in range(1, max_columns):
                        if _safe_text(value(row_index, column_index + offset)) == customer:
                            buyer_column = column_index + offset
                            break

    if buyer_row is not None and buyer_column is not None:
        for offset in range(1, 5):
            address = _safe_text(value(buyer_row + offset, buyer_column))
            if address:
                result["customer_address"] = address
                break

    header_row: int | None = None
    columns: dict[str, int] = {}
    for row_index, row in enumerate(rows):
        if any(re.search(r"ITEM\s*NO", _safe_text(cell), re.IGNORECASE) for cell in row):
            header_row = row_index
            for column_index, raw in enumerate(row):
                text = _safe_text(raw).upper()
                if re.search(r"ITEM\s*NO", text):
                    columns["item_no"] = column_index
                elif "PRODUCT CODE" in text:
                    columns["product_code"] = column_index
                elif "PRODUCT NAME" in text:
                    columns["product_name"] = column_index
                elif re.search(r"H\.?S\.?\s*CODE", text):
                    columns["hs_code"] = column_index
                elif re.fullmatch(r"QUANTITY|QTY", text):
                    columns["quantity"] = column_index
                elif "UNIT PRICE" in text:
                    columns["ci_unit_price"] = column_index
                elif text == "UNIT":
                    columns["unit"] = column_index
                elif "AMOUNT" in text and "USD" in text:
                    columns["amount_usd"] = column_index
            break

    if header_row is not None:
        item_number_column = columns.get("item_no", 0)
        for row_index in range(header_row + 1, len(rows)):
            marker = _safe_text(value(row_index, item_number_column))
            if not marker:
                break
            if "TOTAL" in marker.upper():
                result["quantity_total"] = _decimal(
                    value(row_index, columns.get("quantity", -1))
                )
                if result["fob_amount_usd"] is None and "amount_usd" in columns:
                    result["fob_amount_usd"] = _decimal(
                        value(row_index, columns["amount_usd"]),
                        quantum=MONEY,
                    )
                break
            try:
                int(float(marker))
            except ValueError:
                continue
            result["items"].append(
                {
                    "product_name": _safe_text(value(row_index, columns.get("product_name", -1))),
                    "product_code": _safe_text(value(row_index, columns.get("product_code", -1))),
                    "hs_code": _safe_text(
                        value(row_index, columns.get("hs_code", -1))
                    ).removesuffix(".0"),
                    "unit": _safe_text(value(row_index, columns.get("unit", -1))),
                    "quantity": _decimal(value(row_index, columns.get("quantity", -1))),
                    "ci_unit_price": _decimal(
                        value(row_index, columns.get("ci_unit_price", -1)),
                        quantum=PRICE,
                    ),
                }
            )

    for row_index, row in enumerate(rows):
        for column_index, raw in enumerate(row):
            text = _safe_text(raw).upper()
            if (
                result["fob_amount_usd"] is None
                and re.search(r"\bFOB\b", text)
                and re.search(r"AMOUNT|AMT|PRICE", text)
                and "UNIT" not in text
            ):
                for offset in range(1, 13):
                    amount = _decimal(value(row_index, column_index + offset), quantum=MONEY)
                    if amount is not None:
                        result["fob_amount_usd"] = amount
                        break
            if text in {"FREIGHT", "INSURANCE"}:
                field = "freight_usd" if text == "FREIGHT" else "insurance_usd"
                if result[field] is None:
                    for offset in range(1, 6):
                        amount = _decimal(value(row_index, column_index + offset), quantum=MONEY)
                        if amount is not None:
                            result[field] = amount
                            break

    if not result["quantity_total"]:
        result["quantity_total"] = sum(
            (item["quantity"] or Decimal("0") for item in result["items"]),
            start=Decimal("0"),
        )
    if not result["ci_no"]:
        raise TaxInvoiceRecognitionError("C/I number was not recognized from invoice workbook")
    return result


def parse_customs_pdf(content: bytes) -> dict[str, Any]:
    """出口报关单 → 识别结果 dict（口径见 customs_declaration 模块）。

    这层只做两件事：把 `CustomsDeclarationError` 翻成 `TaxInvoiceRecognitionError`
    （上传接口靠它给 4xx，不能让新异常一路冒成 500），以及把结构化结果摊平成
    service/combine 一直在用的那套键名。真正的识别逻辑全在 customs_declaration 里，
    那里能脱离 FastAPI 单测。
    """
    try:
        parsed = parse_export_declaration(content)
    except CustomsDeclarationError as exc:
        raise TaxInvoiceRecognitionError(str(exc)) from exc

    submission = parsed.fields["submission_date"]
    return {
        # ── 原有键，调用方不用改 ────────────────────────────────────────────
        "cdn": parsed.value("cdn") or "",
        "ci_no": parsed.value("ci_no") or "",
        "ci_date": parsed.value("ci_date"),
        "submission_date": submission.value,
        "submission_date_confidence": submission.confidence,
        "submission_date_source": submission.source,
        "submission_date_low_confidence": submission.review_required,
        # 计价链路一直用这个键做 FOB 验算。优先用报关单自印合计（有独立来源、
        # 能和行级互证），没有自印合计时退回行级加总。
        "customs_fob_usd_total": (
            parsed.value("customs_fob_usd_printed_total")
            or parsed.value("customs_fob_usd_line_total")
        ),
        # ── 业务 2026-07-30 新增：海关汇率 / 货代 / 出口泰铢金额 ──────────────
        "declaration_ref_no": parsed.value("declaration_ref_no") or "",
        "customs_exchange_rate": parsed.value("customs_exchange_rate"),
        # 业务口径：报关单印英文就写英文，只印泰文就写泰文。forwarder_name 是落库/
        # 导出用的那一个值；th/en 两个原文字段留着做取值依据。
        "forwarder_name": parsed.value("forwarder_name") or "",
        "forwarder_name_th": parsed.value("forwarder_name_th") or "",
        "forwarder_name_en": parsed.value("forwarder_name_en") or "",
        "forwarder_tax_no": parsed.value("forwarder_tax_no") or "",
        "customs_fob_thb_total": (
            parsed.value("customs_fob_thb_printed_total")
            or parsed.value("customs_fob_thb_line_total")
        ),
        "customs_fob_thb_line_total": parsed.value("customs_fob_thb_line_total"),
        "customs_fob_thb_printed_total": parsed.value("customs_fob_thb_printed_total"),
        "customs_fob_usd_line_total": parsed.value("customs_fob_usd_line_total"),
        "customs_fob_usd_printed_total": parsed.value("customs_fob_usd_printed_total"),
        "customs_items": [
            {
                "line_number": item.line_number,
                "fob_usd": item.fob_usd,
                "fob_thb": item.fob_thb,
            }
            for item in parsed.items
        ],
        # 识别阶段发现的不一致（行级 vs 自印合计、泰铢 vs 汇率×美元）。
        # 只报不改：擅自挑一个"看起来对的"就等于把这道核对做没了。
        "customs_warnings": list(parsed.warnings),
        # 每个字段的原文与可信度，供复核页并排显示"系统读到的 vs 报关单原文"。
        "customs_field_evidence": {
            name: field.as_dict() for name, field in parsed.fields.items()
        },
    }


def parse_bot_fx_workbook(content: bytes, file_name: str) -> dict[date, Decimal]:
    rows, _ = _grid(content, file_name)
    date_row: int | None = None
    rate_row: int | None = None
    for index, row in enumerate(rows):
        second = _safe_text(row[1] if len(row) > 1 else "").upper()
        third = _safe_text(row[2] if len(row) > 2 else "").upper()
        if date_row is None and re.search(r"\d{1,2}\s+[A-Z]{3}\s+\d{4}", third):
            date_row = index
        if "REFERENCE RATE" in second and "USD" in second:
            rate_row = index + 3
            break
    if date_row is None or rate_row is None or rate_row >= len(rows):
        raise TaxInvoiceRecognitionError("BOT workbook date/rate rows were not recognized")
    rates: dict[date, Decimal] = {}
    column_count = max(len(rows[date_row]), len(rows[rate_row]))
    for column in range(2, column_count):
        raw_date = _safe_text(
            rows[date_row][column] if column < len(rows[date_row]) else ""
        ).upper()
        raw_rate = rows[rate_row][column] if column < len(rows[rate_row]) else None
        match = re.match(r"(\d{1,2})\s+([A-Z]{3})\s+(\d{4})", raw_date)
        rate = _decimal(raw_rate, quantum=PRICE)
        if not match or rate is None or match.group(2) not in MONTHS:
            continue
        rates[
            date(int(match.group(3)), MONTHS[match.group(2)], int(match.group(1)))
        ] = rate
    return rates


def parse_sample_workbook(content: bytes, file_name: str) -> list[dict[str, Any]]:
    rows, _ = _grid(content, file_name)
    if len(rows) < 2:
        raise TaxInvoiceRecognitionError("sample workbook has no data rows")
    headers = [_safe_text(value) for value in rows[0]]
    index = {name: position for position, name in enumerate(headers) if name}
    required = {
        "CDN",
        "Customer Name",
        "Customer Address",
        "Product Name or Service",
        "Product Code",
        "Unit",
        "Quantity",
        "FX Date",
        "FX Rate",
        "FOB Rev USD",
        "FOB Rev THB",
    }
    missing = sorted(required.difference(index))
    if missing:
        raise TaxInvoiceRecognitionError(
            f"sample workbook is missing columns: {', '.join(missing)}"
        )

    def cell(row: list[object], name: str) -> object:
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    # 行号按 Excel 的 1-based 算并跳过表头，报错时能直接说"第 N 行"。
    # 分组会把多行合成一张税票，源行号丢了就只能报业务键，对着表格找不着人。
    groups: dict[str, list[tuple[int, list[object]]]] = {}
    for excel_row, row in enumerate(rows[1:], start=2):
        if not any(_safe_text(value) for value in row):
            continue
        key = (
            _safe_text(cell(row, "DocumentNo"))
            or _safe_text(cell(row, "CDN"))
            or _safe_text(cell(row, "C/I No."))
        )
        if not key:
            raise TaxInvoiceRecognitionError(
                f"sample row {excel_row} has no DocumentNo, CDN or C/I No."
            )
        groups.setdefault(key, []).append((excel_row, row))

    invoices: list[dict[str, Any]] = []
    for group in groups.values():
        source_rows = [excel_row for excel_row, _ in group]
        group_rows = [row for _, row in group]
        first = group_rows[0]
        exchange_target_date = _parse_date(cell(first, "FX Date"))
        explicit_invoice_date = _parse_date(
            cell(first, "Invoice Date") or cell(first, "Submission Date")
        )
        invoice_date = explicit_invoice_date or exchange_target_date
        exchange_rate = _decimal(cell(first, "FX Rate"), quantum=PRICE)
        items: list[dict[str, Any]] = []
        for line_number, row in enumerate(group_rows, start=1):
            items.append(
                {
                    "line_number": line_number,
                    "product_name": _safe_text(cell(row, "Product Name or Service")),
                    "product_code": _safe_text(cell(row, "Product Code")),
                    "hs_code": _safe_text(cell(row, "HS CODE")).removesuffix(".0"),
                    "unit": _safe_text(cell(row, "Unit")),
                    "quantity": _decimal(cell(row, "Quantity")),
                    "ci_unit_price": _decimal(cell(row, "CI Unit Price"), quantum=PRICE),
                    "fob_unit_price_usd": _decimal(
                        cell(row, "FOB Unit Price USD")
                        or cell(row, "Unit Price USD"),
                        quantum=PRICE,
                    ),
                    "fob_revenue_usd": _decimal(cell(row, "FOB Rev USD"), quantum=MONEY),
                    "fob_revenue_thb": _decimal(cell(row, "FOB Rev THB"), quantum=MONEY),
                }
            )
        document_no = _safe_text(cell(first, "DocumentNo"))
        if document_no and not re.fullmatch(r"ZWT-IV\d{8}-\d{2,}", document_no):
            raise TaxInvoiceRecognitionError(
                f"row {source_rows[0]}: invalid existing DocumentNo: {document_no}"
            )
        invoices.append(
            {
                "source_rows": source_rows,
                "document_no": document_no or None,
                "ci_no": _safe_text(cell(first, "C/I No.")) or "LEGACY-NO-CI",
                "cdn": _safe_text(cell(first, "CDN")) or None,
                "ci_date": _parse_date(cell(first, "CI/PI Date")),
                "invoice_date": invoice_date,
                "exchange_target_date": exchange_target_date,
                "exchange_rate_date": exchange_target_date if exchange_rate else None,
                "revenue_period": (
                    _safe_text(cell(first, "RevRec Period"))
                    or (invoice_date.strftime("%Y%m") if invoice_date else None)
                ),
                "currency": "USD",
                "exchange_rate": exchange_rate,
                "customer_name": _safe_text(cell(first, "Customer Name")),
                "customer_address": _safe_text(cell(first, "Customer Address")),
                "tax_id": _safe_text(cell(first, "TAX ID")) or None,
                "po_no": _safe_text(cell(first, "PO No")) or None,
                "incoterms": _safe_text(cell(first, "INCOTERMS")).upper() or None,
                "payment_term": _safe_text(cell(first, "Payment Term")) or None,
                "items": items,
                "fob_revenue_usd_total": sum(
                    (item["fob_revenue_usd"] or Decimal("0") for item in items),
                    start=Decimal("0"),
                ).quantize(MONEY),
                "fob_revenue_thb_total": sum(
                    (item["fob_revenue_thb"] or Decimal("0") for item in items),
                    start=Decimal("0"),
                ).quantize(MONEY),
                "is_dap": _safe_text(cell(first, "INCOTERMS")).upper() == "DAP",
                "fob_verification_failed": False,
                "submission_date_low_confidence": explicit_invoice_date is None,
                "submission_date_confidence": (
                    "high" if explicit_invoice_date else "legacy_fallback"
                ),
                "submission_date_source": (
                    "sample_invoice_date"
                    if explicit_invoice_date
                    else "legacy_fx_date_fallback"
                ),
            }
        )
    return invoices


def lookup_fx_rate(
    rates: Mapping[date, Decimal],
    target_date: date,
    *,
    maximum_lookback_days: int = 9,
) -> tuple[Decimal | None, date]:
    for offset in range(maximum_lookback_days + 1):
        candidate = target_date - timedelta(days=offset)
        if candidate in rates:
            return rates[candidate], candidate
    return None, target_date


def recompute_line_thb(
    fob_usd: Decimal | None,
    rate: Decimal | None,
) -> Decimal | None:
    """一行 USD 金额 × 汇率 → THB，四舍五入到分。

    THB 折算的舍入口径只应有这一处：识别建单（combine_invoice_and_customs）与
    事后的汇率重匹配（service.match_exchange_rate）都从这里取，别再各写一遍
    quantize——两处一旦漂移，台账和 PDF 印出来就会差一分，而这种差异是静默的。
    """
    if fob_usd is None or rate is None:
        return None
    return (fob_usd * rate).quantize(MONEY, rounding=ROUND_HALF_UP)


def combine_invoice_and_customs(
    invoice: Mapping[str, Any],
    customs: Mapping[str, Any],
    rates: Mapping[date, Decimal],
    *,
    currency: str = "USD",
) -> dict[str, Any]:
    invoice_date = customs.get("submission_date")
    if not isinstance(invoice_date, date):
        invoice_date = None
    exchange_target_date = invoice_date
    exchange_rate: Decimal | None = None
    exchange_rate_date: date | None = None
    if exchange_target_date:
        exchange_rate, matched = lookup_fx_rate(rates, exchange_target_date)
        exchange_rate_date = matched if exchange_rate is not None else None

    incoterms = _safe_text(invoice.get("incoterms")).upper()
    quantity_total = invoice.get("quantity_total") or Decimal("0")
    freight = invoice.get("freight_usd") or Decimal("0")
    insurance = invoice.get("insurance_usd") or Decimal("0")
    items: list[dict[str, Any]] = []
    for line_number, source in enumerate(invoice.get("items") or [{}], start=1):
        quantity = source.get("quantity")
        unit_price = source.get("ci_unit_price")
        fob_unit_price: Decimal | None = None
        if unit_price is not None:
            if incoterms == "FOB":
                fob_unit_price = unit_price.quantize(PRICE, rounding=ROUND_HALF_UP)
            elif incoterms == "DAP" and quantity_total:
                fob_unit_price = (
                    unit_price - ((freight + insurance) / quantity_total)
                ).quantize(PRICE, rounding=ROUND_HALF_UP)
            else:
                fob_unit_price = (unit_price * Decimal("1.03")).quantize(
                    PRICE,
                    rounding=ROUND_HALF_UP,
                )
        fob_usd = (
            (fob_unit_price * quantity).quantize(MONEY, rounding=ROUND_HALF_UP)
            if fob_unit_price is not None and quantity is not None
            else None
        )
        fob_thb = recompute_line_thb(fob_usd, exchange_rate)
        items.append(
            {
                **source,
                "line_number": line_number,
                "fob_unit_price_usd": fob_unit_price,
                "fob_revenue_usd": fob_usd,
                "fob_revenue_thb": fob_thb,
                # 报关单该行的 USD 在下方按行位置回填；先占位保证每个 item dict
                # 键一致，TaxInvoiceItem(**item) 才稳定。
                "customs_fob_usd": None,
            }
        )

    # 逐行核对口径（2026-07-30）：报关单第 N 行 ↔ 发票第 N 行，顺序一致。把报关单
    # 该行自印的 FOB USD 挂到发票行上落库，复核台才能并排展示、标红。发票行多于
    # 报关单行时多出的行保持 None（没得核对）。只搬值不改值，和 _check_customs_lines
    # 是同一份 customs_items、同一套按位置对应。
    for item, customs_item in zip(items, customs.get("customs_items") or [], strict=False):
        item["customs_fob_usd"] = customs_item.get("fob_usd")

    calculated_total = sum(
        (item["fob_revenue_usd"] or Decimal("0") for item in items),
        start=Decimal("0"),
    ).quantize(MONEY)
    invoice_total = invoice.get("fob_amount_usd")
    customs_total = customs.get("customs_fob_usd_total")
    verification_failed = (
        (invoice_total is not None and calculated_total != invoice_total)
        or (customs_total is not None and calculated_total != customs_total)
    )
    line_warnings = _check_customs_lines(items, customs.get("customs_items") or [])
    return {
        "ci_no": invoice.get("ci_no"),
        "cdn": customs.get("cdn"),
        "ci_date": invoice.get("ci_date"),
        "invoice_date": invoice_date,
        "exchange_target_date": exchange_target_date,
        "exchange_rate_date": exchange_rate_date,
        "exchange_rate": exchange_rate,
        "revenue_period": invoice_date.strftime("%Y%m") if invoice_date else None,
        "currency": currency.upper(),
        "customer_name": invoice.get("customer_name"),
        "customer_address": invoice.get("customer_address"),
        "po_no": invoice.get("po_no"),
        "incoterms": incoterms,
        "payment_term": invoice.get("payment_term"),
        "items": items,
        "fob_revenue_usd_total": calculated_total,
        "fob_revenue_thb_total": sum(
            (item["fob_revenue_thb"] or Decimal("0") for item in items),
            start=Decimal("0"),
        ).quantize(MONEY),
        "is_dap": incoterms == "DAP",
        "fob_verification_failed": verification_failed,
        "submission_date_low_confidence": customs.get(
            "submission_date_low_confidence",
            False,
        ),
        "submission_date_confidence": customs.get("submission_date_confidence", ""),
        "submission_date_source": customs.get("submission_date_source", ""),
        # ── 报关单侧留痕（业务 2026-07-30 新增）─────────────────────────────
        # 这些值不参与计价——计价口径仍是 BOT 汇率表，不能变。它们是核对用的：
        # 海关按哪个汇率折的泰铢、谁报的关、报关单自己印的泰铢金额是多少。
        # 全部要落到导出核对表里，所以必须一路带到 service 层。
        "declaration_ref_no": customs.get("declaration_ref_no") or None,
        "customs_exchange_rate": customs.get("customs_exchange_rate"),
        "forwarder_name": customs.get("forwarder_name") or None,
        "forwarder_name_th": customs.get("forwarder_name_th") or None,
        "forwarder_name_en": customs.get("forwarder_name_en") or None,
        "forwarder_tax_no": customs.get("forwarder_tax_no") or None,
        "customs_fob_usd_total": customs.get("customs_fob_usd_total"),
        "customs_fob_thb_line_total": customs.get("customs_fob_thb_line_total"),
        "customs_fob_thb_printed_total": customs.get("customs_fob_thb_printed_total"),
        # 非落库字段：service 摘掉它，只用来决定这张票要不要进 needs_review。
        "customs_warnings": [
            *(customs.get("customs_warnings") or []),
            *line_warnings,
        ],
    }


def _check_customs_lines(
    items: list[dict[str, Any]],
    customs_items: list[Mapping[str, Any]],
) -> list[str]:
    """逐行核对：报关单第 N 行 ↔ 商业发票第 N 行。

    业务口径（2026-07-30）：一份关单可能有两三行，**每行对应 commercial invoice
    的一行**，顺序一致。所以行数不等、或者某一行的 FOB USD 对不上，都是要人看的：
    合计对得上但分行错位（比如两行金额互换）在总额上完全看不出来。

    只报差异不改值——识别阶段擅自"对齐"就等于把这道核对做没了。
    关单读不到明细（老模板、文本层坏）时不报：那是"没得核对"，不是"核对不过"。
    """
    if not customs_items:
        return []

    warnings: list[str] = []
    if len(customs_items) != len(items):
        warnings.append(
            f"报关单有 {len(customs_items)} 个申报行，发票有 {len(items)} 行商品，"
            "行数不一致，请人工核对逐行对应关系"
        )

    for customs_item, item in zip(customs_items, items, strict=False):
        declared = customs_item.get("fob_usd")
        calculated = item.get("fob_revenue_usd")
        if declared is None or calculated is None:
            continue
        if declared != calculated:
            warnings.append(
                f"第 {item.get('line_number')} 行 FOB USD 不一致："
                f"系统算出 {calculated}，报关单该行为 {declared}"
            )
    return warnings
