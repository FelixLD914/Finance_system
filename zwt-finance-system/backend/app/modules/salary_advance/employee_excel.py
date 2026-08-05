from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.modules.salary_advance.service import SalaryAdvanceStateError

EMPLOYEE_TEMPLATE_SHEET = "员工主数据 Employee Master Data"

EMPLOYEE_HEADERS: dict[str, tuple[str, ...]] = {
    "emp_id": ("工号", "Emp.ID", "Employee ID", "emp_id"),
    "first_name": ("名 First Name", "名", "First Name", "first_name"),
    "surname": ("姓 Surname", "姓", "Surname", "surname"),
    "en_name": ("英文名", "EN Name", "English Name", "英文名字", "en_name"),
    "chinese_name": ("中文名", "Chinese Name", "chinese_name"),
    "department": ("部门", "Dept", "Department", "部门 Dept", "department"),
    "position": ("职位", "Position", "职位 Position", "position"),
    "start_date": ("入职日期", "Start Date", "开始工作日期", "start_date"),
    "is_active": ("是否在职", "Is Active", "Active", "is_active"),
}

EMPLOYEE_FIELD_ORDER: tuple[str, ...] = tuple(EMPLOYEE_HEADERS)

EMPLOYEE_EXAMPLE: dict[str, str] = {
    "emp_id": "EMP001",
    "first_name": "Somchai",
    "surname": "Saelim",
    "en_name": "Somchai Saelim",
    "chinese_name": "宋柴",
    "department": "IT",
    "position": "Software Engineer",
    "start_date": "2024-01-15",
    "is_active": "是",
}

EMPLOYEE_NOTES: tuple[tuple[str, str], ...] = (
    ("emp_id", "必填。员工工号（系统全局唯一）。"),
    ("first_name", "导入可空，但开预支单必填。英文名 / 名字。"),
    ("surname", "导入可空，但开预支单必填。英文姓氏。"),
    ("en_name", "选填。留空时自动取「名 + 姓」拼接。"),
    ("chinese_name", "选填。中文名。"),
    ("department", "导入可空，但开预支单必填。所属部门。"),
    ("position", "导入可空，但开预支单必填。职位/岗位名称。"),
    ("start_date", "导入可空，但开预支单必填。格式 YYYY-MM-DD，认不出的写法会整单退回。"),
    ("is_active", "选填。是否在职（'是'/'否'，留空默认'是'；填'否'会同步为离职）。"),
)


def _header_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def map_employee_headers(values: list[Any] | tuple[Any, ...]) -> dict[str, int]:
    alias_index = {
        _header_key(alias): field
        for field, aliases in EMPLOYEE_HEADERS.items()
        for alias in aliases
    }
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        field = alias_index.get(_header_key(value))
        if field is not None and field not in mapping:
            mapping[field] = index
    return mapping


def build_employee_import_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EMPLOYEE_TEMPLATE_SHEET
    header = [EMPLOYEE_HEADERS[field][0] for field in EMPLOYEE_FIELD_ORDER]
    sheet.append(header)
    sheet.append([EMPLOYEE_EXAMPLE.get(field) for field in EMPLOYEE_FIELD_ORDER])

    for index, name in enumerate(header, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = max(14, len(name) + 4)

    notes = workbook.create_sheet("说明 Notes")
    notes.append(["列名 Column", "说明 Note"])
    for field, note in EMPLOYEE_NOTES:
        notes.append([EMPLOYEE_HEADERS[field][0], note])
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 60

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y.%m.%d")
TRUTHY = ("true", "1", "是", "y", "yes", "active", "在职")
FALSY = ("false", "0", "否", "n", "no", "inactive", "离职")


# 认不出来的值必须报错，不能静默落成 None/False。
# 静默的代价出现在别处：入职日期悄悄变空，导入照样报"成功"，
# 等到开预支单时才以"入职日期缺失或无法解析"炸出来——那时人已经不在导入这一步了。
def _parse_date(value: Any, row_idx: int) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    val_str = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    raise SalaryAdvanceStateError(
        f"第 {row_idx} 行入职日期 '{val_str}' 无法识别，"
        f"支持的格式：{'、'.join(DATE_FORMATS)}"
    )


def _parse_bool(value: Any, row_idx: int) -> bool:
    if value is None or str(value).strip() == "":
        return True
    if isinstance(value, bool):
        return value
    val_str = str(value).strip().lower()
    if val_str in TRUTHY:
        return True
    if val_str in FALSY:
        return False
    raise SalaryAdvanceStateError(
        f"第 {row_idx} 行是否在职 '{value}' 无法识别，请填 '是' 或 '否'"
    )


def parse_employee_sheet(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise SalaryAdvanceStateError(f"Excel 文件无法解析: {exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SalaryAdvanceStateError("Excel 工作表为空")

    mapping = map_employee_headers(rows[0])
    if "emp_id" not in mapping:
        raise SalaryAdvanceStateError("未找到必需的 '工号' (emp_id) 列")

    parsed_rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        raw_emp_id = row[mapping["emp_id"]] if mapping["emp_id"] < len(row) else None
        emp_id = str(raw_emp_id).strip() if raw_emp_id is not None else ""
        if not emp_id:
            raise SalaryAdvanceStateError(f"第 {row_idx} 行缺失必需的工号 (emp_id)")

        def _get_val(field: str, r: tuple[Any, ...]) -> Any:
            idx = mapping.get(field)
            if idx is not None and idx < len(r):
                val = r[idx]
                return str(val).strip() if val is not None else None
            return None

        first_name = _get_val("first_name", row)
        surname = _get_val("surname", row)
        en_name = _get_val("en_name", row)
        chinese_name = _get_val("chinese_name", row)
        department = _get_val("department", row)
        position = _get_val("position", row)

        raw_start_date = (
            row[mapping["start_date"]]
            if "start_date" in mapping and mapping["start_date"] < len(row)
            else None
        )
        start_date = _parse_date(raw_start_date, row_idx)

        raw_is_active = (
            row[mapping["is_active"]]
            if "is_active" in mapping and mapping["is_active"] < len(row)
            else None
        )
        is_active = _parse_bool(raw_is_active, row_idx)

        parsed_rows.append(
            {
                "emp_id": emp_id,
                "first_name": first_name or None,
                "surname": surname or None,
                "en_name": en_name or None,
                "chinese_name": chinese_name or None,
                "department": department or None,
                "position": position or None,
                "start_date": start_date,
                "is_active": is_active,
            }
        )

    return parsed_rows
