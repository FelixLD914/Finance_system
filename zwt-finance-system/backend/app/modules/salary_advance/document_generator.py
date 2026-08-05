from __future__ import annotations

import hashlib
import re
import tempfile
from collections.abc import Iterator
from contextlib import contextmanager, suppress
from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image
from pypdf import PdfReader, PdfWriter

from app.core.signature_image import SignatureImageError
from app.core.signature_image import build_blue_signature as prepare_signature_stamp
from app.modules.salary_advance import pdf_layout
from app.modules.salary_advance.validation import safe_excel_text

CURRENT_RECORD_FIELDS = (
    "period",
    "emp_id",
    "en_name",
    "first_name",
    "surname",
    "chinese_name",
    "applicant_display_name",
    "department",
    "position",
    "start_date",
    "reason",
    "advance_amount",
    "advance_amount_words_th",
    "monthly_deduction",
    "monthly_deduction_words_th",
    "request_date",
    "finance_comment",
    "finance_display_name",
    "finance_date",
    "approval_status",
    "md_display_name",
    "md_date",
    "applicant_signature_mode",
    "finance_signature_code",
    "md_signature_code",
    "output_filename",
)

# 职位名称经常一行放不下（例：MANUFACTURING TECHNOLOGY SUPERVISOR 2）。
# 表单模板里 K8:L9 已按三行高度制版并改成底部对齐，PDF 叠加层必须用同样的行数
# 折行，两种产物才长得一样；改这里的行数要同步改模板第 9 行的行高。
WRAPPED_FIELDS = {"position": 3}
TOP_ALIGNED_FIELDS = {"position"}

LINE_SPACING = 1.2
MIN_FONT_SIZE = 7.0

FORMULA_CELLS = (
    "C8",
    "G8",
    "K8",
    "D10",
    "F12",
    "A15",
    "F17",
    "I17",
    "E22",
    "H22",
    "E24",
    "H24",
    "H29",
    "J30",
    "E32",
    "H37",
    "J38",
    "B42",
    "H42",
    "H47",
    "J48",
)


class SalaryAdvanceDocumentError(RuntimeError):
    pass


@dataclass(frozen=True)
class GenerationSnapshot:
    normalized_data: dict[str, Any]
    # None = 这一位不盖章。沿用 WHT / TAX INV 的 signature_source_path 约定。
    # 产线路径不会出现 None——SalaryAdvanceDocumentService._resolve_signature
    # 取不到签名是直接报错的（"绝不静默换人"，见 test_salary_advance_signature_
    # resolution）；用得上 None 的只有制签名预览底图的脚本，它要的正是一份未盖章的页。
    finance_signature_path: Path | None
    md_signature_path: Path | None
    finance_signature_version: dict[str, Any]
    md_signature_version: dict[str, Any]
    # 签名库维护页里那个"确认应用签名尺寸到系统开票"存下来的比例
    # (signature_assets.scale_percent)。两个签名位可以是两张不同的签名资产，
    # 各带各的比例，所以分开传，不能合成一个。
    # 默认 100 = 原尺寸：老调用方不传时行为与加这个参数之前完全一致。
    finance_scale_percent: int = 100
    md_scale_percent: int = 100


def sha256_path(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _date(value: Any) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value)).strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def _amount(value: Any) -> str:
    if value in (None, ""):
        return ""
    number = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{number:,.2f}"


def pdf_field_values(normalized: dict[str, Any]) -> dict[str, str]:
    applicant = str(normalized.get("applicant_display_name") or "")
    finance = str(normalized.get("finance_display_name") or "")
    md_name = str(normalized.get("md_display_name") or "")
    return {
        "first_name": str(normalized.get("first_name") or ""),
        "surname": str(normalized.get("surname") or ""),
        "position": str(normalized.get("position") or ""),
        "department": str(normalized.get("department") or ""),
        "start_date": _date(normalized.get("start_date")),
        "reason": str(normalized.get("reason") or ""),
        "advance_amount": _amount(normalized.get("advance_amount")),
        "advance_amount_words_th": str(normalized.get("advance_amount_words_th") or ""),
        "advance_amount_repeat": _amount(normalized.get("advance_amount")),
        "advance_words_repeat": str(normalized.get("advance_amount_words_th") or ""),
        "monthly_deduction": _amount(normalized.get("monthly_deduction")),
        "monthly_deduction_words_th": str(
            normalized.get("monthly_deduction_words_th") or ""
        ),
        "applicant_display_name": f"( {applicant} )" if applicant else "",
        "request_date": _date(normalized.get("request_date")),
        "finance_comment": str(normalized.get("finance_comment") or ""),
        "finance_display_name": f"( {finance} )" if finance else "",
        "finance_date": _date(normalized.get("finance_date")),
        "md_display_name": f"( {md_name} )" if md_name else "",
        "md_date": _date(normalized.get("md_date")),
    }


def _validate_signature(path: Path) -> None:
    if not path.is_file():
        raise SalaryAdvanceDocumentError(f"签名图片不存在：{path.name}")
    try:
        with Image.open(path) as image:
            image.verify()
    except Exception as exc:
        raise SalaryAdvanceDocumentError(f"签名图片无效：{path.name}") from exc


@contextmanager
def _stamped_signature(path: Path | None) -> Iterator[Path | None]:
    """把签名图按出票口径处理好（清底、去扫描下划线、转蓝墨、裁到墨迹外接框），
    用完即删。`path` 为 None 表示这一位不盖章，原样返回 None。

    与 WHT / TAX INV 走同一份 `core.signature_image`：同一张签名在三张单据上
    必须是同一支笔写的。**裁掉四周留白尤其重要**——drawImage 是等比适配整张图，
    留白会照比例吃掉签名框，不裁的话同一张图印在工资预支的 90pt 框里只有 34pt
    （实测），用户反馈"签名过小"就是这么来的。
    """
    if path is None:
        yield None
        return
    _validate_signature(path)
    with tempfile.TemporaryDirectory(prefix="zwt-salary-sig-") as workspace:
        prepared = Path(workspace) / "signature.png"
        try:
            prepare_signature_stamp(path, prepared)
        except SignatureImageError as exc:
            raise SalaryAdvanceDocumentError(
                f"签名图片无法用于套印：{path.name}（{exc}）"
            ) from exc
        yield prepared


def _add_workbook_signature(
    worksheet: Any,
    path: Path | None,
    anchor: str,
    scale_percent: int = 100,
) -> None:
    """xlsx 那一份也认同一个 scale_percent。

    xlsx 与 pdf 是同一张单据的两种产出，尺寸各走各的会让同一条记录导出两个样子。
    110x35 是签名框在 Excel 里的可用区域（约等于 pdf_layout 的 90x28 点），
    scale_percent 在这个"贴合尺寸"之上再乘一次，和 PDF 侧的含义一致。
    """
    if path is None:
        return
    _validate_signature(path)
    # 立刻读进内存，不要把路径交给 openpyxl：它到 workbook.save() 才真正读文件，
    # 而调用方传进来的是 _stamped_signature 那个用完即删的临时文件，
    # 拖到 save 的时候早就没了（save 在 with 块之外）。
    image = OpenpyxlImage(BytesIO(path.read_bytes()))
    ratio = (scale_percent or 100) / 100.0
    scale = min(110 / image.width, 35 / image.height) * ratio
    image.width = max(1, int(image.width * scale))
    image.height = max(1, int(image.height * scale))
    worksheet.add_image(image, anchor)


def render_salary_advance_workbook(
    template_path: Path,
    snapshot: GenerationSnapshot,
) -> bytes:
    if not template_path.is_file():
        raise SalaryAdvanceDocumentError(f"工资预支 Excel 模板不存在：{template_path}")
    workbook = load_workbook(template_path)
    try:
        if "当前记录" not in workbook.sheetnames or "表单模板" not in workbook.sheetnames:
            raise SalaryAdvanceDocumentError("Excel 模板缺少 当前记录 或 表单模板 工作表")
        current = workbook["当前记录"]
        normalized = snapshot.normalized_data
        for row, field in enumerate(CURRENT_RECORD_FIELDS, start=2):
            value = normalized.get(field)
            if field in {"start_date", "request_date", "finance_date", "md_date"} and value:
                with suppress(ValueError):
                    value = datetime.fromisoformat(str(value)).date()
            elif field in {"advance_amount", "monthly_deduction"} and value not in (None, ""):
                value = Decimal(str(value))
            elif isinstance(value, str):
                value = safe_excel_text(value)
            current.cell(row=row, column=2, value=value)

        form = workbook["表单模板"]
        form.row_dimensions[17].height = 32
        form.row_dimensions[32].height = 32
        form["H34"] = None
        form["H44"] = None
        for row in range(26, 29):
            for column in range(8, 11):
                cell = form.cell(row=row, column=column)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        no_fill = PatternFill(fill_type=None)
        for row_start in (34, 44):
            for row in range(row_start, row_start + 3):
                for column in range(8, 11):
                    form.cell(row=row, column=column).fill = no_fill
        approval = normalized.get("approval_status")
        form["B42"] = "☑" if approval == "Approve" else "☐"
        form["H42"] = "☑" if approval == "Not approved" else "☐"
        # xlsx 与 pdf 是同一张单据的两种产出，签名必须经过同一道处理，
        # 否则同一条记录导出两个样子（一份裁过、一份带着留白）。
        with (
            _stamped_signature(snapshot.finance_signature_path) as finance_stamp,
            _stamped_signature(snapshot.md_signature_path) as md_stamp,
        ):
            _add_workbook_signature(
                form, finance_stamp, "H34", snapshot.finance_scale_percent
            )
            _add_workbook_signature(form, md_stamp, "H44", snapshot.md_scale_percent)

        form.page_setup.orientation = "portrait"
        form.page_setup.paperSize = form.PAPERSIZE_A4
        form.page_setup.fitToWidth = 1
        form.page_setup.fitToHeight = 1
        form.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        form.print_area = "A1:L50"

        if workbook.calculation is None:
            workbook.calculation = CalcProperties()
        workbook.calculation.fullCalcOnLoad = True
        for sheet_name in ("使用说明", "导入数据模板", "字段映射", "签名配置"):
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        workbook._sheets = [form, current]
        workbook.active = form
        current.sheet_state = "hidden"

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def render_clean_template_workbook(template_path: Path) -> bytes:
    """制版用：保留固定 Logo/标签/边框，清空所有动态字段和签名区域。"""

    workbook = load_workbook(template_path)
    try:
        form = workbook["表单模板"]
        for coordinate in FORMULA_CELLS:
            form[coordinate] = None
        for start in (26, 34, 44):
            for row in range(start, start + 3):
                for column in range(8, 11):
                    cell = form.cell(row=row, column=column)
                    if not isinstance(cell, MergedCell):
                        cell.value = None
        form.page_setup.orientation = "portrait"
        form.page_setup.paperSize = form.PAPERSIZE_A4
        form.page_setup.fitToWidth = 1
        form.page_setup.fitToHeight = 1
        form.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        form.print_area = "A1:L50"
        for sheet in workbook.worksheets:
            sheet.sheet_state = "hidden" if sheet.title != "表单模板" else "visible"
        workbook.active = form
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


# 底版是 Excel 从模板导出的，表格里的标签用的是 TH SarabunPSK；叠加层必须注册同一个字体，
# 否则同字号下填入的字要宽/粗一半，和表格明显不是一套。注册名带上字体名是有意为之——
# reportlab 按名字缓存，名字含糊时换了文件也不会重新加载。
FORM_FONT_NAME = "SalaryAdvanceThSarabunPSK"


def _font_for(text: str) -> str:
    if re.search(r"[\u4e00-\u9fff]", text):
        return "STSong-Light"
    return FORM_FONT_NAME


def _fit(text: str, font: str, size: float, width: float) -> tuple[float, str]:
    from reportlab.pdfbase.pdfmetrics import stringWidth

    if not text:
        return size, text
    fitted_size = size
    measured = stringWidth(text, font, fitted_size)
    if measured > width:
        fitted_size = max(MIN_FONT_SIZE, fitted_size * width / measured)
    while text and stringWidth(text, font, fitted_size) > width:
        text = text[:-1]
    return fitted_size, text


def _wrap_by_words(text: str, font: str, size: float, width: float) -> tuple[list[str], bool]:
    """按空格/单词折行，绝不硬截断单词。
    若某个单词在当前字号下单词宽度超过 width，返回 has_oversized_word = True。
    """
    from reportlab.pdfbase.pdfmetrics import stringWidth

    words = text.split()
    if not words:
        return [], False

    lines: list[str] = []
    has_oversized_word = False
    current_line = ""

    for word in words:
        word_w = stringWidth(word, font, size)
        if word_w > width:
            has_oversized_word = True
        if not current_line:
            current_line = word
        elif stringWidth(f"{current_line} {word}", font, size) <= width:
            current_line = f"{current_line} {word}"
        else:
            lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)

    return lines, has_oversized_word


def _fit_lines(
    text: str,
    font: str,
    size: float,
    width: float,
    max_lines: int,
) -> tuple[float, list[str]]:
    if not text:
        return size, []
    if max_lines <= 1:
        fitted_size, fitted = _fit(text, font, size, width)
        return fitted_size, [fitted]

    fitted_size = size
    while fitted_size > MIN_FONT_SIZE:
        lines, has_oversized_word = _wrap_by_words(text, font, fitted_size, width)
        if len(lines) <= max_lines and not has_oversized_word:
            return fitted_size, lines
        fitted_size = max(MIN_FONT_SIZE, round(fitted_size - 0.2, 2))

    # 达到最小字号 MIN_FONT_SIZE 时，依然按单词折行，绝不截断单词
    lines, _ = _wrap_by_words(text, font, fitted_size, width)
    return fitted_size, lines


def _draw_signature(
    canvas: Any,
    path: Path | None,
    box: tuple[float, float, float, float],
    scale_percent: int = 100,
) -> None:
    """在签名框内套印签名，按 scale_percent **围绕框中心**缩放。

    缩放口径与 wht / tax_invoice 两处 drawImage 逐字一致：签名库维护页的预览
    就是照这个算式画的，三处但凡有一处不同，用户照着预览调出来的比例在那张
    单据上就是错的。加这个参数之前，工资预支单是这三种单据里唯一不认
    scale_percent 的——维护页照样弹"已成功应用至系统开票"，出票尺寸纹丝不动。
    """
    from reportlab.lib.utils import ImageReader

    if path is None:
        return
    _validate_signature(path)
    x, y, width, height = box
    ratio = (scale_percent or 100) / 100.0
    scaled_w = width * ratio
    scaled_h = height * ratio
    canvas.drawImage(
        ImageReader(str(path)),
        x + (width - scaled_w) / 2.0,
        y + (height - scaled_h) / 2.0,
        width=scaled_w,
        height=scaled_h,
        preserveAspectRatio=True,
        anchor="c",
        mask="auto",
    )


def _validate_layout(template_path: Path, underlay_path: Path) -> None:
    if pdf_layout.LAYOUT_VERSION == "unbuilt":
        raise SalaryAdvanceDocumentError(
            "工资预支 PDF 坐标尚未制版，请先运行 build_salary_advance_underlay.py"
        )
    xlsx_hash = sha256_path(template_path)
    pdf_hash = sha256_path(underlay_path)
    if xlsx_hash != pdf_layout.SOURCE_XLSX_SHA256:
        raise SalaryAdvanceDocumentError("工资预支 XLSX 模板与 PDF 坐标版本不匹配")
    if pdf_hash != pdf_layout.PDF_UNDERLAY_SHA256:
        raise SalaryAdvanceDocumentError("工资预支 PDF 底版与坐标版本不匹配")


def export_pdf_from_template(
    underlay_path: Path,
    output_path: Path,
    snapshot: GenerationSnapshot,
    form_font_path: Path,
    *,
    xlsx_template_path: Path | None = None,
) -> None:
    if not underlay_path.is_file():
        raise SalaryAdvanceDocumentError(f"工资预支 PDF 底版不存在：{underlay_path}")
    if not form_font_path.is_file():
        raise SalaryAdvanceDocumentError(f"表单字体不存在：{form_font_path}")
    if xlsx_template_path is not None:
        _validate_layout(xlsx_template_path, underlay_path)

    from reportlab.lib.colors import black
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    if FORM_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(FORM_FONT_NAME, str(form_font_path)))
    if "STSong-Light" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    reader = PdfReader(str(underlay_path))
    if len(reader.pages) != pdf_layout.PAGE_COUNT:
        raise SalaryAdvanceDocumentError(
            f"工资预支 PDF 底版必须为 {pdf_layout.PAGE_COUNT} 页"
        )
    page = reader.pages[0]
    buffer = BytesIO()
    overlay = canvas.Canvas(
        buffer,
        pagesize=(float(page.mediabox.width), float(page.mediabox.height)),
    )
    overlay.setFillColor(black)
    for field, value in pdf_field_values(snapshot.normalized_data).items():
        anchor = pdf_layout.TEXT_ANCHORS.get(field)
        if anchor is None or not value:
            continue
        font = _font_for(value)
        size, lines = _fit_lines(
            value,
            font,
            anchor.size,
            anchor.max_width,
            WRAPPED_FIELDS.get(field, 1),
        )
        overlay.setFont(font, size)
        if field in TOP_ALIGNED_FIELDS:
            # 顶部基线对齐：第 0 行锚在 first_name.y（与左侧 Name / Surname 顶行水平线 100% 精准齐平），后续行向下排列
            top_y = pdf_layout.TEXT_ANCHORS.get("first_name", anchor).y
            for offset, line in enumerate(lines):
                line_y = top_y - offset * size * LINE_SPACING
                if anchor.align == "right":
                    overlay.drawRightString(anchor.x, line_y, line)
                elif anchor.align == "center":
                    overlay.drawCentredString(anchor.x, line_y, line)
                else:
                    overlay.drawString(anchor.x, line_y, line)
        else:
            # 底部基线对齐：第 0 行在最下侧
            for offset, line in enumerate(reversed(lines)):
                line_y = anchor.y + offset * size * LINE_SPACING
                if anchor.align == "right":
                    overlay.drawRightString(anchor.x, line_y, line)
                elif anchor.align == "center":
                    overlay.drawCentredString(anchor.x, line_y, line)
                else:
                    overlay.drawString(anchor.x, line_y, line)

    approval = snapshot.normalized_data.get("approval_status")
    for name, selected in (
        ("approve", approval == "Approve"),
        ("not_approved", approval == "Not approved"),
    ):
        x, y, size = pdf_layout.CHECKBOX_ANCHORS[name]
        overlay.setLineWidth(0.8)
        overlay.rect(x, y, size, size, stroke=1, fill=0)
        if selected:
            overlay.setLineWidth(1.3)
            overlay.line(x + 1.5, y + size * 0.48, x + size * 0.42, y + 1.5)
            overlay.line(x + size * 0.42, y + 1.5, x + size - 1.2, y + size - 1.2)

    with (
        _stamped_signature(snapshot.finance_signature_path) as finance_stamp,
        _stamped_signature(snapshot.md_signature_path) as md_stamp,
    ):
        _draw_signature(
            overlay,
            finance_stamp,
            pdf_layout.SIGNATURE_BOXES["finance"],
            snapshot.finance_scale_percent,
        )
        _draw_signature(
            overlay,
            md_stamp,
            pdf_layout.SIGNATURE_BOXES["md"],
            snapshot.md_scale_percent,
        )
    overlay.save()
    buffer.seek(0)

    writer = PdfWriter()
    writer.add_page(page)
    writer.pages[0].merge_page(PdfReader(buffer).pages[0])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as target:
        writer.write(target)

    output_reader = PdfReader(str(output_path))
    if len(output_reader.pages) != 1 or output_path.stat().st_size < 1_000:
        raise SalaryAdvanceDocumentError("生成的工资预支 PDF 未通过一页、非空校验")


def document_file_stem(sequence: int, normalized: dict[str, Any]) -> str:
    period = normalized.get("period") or ""
    emp_id = normalized.get("emp_id") or ""
    name = normalized.get("chinese_name") or normalized.get("en_name") or ""
    raw = f"{sequence}.{period}+Salary Advance Form+{emp_id}+{name}"
    return re.sub(r'[\\/:*?"<>|]', "", raw)[:220]
