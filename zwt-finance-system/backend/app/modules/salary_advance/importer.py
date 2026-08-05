from __future__ import annotations

import hashlib
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.modules.salary_advance.validation import (
    FIELD_ORDER,
    HEADER_ALIASES,
    data_fingerprint,
    map_headers,
    validate_and_normalize_record,
)

MAX_ARCHIVE_FILES = 2_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024

# 生成的导入模板与解析时优先认的工作表同名，用一个常量绑死，避免两边各写一遍走偏。
IMPORT_TEMPLATE_SHEET = "导入数据模板"


class SalaryAdvanceImportError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedRecord:
    source_row_no: int
    period: str
    emp_id: str
    raw_data: dict[str, Any]
    normalized_data: dict[str, Any]
    data_fingerprint: str
    validation_status: str
    validation_errors: list[dict[str, str]]
    validation_warnings: list[dict[str, str]]


@dataclass(frozen=True)
class ParsedBatch:
    source_sha256: str
    records: list[ParsedRecord]

    @property
    def valid_rows(self) -> int:
        return sum(record.validation_status == "valid" for record in self.records)

    @property
    def warning_rows(self) -> int:
        return sum(record.validation_status == "warning" for record in self.records)

    @property
    def invalid_rows(self) -> int:
        return sum(record.validation_status == "invalid" for record in self.records)


def _json_value(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _validate_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            files = archive.infolist()
            if len(files) > MAX_ARCHIVE_FILES:
                raise SalaryAdvanceImportError("Excel 压缩包包含过多文件")
            total = 0
            for item in files:
                path = item.filename.replace("\\", "/")
                if path.startswith("/") or ".." in path.split("/"):
                    raise SalaryAdvanceImportError("Excel 压缩包包含越界路径")
                total += item.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise SalaryAdvanceImportError("Excel 解压后内容超过安全限制")
                if path.lower().endswith("vbaproject.bin"):
                    raise SalaryAdvanceImportError("不接受包含宏的 Excel 文件")
    except zipfile.BadZipFile as exc:
        raise SalaryAdvanceImportError("上传文件不是有效的 .xlsx 工作簿") from exc


def parse_salary_advance_workbook(
    content: bytes,
    *,
    period: str,
    existing_keys: set[tuple[str, str]] | None = None,
) -> ParsedBatch:
    _validate_archive(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise SalaryAdvanceImportError(f"无法解析 Excel：{exc}") from exc

    try:
        if IMPORT_TEMPLATE_SHEET in workbook.sheetnames:
            worksheet = workbook[IMPORT_TEMPLATE_SHEET]
        elif workbook.sheetnames:
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            raise SalaryAdvanceImportError("工作簿中没有工作表")

        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise SalaryAdvanceImportError("工作表为空")
        mapping = map_headers(header)
        missing = [field for field in ("period", "emp_id") if field not in mapping]
        if missing:
            raise SalaryAdvanceImportError(f"缺少必要列：{', '.join(missing)}")

        parsed: list[ParsedRecord] = []
        seen: dict[tuple[str, str], int] = {}
        external = existing_keys or set()
        for row_number, row in enumerate(rows, start=2):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            raw = {
                field: row[column]
                for field, column in mapping.items()
                if column < len(row)
            }
            row_period = raw.get("period")
            if row_period is not None:
                try:
                    row_period_text = str(int(float(row_period)))
                except (TypeError, ValueError):
                    row_period_text = str(row_period).strip()
                if row_period_text and row_period_text != period:
                    continue

            status, errors, warnings, normalized = validate_and_normalize_record(
                raw,
                batch_period=period,
            )
            key = (normalized.get("period", ""), normalized.get("emp_id", ""))
            if all(key):
                if key in seen:
                    status = "invalid"
                    errors.append(
                        {
                            "field": "emp_id",
                            "code": "DUPLICATE_PERIOD_EMPLOYEE",
                            "message": f"同一批次期间+工号重复，首次出现在第 {seen[key]} 行",
                        }
                    )
                elif key in external:
                    status = "invalid"
                    errors.append(
                        {
                            "field": "emp_id",
                            "code": "DUPLICATE_ACTIVE_RECORD",
                            "message": "系统中已存在相同期间和工号的有效记录",
                        }
                    )
                seen[key] = row_number

            parsed.append(
                ParsedRecord(
                    source_row_no=row_number,
                    period=normalized.get("period") or period,
                    emp_id=normalized.get("emp_id") or "",
                    raw_data=_json_value(raw),
                    normalized_data=_json_value(normalized),
                    data_fingerprint=data_fingerprint(normalized),
                    validation_status=status,
                    validation_errors=errors,
                    validation_warnings=warnings,
                )
            )
    finally:
        workbook.close()

    if not parsed:
        raise SalaryAdvanceImportError("所选期间没有可导入的数据行")
    return ParsedBatch(
        source_sha256=hashlib.sha256(content).hexdigest(),
        records=parsed,
    )


# 示例行按字段给值，导出时按 FIELD_ORDER 排列；长度由测试钉住必须与列数一致。
# 只填必填与常用列，可留空的（泰文大写、中文名、输出文件名等）留空以示意。
_TEMPLATE_EXAMPLE: dict[str, Any] = {
    "row_no": 1,
    "period": "202607",
    "emp_id": "E001",
    "en_name": "SOMCHAI TEST",
    "department": "Finance",
    "position": "Accountant",
    "advance_amount": 12500.5,
    "request_date": "2026-07-20",
    "start_date": "2024-01-15",
    "first_name": "SOMCHAI",
    "surname": "TEST",
    "finance_display_name": "邢兰慧",
    "md_display_name": "龚尧文",
    "reason": "Living expenses",
    "monthly_deduction": 2500.1,
    "finance_comment": "Verified",
    "approval_status": "Approve",
    "applicant_signature_mode": "Handwritten",
    "finance_signature_code": "FIN_XING_LANHUI",
    "md_signature_code": "MD_GONG_YAOWEN",
    "finance_date": "2026-07-20",
    "md_date": "2026-07-20",
}

# 逐列说明，键是字段名（渲染时换成规范列头）。文字落在 Excel 单元格里不渲染
# markdown，别写 ** 强调，会显示成星号。
_TEMPLATE_NOTES: tuple[tuple[str, str], ...] = (
    ("period", "必填。期间 YYYYMM，例如 202607；与导入时所选期间不一致的行会被跳过。"),
    ("emp_id", "必填。员工工号；同一批次内「期间 + 工号」不能重复。"),
    ("first_name", "必填。名。"),
    ("surname", "必填。姓。"),
    ("en_name", "可留空。留空时自动用「名 姓」大写拼接。"),
    ("department", "必填。部门。"),
    ("position", "必填。职位；超过 50 字会提示表单里可能缩小显示。"),
    ("advance_amount", "必填。预支金额，必须大于 0，不要带千分位符号。"),
    ("monthly_deduction", "可留空。留空默认等于预支金额；必须大于 0 且不超过预支金额。"),
    ("request_date", "必填。申请/签字日期，格式 YYYY-MM-DD。"),
    ("start_date", "必填。入职日期，YYYY-MM-DD，不能晚于申请日期。"),
    ("approval_status", "可留空。Approve / Not approved / Pending，留空按 Pending。"),
    (
        "finance_display_name",
        "可留空。单据上印的签字人姓名取自签名库里那张签名的「签名人姓名」，不取本列。",
    ),
    (
        "md_display_name",
        "可留空。单据上印的签字人姓名取自签名库里那张签名的「签名人姓名」，不取本列。",
    ),
    (
        "finance_signature_code",
        "可留空。盖哪张章在开具时选定；这里填了签名库中的代码就按它盖。",
    ),
    (
        "md_signature_code",
        "可留空。盖哪张章在开具时选定；这里填了签名库中的代码就按它盖。",
    ),
    ("advance_amount_words_th", "可留空。系统按金额自动换算泰文大写；填了不一致会提示。"),
    ("applicant_signature_mode", "只能填 Handwritten（申请人手写签名）。"),
    ("output_filename", "可留空。默认按「期间_工号_姓名_工资预支单」命名。"),
)


def _template_header() -> list[str]:
    # 用每个字段的第一个别名做规范列头——map_headers 能原样映射回来，模板与解析不脱节。
    return [HEADER_ALIASES[field][0] for field in FIELD_ORDER]


def build_import_template_workbook() -> bytes:
    """生成工资预支导入模板：第一张表列头 + 一行示例，第二张表逐列填写说明。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = IMPORT_TEMPLATE_SHEET
    header = _template_header()
    sheet.append(header)
    sheet.append([_TEMPLATE_EXAMPLE.get(field) for field in FIELD_ORDER])
    for index, name in enumerate(header, start=1):
        sheet.column_dimensions[
            sheet.cell(row=1, column=index).column_letter
        ].width = max(14, len(name) + 4)

    notes = workbook.create_sheet("说明 Notes")
    notes.append(["列名 Column", "说明 Note"])
    for field, note in _TEMPLATE_NOTES:
        notes.append([HEADER_ALIASES[field][0], note])
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 82

    buffer = BytesIO()
    try:
        workbook.save(buffer)
    finally:
        workbook.close()
    return buffer.getvalue()

