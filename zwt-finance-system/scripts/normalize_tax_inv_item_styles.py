"""把 TAX INV 模板三联明细区的数字格式与水平对齐统一成正本的样式。

模板一页横排三联，两联 COPY 用 "=B23" 这类公式镜像第一联。公式只搬值不搬
样式，而模板给两联 COPY 留的是另一套样式，于是同一张单据的三联印出来不一样：

    数量 1101.25   正本 1,101.25   副联 1,101      小数被抹掉
    数量 4820.5    正本 4,820.5    副联 4,821      甚至进位
    FOB Rev USD    正本两位小数    副联 31~40 行是三位小数

税务副联和正本是同一份单据，这属于模板缺陷。

对齐必须和格式一起改：副联原本的会计格式（"_ * #,##0.00_ "）靠 "* " 重复
填充把数字顶到右边，等于强行压住了单元格自己声明的水平对齐。一旦只把格式
换成普通的 "#,##0.00"，被压住的对齐就会冒出来——模板里 COPY#1 的汇率列声明
的是 left，会当场变成左对齐。所以这两项要么都改，要么都不改。

    cd zwt-finance-system\\backend
    .\\.venv\\Scripts\\python.exe ..\\scripts\\normalize_tax_inv_item_styles.py          # 预演
    .\\.venv\\Scripts\\python.exe ..\\scripts\\normalize_tax_inv_item_styles.py --apply  # 写入

用 Excel COM 而不是 openpyxl 改写：openpyxl 存盘会把三联共用的那张 PNG 拆成
三份（199KB -> 331KB），Excel 原样保留。两者都会丢掉 xl/cellimages.xml，
但那是 WPS 留下的空壳（无 DISPIMG 引用），丢了不影响版面。

列宽三联本来就各不相同（如 THB 金额列 14.36 / 15.64 / 17.18），那是模板一直
以来的排版，不在本脚本的处理范围内——改列宽会动到 print_area 的分页。

模板改动后底版必须重新制版：sha256 变了，scripts/build_tax_inv_underlay.py
记在 pdf_layout.py 里的源模板校验和要跟着更新。
"""

from __future__ import annotations

import argparse
import hashlib
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parents[1] / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import load_workbook  # noqa: E402

from app.modules.tax_invoice.document_generator import (  # noqa: E402
    COPY_COLUMN_OFFSETS,
    DATA_COLUMNS,
    DATA_START_ROW,
    ITEM_NUMBER_FORMATS,
    MAX_ITEMS,
)

DEFAULT_TEMPLATE = Path(r"D:\ZWTFinance\data\templates\TAX-INV-Template.xlsx")
ITEM_ROWS = range(DATA_START_ROW, DATA_START_ROW + MAX_ITEMS)

# Excel 的 xlHAlign 常量。openpyxl 用 None 表示"常规"：数字在常规下本来就
# 靠右，和 right 视觉一致，但仍然照抄正本的声明，免得又多出第二套事实。
XL_HALIGN = {None: 1, "general": 1, "left": -4131, "center": -4108, "right": -4152}


class NormalizeError(RuntimeError):
    pass


@dataclass(frozen=True)
class Change:
    coordinate: str
    kind: str  # "format" 或 "align"
    current: str
    wanted: str


def sha256_of(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def plan(template_path: Path) -> list[Change]:
    """列出所有要改的单元格。

    数字格式以 renderer 的 ITEM_NUMBER_FORMATS 为准，三联一律覆盖，顺带把
    第一联模板里那套不统一的格式也理平；水平对齐以第一联当前的声明为准，
    只同步两联 COPY——正本怎么排是模板的事，本脚本不替它做主。
    """
    workbook = load_workbook(template_path)
    worksheet = workbook["Invoice"]
    changes: list[Change] = []
    for row in ITEM_ROWS:
        for column, wanted_format in ITEM_NUMBER_FORMATS.items():
            for copy_offset in COPY_COLUMN_OFFSETS:
                cell = worksheet.cell(row=row, column=column + copy_offset)
                if cell.number_format != wanted_format:
                    changes.append(
                        Change(cell.coordinate, "format", cell.number_format, wanted_format)
                    )
        for column in DATA_COLUMNS:
            wanted_align = worksheet.cell(row=row, column=column).alignment.horizontal
            for copy_offset in COPY_COLUMN_OFFSETS[1:]:
                cell = worksheet.cell(row=row, column=column + copy_offset)
                if cell.alignment.horizontal != wanted_align:
                    changes.append(
                        Change(
                            cell.coordinate,
                            "align",
                            str(cell.alignment.horizontal),
                            str(wanted_align),
                        )
                    )
    workbook.close()
    return changes


def report(changes: list[Change]) -> None:
    if not changes:
        print("  三联明细区样式已经一致，无需改动。")
        return
    grouped: dict[tuple[str, str, str], list[str]] = {}
    for change in changes:
        key = (change.kind, change.current, change.wanted)
        grouped.setdefault(key, []).append(change.coordinate)
    for (kind, current, wanted), cells in sorted(grouped.items(), key=lambda kv: -len(kv[1])):
        cells.sort(key=lambda name: (len(name), name))
        shown = ", ".join(cells[:8]) + (f" … 共 {len(cells)} 格" if len(cells) > 8 else "")
        print(f"\n  [{kind}] {current!r}\n    -> {wanted!r}\n    {shown}")


def apply_with_excel(template_path: Path, changes: list[Change]) -> None:
    """用 Excel COM 逐格改样式。只动 NumberFormat / HorizontalAlignment。"""
    try:
        import win32com.client as win32
    except ImportError as exc:  # pragma: no cover - 只在开发机上跑
        raise NormalizeError("需要 pywin32，且必须在装了 Excel 的机器上运行") from exc

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(template_path))
        worksheet = workbook.Worksheets("Invoice")
        for change in changes:
            target = worksheet.Range(change.coordinate)
            if change.kind == "format":
                target.NumberFormat = change.wanted
            else:
                wanted = None if change.wanted == "None" else change.wanted
                target.HorizontalAlignment = XL_HALIGN[wanted]
        workbook.Save()
        workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()


def verify(template_path: Path) -> None:
    remaining = plan(template_path)
    if remaining:
        raise NormalizeError(
            f"改完之后仍有 {len(remaining)} 处不一致，"
            f"例如 {[change.coordinate for change in remaining[:5]]}"
        )
    print("  复核通过：三联明细区的数字格式与水平对齐逐格一致。")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="真正写入模板；不加就只预演，不碰文件",
    )
    args = parser.parse_args()

    template_path: Path = args.template
    if not template_path.is_file():
        print(f"找不到模板：{template_path}")
        return 1

    before = sha256_of(template_path)
    print(f"模板   : {template_path}")
    print(f"sha256 : {before}")
    print(f"大小   : {template_path.stat().st_size} 字节")

    changes = plan(template_path)
    formats = sum(1 for change in changes if change.kind == "format")
    print(f"\n需要修改 {len(changes)} 处（格式 {formats}，对齐 {len(changes) - formats}）：")
    report(changes)

    if not changes:
        return 0
    if not args.apply:
        print("\n以上仅为预演。确认无误后加 --apply 写入。")
        return 0

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = template_path.with_suffix(f".{stamp}.bak.xlsx")
    shutil.copy2(template_path, backup)
    print(f"\n已备份 -> {backup}")

    apply_with_excel(template_path, changes)
    after = sha256_of(template_path)
    print(f"已写入 : {template_path}")
    print(f"新大小 : {template_path.stat().st_size} 字节")
    verify(template_path)

    print("\n" + "=" * 72)
    print(f"源模板 sha256 变了：\n  {before}\n  -> {after}")
    print("底版必须重新制版，否则 pdf_layout.py 记的校验和对不上：")
    print("  .\\.venv\\Scripts\\python.exe ..\\scripts\\build_tax_inv_underlay.py")
    print("=" * 72)
    return 0


if __name__ == "__main__":
    sys.exit(main())
