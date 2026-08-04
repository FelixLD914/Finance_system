r"""一次性制版：从工资预支 Excel 模板生成静态 PDF 底版和坐标表。

只在装有 Excel 的开发/制版机运行；生产 API 不导入 pywin32，也不调用 Office。

    cd zwt-finance-system\backend
    .\.venv\Scripts\python.exe ..\scripts\build_salary_advance_underlay.py
"""

from __future__ import annotations

import argparse
import hashlib
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

BACKEND_ROOT = Path(__file__).resolve().parents[1] / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.salary_advance.document_generator import (  # noqa: E402
    render_clean_template_workbook,
)

DEFAULT_TEMPLATE = BACKEND_ROOT / "app" / "assets" / "templates" / "Salary-Advance-Template.xlsx"
DEFAULT_UNDERLAY = BACKEND_ROOT / "app" / "assets" / "templates" / "Salary-Advance-Template.pdf"
DEFAULT_LAYOUT = BACKEND_ROOT / "app" / "modules" / "salary_advance" / "pdf_layout.py"

FIELD_CELLS = {
    "first_name": "C8",
    "surname": "G8",
    "position": "K8",
    "department": "D10",
    "start_date": "F12",
    "reason": "A15",
    "advance_amount": "F17",
    "advance_amount_words_th": "I17",
    "advance_amount_repeat": "E22",
    "advance_words_repeat": "H22",
    "monthly_deduction": "E24",
    "monthly_deduction_words_th": "H24",
    "applicant_display_name": "H29",
    "request_date": "J30",
    "finance_comment": "E32",
    "finance_display_name": "H37",
    "finance_date": "J38",
    "md_display_name": "H47",
    "md_date": "J48",
}

MAX_WIDTHS = {
    "first_name": 105.0,
    "surname": 105.0,
    # K8:L9 实测可绘宽度 113.04pt，留出左右各 1.5pt 内边距；
    # 这个值同时决定 PDF 叠加层的折行位置，改列宽后要重新量。
    "position": 136.0,
    "department": 300.0,
    "start_date": 150.0,
    "reason": 500.0,
    "advance_amount": 120.0,
    "advance_amount_words_th": 230.0,
    "advance_amount_repeat": 120.0,
    "advance_words_repeat": 230.0,
    "monthly_deduction": 120.0,
    "monthly_deduction_words_th": 230.0,
    "applicant_display_name": 220.0,
    "request_date": 100.0,
    "finance_comment": 320.0,
    "finance_display_name": 160.0,
    "finance_date": 100.0,
    "md_display_name": 160.0,
    "md_date": 100.0,
}


class PlateError(RuntimeError):
    pass


@dataclass(frozen=True)
class WordBox:
    text: str
    x0: float
    x1: float
    baseline: float
    size: float


def export_pdf(xlsx_path: Path, pdf_path: Path) -> None:
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(xlsx_path), UpdateLinks=0, ReadOnly=True)
        try:
            worksheet = workbook.Worksheets("表单模板")
            worksheet.Activate()
            worksheet.PageSetup.PrintArea = "$A$1:$L$50"
            worksheet.PageSetup.Orientation = 1  # xlPortrait
            worksheet.PageSetup.PaperSize = 9  # xlPaperA4
            worksheet.PageSetup.Zoom = False
            worksheet.PageSetup.FitToPagesWide = 1
            worksheet.PageSetup.FitToPagesTall = 1
            worksheet.ExportAsFixedFormat(0, str(pdf_path))
        finally:
            workbook.Close(False)
    finally:
        excel.Quit()
    if not pdf_path.is_file():
        raise PlateError(f"Excel 未生成 {pdf_path}")


def _probe_workbook(template_path: Path, output_path: Path) -> dict[str, str]:
    workbook = load_workbook(template_path)
    try:
        form = workbook["表单模板"]
        probes: dict[str, str] = {}
        for index, (field, cell) in enumerate(FIELD_CELLS.items(), start=1):
            value = f"Z{index:02d}Q"
            probes[field] = value
            form[cell] = value
        form["B42"] = "ZAP"
        form["H42"] = "ZNO"
        form["H34"] = "ZFIN"
        form["H44"] = "ZMDX"
        for sheet in workbook.worksheets:
            sheet.sheet_state = "hidden" if sheet.title != "表单模板" else "visible"
        workbook.active = form
        form.print_area = "A1:L50"
        form.page_setup.orientation = "portrait"
        form.page_setup.paperSize = form.PAPERSIZE_A4
        form.page_setup.fitToWidth = 1
        form.page_setup.fitToHeight = 1
        workbook.save(output_path)
        return probes
    finally:
        workbook.close()


def _word_boxes(pdf_path: Path) -> tuple[tuple[float, float], list[WordBox]]:
    import pdfplumber

    with pdfplumber.open(pdf_path) as document:
        if len(document.pages) != 1:
            raise PlateError(f"工资预支底版必须为一页，当前为 {len(document.pages)} 页")
        page = document.pages[0]
        boxes: list[WordBox] = []
        for word in page.extract_words(keep_blank_chars=False):
            characters = [
                char
                for char in page.chars
                if char["x0"] >= word["x0"] - 0.1
                and char["x1"] <= word["x1"] + 0.1
                and char["bottom"] > word["top"] - 0.5
                and char["top"] < word["bottom"] + 0.5
            ]
            if not characters:
                continue
            baseline = round(
                sum(float(char["matrix"][5]) for char in characters) / len(characters),
                2,
            )
            boxes.append(
                WordBox(
                    text=word["text"],
                    x0=round(float(word["x0"]), 2),
                    x1=round(float(word["x1"]), 2),
                    baseline=baseline,
                    size=round(float(characters[0]["size"]), 2),
                )
            )
        return (round(float(page.width), 2), round(float(page.height), 2)), boxes


def _find(boxes: list[WordBox], text: str) -> WordBox:
    found = [box for box in boxes if box.text == text]
    if not found:
        found = [box for box in boxes if box.text.startswith(text)]
    if len(found) != 1:
        visible_probes = [box.text for box in boxes if box.text.startswith("Z")]
        raise PlateError(
            f"探针 {text!r} 应定位一次，实际 {len(found)} 次；"
            f"已识别探针：{visible_probes}"
        )
    return found[0]


def _layout_source(
    *,
    source_hash: str,
    underlay_hash: str,
    page_size: tuple[float, float],
    probes: dict[str, str],
    boxes: list[WordBox],
    template_path: Path,
) -> str:
    workbook = load_workbook(template_path, data_only=False)
    try:
        form = workbook["表单模板"]
        anchors = []
        for field, cell in FIELD_CELLS.items():
            box = _find(boxes, probes[field])
            alignment = form[cell].alignment.horizontal or "left"
            if alignment == "center":
                x = (box.x0 + box.x1) / 2
            elif alignment == "right":
                x = box.x1
            else:
                x = box.x0
                alignment = "left"
            anchors.append(
                f'    "{field}": TextAnchor({x:.2f}, {box.baseline:.2f}, '
                f'{box.size:.2f}, {MAX_WIDTHS[field]:.2f}, "{alignment}"),'
            )
    finally:
        workbook.close()

    approve = _find(boxes, "ZAP")
    not_approved = _find(boxes, "ZNO")
    finance = _find(boxes, "ZFIN")
    md = _find(boxes, "ZMDX")
    checkbox_size = 8.5
    checkbox_rows = [
        f'    "approve": ({approve.x0:.2f}, '
        f"{approve.baseline - 1.5:.2f}, {checkbox_size:.2f}),",
        f'    "not_approved": ({not_approved.x0:.2f}, '
        f"{not_approved.baseline - 1.5:.2f}, {checkbox_size:.2f}),",
    ]
    signature_rows = [
        f'    "finance": ({(finance.x0 + finance.x1) / 2 - 45:.2f}, '
        f"{finance.baseline - 12:.2f}, 90.00, 28.00),",
        f'    "md": ({(md.x0 + md.x1) / 2 - 45:.2f}, '
        f"{md.baseline - 12:.2f}, 90.00, 28.00),",
    ]
    version = hashlib.sha256(
        f"{source_hash}:{underlay_hash}:{anchors}".encode()
    ).hexdigest()[:16]
    return f'''"""自动生成；源文件：Salary-Advance-Template.xlsx。请勿手工改坐标。"""

from dataclasses import dataclass


@dataclass(frozen=True)
class TextAnchor:
    x: float
    y: float
    size: float
    max_width: float
    align: str = "left"


SOURCE_XLSX_SHA256 = "{source_hash}"
PDF_UNDERLAY_SHA256 = "{underlay_hash}"
LAYOUT_VERSION = "{version}"
PAGE_COUNT = 1
PAGE_SIZE = ({page_size[0]:.2f}, {page_size[1]:.2f})

TEXT_ANCHORS = {{
{chr(10).join(anchors)}
}}

CHECKBOX_ANCHORS = {{
{chr(10).join(checkbox_rows)}
}}

SIGNATURE_BOXES = {{
{chr(10).join(signature_rows)}
}}
'''


def build(template_path: Path, underlay_path: Path, layout_path: Path) -> None:
    if not template_path.is_file():
        raise PlateError(f"模板不存在：{template_path}")
    with tempfile.TemporaryDirectory(prefix="salary-advance-plate-") as temporary:
        root = Path(temporary)
        clean_xlsx = root / "clean.xlsx"
        clean_xlsx.write_bytes(render_clean_template_workbook(template_path))
        underlay_path.parent.mkdir(parents=True, exist_ok=True)
        export_pdf(clean_xlsx, underlay_path)

        probe_xlsx = root / "probe.xlsx"
        probe_pdf = root / "probe.pdf"
        probes = _probe_workbook(template_path, probe_xlsx)
        export_pdf(probe_xlsx, probe_pdf)
        page_size, boxes = _word_boxes(probe_pdf)

    source_hash = hashlib.sha256(template_path.read_bytes()).hexdigest()
    underlay_hash = hashlib.sha256(underlay_path.read_bytes()).hexdigest()
    layout_path.write_text(
        _layout_source(
            source_hash=source_hash,
            underlay_hash=underlay_hash,
            page_size=page_size,
            probes=probes,
            boxes=boxes,
            template_path=template_path,
        ),
        encoding="utf-8",
    )
    print(f"underlay={underlay_path}")
    print(f"layout={layout_path}")
    print(f"xlsx_sha256={source_hash}")
    print(f"pdf_sha256={underlay_hash}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--underlay", type=Path, default=DEFAULT_UNDERLAY)
    parser.add_argument("--layout", type=Path, default=DEFAULT_LAYOUT)
    arguments = parser.parse_args()
    build(arguments.template.resolve(), arguments.underlay.resolve(), arguments.layout.resolve())


if __name__ == "__main__":
    main()
